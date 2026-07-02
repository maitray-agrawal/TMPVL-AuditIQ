import unittest
import datetime
import io
import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from backend.app.core.db import Base
from backend.app.models.models import Trainee, InvoiceRecord, PaymentLedger, ValidationResult, AuditLog
from backend.app.services.report_service import ReportService

class TestReportService(unittest.TestCase):
    def setUp(self):
        # Configure in-memory database for testing
        self.engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(self.engine)
        Session = sessionmaker(bind=self.engine)
        self.db = Session()

    def tearDown(self):
        self.db.close()
        Base.metadata.drop_all(self.engine)

    def test_generate_approved_invoice_excel(self):
        # Populate dummy data
        t = Trainee(
            id="T001", name="Alice", doj=datetime.date(2026, 1, 1), dol=datetime.date(2026, 6, 1),
            scheme="NAPS", status="SEPARATED"
        )
        self.db.add(t)
        
        inv1 = InvoiceRecord(
            invoice_number="INV-APPROVED", invoice_date=datetime.date(2026, 2, 1), trainee_id="T001",
            billed_name="Alice", billed_joining_amount=1200.0, billed_180_days_amount=600.0,
            billed_other_amount=100.0, billed_total_amount=1900.0,
            approved_joining_amount=1200.0, approved_180_days_amount=0.0, approved_total_amount=1200.0,
            status="APPROVED", file_name="invoice.xlsx"
        )
        # Add another record with approved amounts = 0 (should be excluded)
        inv2 = InvoiceRecord(
            invoice_number="INV-APPROVED", invoice_date=datetime.date(2026, 2, 1), trainee_id="T001",
            billed_name="Alice", billed_joining_amount=1200.0, billed_180_days_amount=600.0,
            billed_other_amount=100.0, billed_total_amount=1900.0,
            approved_joining_amount=0.0, approved_180_days_amount=0.0, approved_total_amount=0.0,
            status="APPROVED", file_name="invoice.xlsx"
        )
        self.db.add_all([inv1, inv2])
        self.db.commit()

        # Generate report
        report_bytes = ReportService.generate_approved_invoice_excel(self.db, "INV-APPROVED")
        
        # Load and verify
        wb = openpyxl.load_workbook(io.BytesIO(report_bytes))
        self.assertIn("Approved Payouts", wb.sheetnames)
        
        ws = wb["Approved Payouts"]
        # Max row is 2 (1 header, 1 data row)
        self.assertEqual(ws.max_row, 2)
        
        # Verify headers
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers[0], "Trainee ID")
        self.assertEqual(headers[1], "Trainee Name")
        self.assertEqual(headers[2], "Scheme")
        
        # Verify row value
        row_vals = [cell.value for cell in ws[2]]
        self.assertEqual(row_vals[0], "T001")
        self.assertEqual(row_vals[1], "Alice")
        self.assertEqual(row_vals[2], "NAPS")
        self.assertEqual(row_vals[3], "2026-01-01")
        self.assertEqual(row_vals[4], "2026-06-01")
        self.assertEqual(row_vals[7], 1200.0) # approved joining
        self.assertEqual(row_vals[8], 0.0) # approved 180 days

        # Verify identical column width calculation (total_rows = 2)
        # Col A (col_idx=1): max_len = len("<Cell 'Approved Payouts'.A2>") = 28. width = max(28+3, 12) = 31
        self.assertEqual(ws.column_dimensions["A"].width, 31.0)
        self.assertEqual(ws.column_dimensions["B"].width, 31.0)

    def test_generate_exception_report_excel(self):
        # Populate exception warnings/errors
        t = Trainee(id="T001", name="Alice", doj=datetime.date(2026, 1, 1), scheme="NAPS")
        self.db.add(t)
        inv = InvoiceRecord(
            invoice_number="INV-EXC", invoice_date=datetime.date(2026, 2, 1), trainee_id="T001",
            billed_name="Alice", file_name="invoice.xlsx"
        )
        self.db.add(inv)
        self.db.commit()

        val = ValidationResult(
            invoice_record_id=inv.id, trainee_id="T001", rule_name="Test Exception Rule",
            status="ERROR", message="Validation error msg", created_at=datetime.datetime(2026, 2, 1, 10, 0, 0)
        )
        self.db.add(val)
        self.db.commit()

        report_bytes = ReportService.generate_exception_report_excel(self.db, "INV-EXC")

        wb = openpyxl.load_workbook(io.BytesIO(report_bytes))
        self.assertIn("Exceptions", wb.sheetnames)
        
        ws = wb["Exceptions"]
        self.assertEqual(ws.max_row, 2)
        
        row_vals = [cell.value for cell in ws[2]]
        self.assertEqual(row_vals[0], "INV-EXC")
        self.assertEqual(row_vals[1], "T001")
        self.assertEqual(row_vals[2], "Alice")
        self.assertEqual(row_vals[3], "Test Exception Rule")
        self.assertEqual(row_vals[4], "ERROR")
        self.assertEqual(row_vals[5], "N/A")
        self.assertEqual(row_vals[6], "Validation error msg")
        self.assertEqual(row_vals[8], "2026-02-01 10:00:00")

        # Verify identical column width calculation (total_rows = 2)
        # Col A (col_idx=1): max_len = len("<Cell 'Exceptions'.A2>") = 22. width = max(22+3, 12) = 25
        self.assertEqual(ws.column_dimensions["A"].width, 25.0)

    def test_generate_fraud_report_excel(self):
        # Populate fraud results
        t = Trainee(id="T001", name="Alice", doj=datetime.date(2026, 1, 1), scheme="NAPS")
        self.db.add(t)
        inv = InvoiceRecord(
            invoice_number="INV-FRD", invoice_date=datetime.date(2026, 2, 1), trainee_id="T001",
            billed_name="Alice", file_name="invoice.xlsx"
        )
        self.db.add(inv)
        self.db.commit()

        val = ValidationResult(
            invoice_record_id=inv.id, trainee_id="T001", rule_name="Test Fraud Rule",
            status="FRAUD", message="Validation fraud msg", created_at=datetime.datetime(2026, 2, 1, 10, 0, 0)
        )
        self.db.add(val)
        self.db.commit()

        report_bytes = ReportService.generate_fraud_report_excel(self.db)

        wb = openpyxl.load_workbook(io.BytesIO(report_bytes))
        self.assertIn("Fraud Incidents", wb.sheetnames)
        
        ws = wb["Fraud Incidents"]
        self.assertEqual(ws.max_row, 2)
        
        row_vals = [cell.value for cell in ws[2]]
        self.assertEqual(row_vals[0], "INV-FRD")
        self.assertEqual(row_vals[1], "T001")
        self.assertEqual(row_vals[2], "Alice")
        self.assertEqual(row_vals[3], "Test Fraud Rule")
        self.assertEqual(row_vals[4], "FRAUD / CRITICAL")
        self.assertEqual(row_vals[5], "N/A")
        self.assertEqual(row_vals[6], "Validation fraud msg")
        self.assertEqual(row_vals[8], "2026-02-01 10:00:00")

        # Verify identical column width calculation (total_rows = 2)
        # Col A (col_idx=1): max_len = len("<Cell 'Fraud Incidents'.A2>") = 27. width = max(27+3, 12) = 30
        self.assertEqual(ws.column_dimensions["A"].width, 30.0)

    def test_generate_payment_summary_excel(self):
        t1 = Trainee(id="T001", name="Alice", doj=datetime.date(2026, 1, 1), scheme="NAPS", status="ACTIVE")
        t2 = Trainee(id="T002", name="Bob", doj=datetime.date(2026, 1, 2), scheme="B.Tech", status="ACTIVE")
        self.db.add_all([t1, t2])
        self.db.commit()

        p1 = PaymentLedger(trainee_id="T001", invoice_number="INV-1", payment_type="JOINING", amount_paid=1200.0, payment_date=datetime.date(2026, 2, 1))
        p2 = PaymentLedger(trainee_id="T001", invoice_number="INV-2", payment_type="180_DAYS", amount_paid=600.0, payment_date=datetime.date(2026, 3, 1))
        p3 = PaymentLedger(trainee_id="T002", invoice_number="INV-1", payment_type="JOINING", amount_paid=1000.0, payment_date=datetime.date(2026, 2, 1))
        self.db.add_all([p1, p2, p3])
        self.db.commit()

        report_bytes = ReportService.generate_payment_summary_excel(self.db)

        wb = openpyxl.load_workbook(io.BytesIO(report_bytes))
        self.assertIn("Payment Ledgers", wb.sheetnames)
        
        ws = wb["Payment Ledgers"]
        # Max row is 3 (1 header, 2 data rows)
        self.assertEqual(ws.max_row, 3)

        # Check Alice's total disbursement (1200 + 600 = 1800)
        alice_row = [cell.value for cell in ws[2]]
        self.assertEqual(alice_row[0], "T001")
        self.assertEqual(alice_row[7], 1200.0)
        self.assertEqual(alice_row[8], 600.0)
        self.assertEqual(alice_row[9], 1800.0)
        self.assertEqual(alice_row[10], 0.0) # Remaining payout limit

        # Check Bob's total disbursement (1000 + 0 = 1000)
        bob_row = [cell.value for cell in ws[3]]
        self.assertEqual(bob_row[0], "T002")
        self.assertEqual(bob_row[7], 1000.0)
        self.assertEqual(bob_row[8], 0.0)
        self.assertEqual(bob_row[9], 1000.0)
        self.assertEqual(bob_row[10], 800.0) # Remaining payout limit (1800 - 1000 = 800)

        # Verify identical column width calculation (total_rows = 3)
        # Col A (col_idx=1): max_len = len("<Cell 'Payment Ledgers'.A3>") = 27. width = max(27+3, 12) = 30
        self.assertEqual(ws.column_dimensions["A"].width, 30.0)

    def test_generate_audit_report_excel(self):
        log = AuditLog(timestamp=datetime.datetime(2026, 2, 1, 12, 0, 0), action="TEST", module="REPORTS", details="Audit log test detail")
        self.db.add(log)
        self.db.commit()

        report_bytes = ReportService.generate_audit_report_excel(self.db)

        wb = openpyxl.load_workbook(io.BytesIO(report_bytes))
        self.assertIn("Audit Trails", wb.sheetnames)
        
        ws = wb["Audit Trails"]
        self.assertEqual(ws.max_row, 2)
        
        row_vals = [cell.value for cell in ws[2]]
        self.assertEqual(row_vals[0], "2026-02-01 12:00:00")
        self.assertEqual(row_vals[1], "TEST")
        self.assertEqual(row_vals[2], "REPORTS")
        self.assertEqual(row_vals[3], "Audit log test detail")

        # Verify identical column width calculation (total_rows = 2)
        # Col A (col_idx=1): max_len = len("<Cell 'Audit Trails'.A2>") = 24. width = max(24+3, 12) = 27
        self.assertEqual(ws.column_dimensions["A"].width, 27.0)

    def test_finance_analytics_data_and_exports(self):
        # 1. Seed Trainees
        t1 = Trainee(id="T001", name="Alice", doj=datetime.date(2026, 1, 1), status="ACTIVE", scheme="NAPS")
        t2 = Trainee(id="T002", name="Bob", doj=datetime.date(2026, 1, 2), status="BLOCKED", scheme="B.Tech")
        self.db.add_all([t1, t2])
        self.db.commit()

        # 2. Seed Invoice Records
        # Record 1: Alice, Billed 1900, Approved 1200, Rejected 700 (which has 100 kit/other)
        inv1 = InvoiceRecord(
            invoice_number="INV-001", invoice_date=datetime.date(2026, 2, 1), trainee_id="T001",
            billed_name="Alice", billed_joining_amount=1200.0, billed_180_days_amount=600.0,
            billed_other_amount=100.0, billed_total_amount=1900.0,
            approved_joining_amount=1200.0, approved_180_days_amount=0.0, approved_total_amount=1200.0,
            status="APPROVED", file_name="vendor_a.xlsx"
        )
        # Record 2: Bob, Billed 1800, Approved 0, Rejected 1800
        inv2 = InvoiceRecord(
            invoice_number="INV-001", invoice_date=datetime.date(2026, 2, 1), trainee_id="T002",
            billed_name="Bob", billed_joining_amount=1200.0, billed_180_days_amount=600.0,
            billed_other_amount=0.0, billed_total_amount=1800.0,
            approved_joining_amount=0.0, approved_180_days_amount=0.0, approved_total_amount=0.0,
            status="REJECTED", file_name="vendor_b.xlsx"
        )
        self.db.add_all([inv1, inv2])
        self.db.commit()

        # 3. Seed Validation Result for Bob (Fraud)
        val = ValidationResult(
            invoice_record_id=inv2.id, trainee_id="T002", rule_name="Employee Blocked",
            status="FRAUD", message="Employee is blocked", reason_code="EMPLOYEE_BLOCKED"
        )
        self.db.add(val)
        self.db.commit()

        # 4. Seed Payment Ledger for Alice
        p1 = PaymentLedger(
            trainee_id="T001", invoice_number="INV-001", payment_type="JOINING",
            amount_paid=1200.0, payment_date=datetime.date(2026, 2, 15)
        )
        self.db.add(p1)
        self.db.commit()

        # 5. Fetch Finance Analytics
        analytics = ReportService.get_finance_analytics_data(self.db)

        # 6. Verify overall counts & amounts
        summary = analytics["overall_summary"]
        self.assertEqual(summary["total_trainees"], 2)
        self.assertEqual(summary["active_trainees"], 1)
        self.assertEqual(summary["blocked_trainees"], 1)
        self.assertEqual(summary["total_billed"], 3700.0)
        self.assertEqual(summary["total_approved"], 1200.0)
        self.assertEqual(summary["total_rejected"], 2500.0)
        self.assertEqual(summary["total_paid"], 1200.0)

        # 7. Verify individual savings & liability
        self.assertEqual(analytics["fraud_savings"]["savings"], 1800.0)
        self.assertEqual(analytics["fraud_savings"]["count"], 1)
        self.assertEqual(analytics["blocked_employee_savings"]["savings"], 1800.0)
        self.assertEqual(analytics["blocked_employee_savings"]["count"], 1)
        self.assertEqual(analytics["kit_savings"]["savings"], 100.0)
        self.assertEqual(analytics["remaining_liability"], 600.0) # Alice limit 1800 - paid 1200 = 600. Bob is blocked so not active liability.

        # 8. Verify Excel Export
        excel_bytes = ReportService.generate_finance_analytics_excel(self.db)
        self.assertTrue(len(excel_bytes) > 0)
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        self.assertIn("Executive Summary", wb.sheetnames)
        self.assertIn("Monthly Spend Summary", wb.sheetnames)
        self.assertIn("Vendor Spend Summary", wb.sheetnames)
        self.assertIn("Category Spend Summary", wb.sheetnames)

        # 9. Verify CSV Export
        csv_bytes = ReportService.generate_finance_analytics_csv(self.db, "monthly_spend")
        self.assertTrue(len(csv_bytes) > 0)
        self.assertIn(b"Month,", csv_bytes)

        csv_summary = ReportService.generate_finance_analytics_csv(self.db, "summary")
        self.assertTrue(len(csv_summary) > 0)
        self.assertIn(b"Fraud Savings,", csv_summary)

        # 10. Verify PDF Export
        pdf_bytes = ReportService.generate_finance_analytics_pdf(self.db)
        self.assertTrue(len(pdf_bytes) > 0)
        self.assertEqual(pdf_bytes[:4], b"%PDF")

if __name__ == "__main__":
    unittest.main()

