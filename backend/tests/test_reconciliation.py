"""
Vendor Claimed

₹1800

↓

Approved

₹1200

↓

Rejected

₹600

↓

Reason

Joining payment exceeds policy

"""







import unittest
import datetime
import io
import openpyxl
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from backend.app.core.db import Base
from backend.app.models.models import Trainee, InvoiceRecord, PaymentLedger, ValidationResult
from backend.app.services.import_service import ImportService
from backend.app.services.validation_service import ValidationService
from backend.app.services.report_service import ReportService

class TestInvoiceReconciliation(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(self.engine)
        Session = sessionmaker(bind=self.engine)
        self.db = Session()

    def tearDown(self):
        self.db.close()
        Base.metadata.drop_all(self.engine)

    def _create_excel_bytes(self, data: list, columns: list) -> bytes:
        df = pd.DataFrame(data, columns=columns)
        out = io.BytesIO()
        df.to_excel(out, index=False)
        return out.getvalue()

    def test_vendor_invoice_reconciliation_flow(self):
        # 1. Seed the trainee database
        trainee_active = Trainee(
            id="T_REC_001",
            name="Alice Trainee",
            ticket_number="TKT_REC_99",
            doj=datetime.date(2025, 12, 1),
            scheme="NAPS",
            status="ACTIVE"
        )
        self.db.add(trainee_active)
        self.db.commit()

        # 2. Prepare mock Vendor Invoice spreadsheet bytes
        # Columns: Ticket Number, Employee Name, Billing Stage, Claimed Amount, Shirt Quantity, Jean Quantity
        headers = [
            "Ticket Number", "Employee Name", "Billing Stage", 
            "Claimed Amount", "Shirt Quantity", "Jean Quantity"
        ]
        
        # We will test three scenarios:
        # Row 1: Valid trainee, valid quantities -> Should pass validation cleanly
        # Row 2: Valid trainee, excess shirt quantity (3 shirts) -> Should raise Warning, but approved (since it's a kit warning, not a rejection unless other items)
        # Row 3: Unknown trainee ticket -> Should flag "Trainee Not Found"
        data = [
            ["TKT_REC_99", "Alice Trainee", "Joining", 1200.0, 2.0, 1.0],
            ["TKT_REC_99", "Alice Trainee", "Six Months", 600.0, 4.0, 1.0],
            ["TKT_REC_999", "Bob Unknown", "Joining", 1200.0, 1.0, 1.0]
        ]
        excel_bytes = self._create_excel_bytes(data, headers)

        # 3. Import workbook
        import_res = ImportService.import_invoice_workbook(
            db=self.db,
            file_content=excel_bytes,
            file_name="vendor_invoice_rec.xlsx",
            invoice_number_override="INV-REC-100"
        )

        self.assertEqual(import_res["records_imported"], 3)
        self.assertEqual(import_res["invoice_number"], "INV-REC-100")

        # 4. Verify imported records and their parsed quantities in extra_data
        records = self.db.query(InvoiceRecord).filter(InvoiceRecord.invoice_number == "INV-REC-100").all()
        self.assertEqual(len(records), 3)

        # Verify Joining Row
        rec_joining = [r for r in records if r.billed_joining_amount == 1200.0 and r.trainee_id == "T_REC_001"][0]
        self.assertEqual(rec_joining.extra_data["shirt_quantity"], 2.0)
        self.assertEqual(rec_joining.extra_data["jean_quantity"], 1.0)
        self.assertEqual(rec_joining.extra_data["ticket_number"], "TKT_REC_99")

        # Verify Six Months Row
        rec_six_months = [r for r in records if r.billed_180_days_amount == 600.0][0]
        self.assertEqual(rec_six_months.extra_data["shirt_quantity"], 4.0)

        # 5. Run validation engine
        validation_res = ValidationService.validate_invoice(self.db, "INV-REC-100")
        self.assertEqual(validation_res["total_records"], 3)

        # Refresh records from database
        self.db.refresh(rec_joining)
        self.db.refresh(rec_six_months)
        
        # Verify first row passed
        self.assertEqual(rec_joining.status, "VALIDATED")
        self.assertEqual(rec_joining.approved_joining_amount, 1200.0)

        # Verify second row (excess shirt qty warning)
        # It has a warning, but should still have approved amount if other conditions pass, and warning status
        self.assertEqual(rec_six_months.status, "VALIDATED")
        self.assertEqual(rec_six_months.approved_180_days_amount, 600.0)
        
        # Check warnings on second row
        warnings = self.db.query(ValidationResult).filter(
            ValidationResult.invoice_record_id == rec_six_months.id
        ).all()
        self.assertTrue(any("Excess kit quantity" in w.message for w in warnings))

        # Verify third row (trainee not found)
        rec_unknown = [r for r in records if r.trainee_id is None][0]
        self.db.refresh(rec_unknown)
        self.assertEqual(rec_unknown.status, "EXCEPTION")
        self.assertEqual(rec_unknown.approved_joining_amount, 0.0)
        self.assertEqual(rec_unknown.approved_total_amount, 0.0)

        # 6. Verify Exception Report in Reconciliation Format
        exception_excel_bytes = ReportService.generate_exception_report_excel(
            db=self.db,
            invoice_number="INV-REC-100",
            format_version="reconciliation"
        )
        
        wb_exc = openpyxl.load_workbook(io.BytesIO(exception_excel_bytes), read_only=True, data_only=True)
        ws_exc = wb_exc.active
        self.assertEqual(ws_exc.title, "Exceptions")
        
        # Read headers
        rows_exc = list(ws_exc.iter_rows(values_only=True))
        headers_exc = rows_exc[0]
        expected_exc_headers = ["Ticket", "Name", "Vendor Claim (₹)", "Approved Amount (₹)", "Rejected Amount (₹)", "Failure Reasons"]
        self.assertEqual(list(headers_exc), expected_exc_headers)
        
        # Verify content: Row 2 (Six Months with warning) and Row 3 (Unknown trainee) must be in the Exception Report
        # Row 2 (index 1) or Row 3 (index 2)
        exc_data_rows = rows_exc[1:]
        self.assertEqual(len(exc_data_rows), 2)
        
        tickets_in_exc = [r[0] for r in exc_data_rows]
        self.assertIn("TKT_REC_99", tickets_in_exc)
        self.assertIn("TKT_REC_999", tickets_in_exc)

        # 7. Verify Payment Summary Report in Reconciliation Format
        summary_excel_bytes = ReportService.generate_payment_summary_excel(
            db=self.db,
            invoice_number="INV-REC-100",
            format_version="reconciliation"
        )
        
        wb_sum = openpyxl.load_workbook(io.BytesIO(summary_excel_bytes), read_only=True, data_only=True)
        ws_sum = wb_sum.active
        self.assertEqual(ws_sum.title, "Payment Summary")
        
        rows_sum = list(ws_sum.iter_rows(values_only=True))
        headers_sum = rows_sum[0]
        expected_sum_headers = [
            "Invoice Total (₹)", "Approved Total (₹)", "Rejected Total (₹)", 
            "Money Saved (₹)", "Number Approved", "Number Rejected"
        ]
        self.assertEqual(list(headers_sum), expected_sum_headers)
        
        # Values checks
        # Invoice Total: 1200 + 600 + 1200 = 3000
        # Approved Total: 1200 + 600 + 0 = 1800
        # Rejected Total: 1200
        # Money Saved: 1200
        # Number Approved: 2 (Row 1 & Row 2 have approved amounts > 0)
        # Number Rejected: 1 (Row 3 has approved amount == 0)
        values_sum = list(rows_sum[1])
        self.assertEqual(values_sum[0], 3000.0)
        self.assertEqual(values_sum[1], 1800.0)
        self.assertEqual(values_sum[2], 1200.0)
        self.assertEqual(values_sum[3], 1200.0)
        self.assertEqual(values_sum[4], 2)
        self.assertEqual(values_sum[5], 1)

    def test_rehire_lifecycle_archiving_and_rule_scoping(self):
        # 1. Seed trainee in a separated state
        trainee = Trainee(
            id="T_REHIRE_01",
            name="Bob Rehire",
            ticket_number="TKT_REHIRE_1",
            doj=datetime.date(2024, 7, 15),
            dol=datetime.date(2025, 1, 15),
            scheme="NAPS",
            status="SEPARATED"
        )
        self.db.add(trainee)
        self.db.commit()

        # Seed separation history record
        from backend.app.models.models import SeparationRecord
        sep_record = SeparationRecord(
            trainee_id="T_REHIRE_01",
            dol=datetime.date(2025, 1, 15),
            reason="Separated",
            file_name="sep_prev.xlsx",
            extra_data={"sheet": "NAPS", "tenure": 184}
        )
        self.db.add(sep_record)

        # Seed historical invoice record and ledger
        inv_record = InvoiceRecord(
            trainee_id="T_REHIRE_01",
            invoice_number="INV-OLD-999",
            invoice_date=datetime.date(2024, 8, 1),
            billed_joining_amount=1200.0,
            billed_180_days_amount=0.0,
            billed_total_amount=1200.0,
            approved_joining_amount=1200.0,
            approved_180_days_amount=0.0,
            approved_total_amount=1200.0,
            status="APPROVED",
            file_name="inv_old.xlsx"
        )
        self.db.add(inv_record)
        self.db.commit()

        ledger_entry = PaymentLedger(
            trainee_id="T_REHIRE_01",
            invoice_number="INV-OLD-999",
            payment_type="JOINING",
            amount_paid=1200.0,
            payment_date=datetime.date(2024, 8, 5)
        )
        self.db.add(ledger_entry)

        val_result = ValidationResult(
            invoice_record_id=inv_record.id,
            trainee_id="T_REHIRE_01",
            rule_name="Joining Limit",
            status="WARNING",
            message="Compliant payment"
        )
        self.db.add(val_result)
        self.db.commit()

        # 2. Simulate rehire with a new DOJ > previous DOL (New DOJ: 2025-06-01)
        headers = ["Trainee ID", "Trainee Name", "Date of Joining", "Category", "Aadhaar", "Ticket Number"]
        data = [
            ["T_REHIRE_01", "Bob Rehire", "2025-06-01", "NAPS", "111122223333", "TKT_REHIRE_2"]
        ]
        excel_bytes = self._create_excel_bytes(data, headers)

        import_res = ImportService.import_bdc_workbook(
            db=self.db,
            file_content=excel_bytes,
            file_name="bdc_rehire.xlsx"
        )
        self.assertEqual(import_res["updated_records"], 1)

        # 3. Verify trainee master is active and previous lifecycle is archived
        self.db.refresh(trainee)
        self.assertEqual(trainee.status, "ACTIVE")
        self.assertEqual(trainee.doj, datetime.date(2025, 6, 1))
        self.assertIsNone(trainee.dol)
        self.assertEqual(trainee.ticket_number, "TKT_REHIRE_2")

        lifecycles = trainee.extra_data.get("lifecycles", [])
        self.assertEqual(len(lifecycles), 1)
        prev_lc = lifecycles[0]
        self.assertEqual(prev_lc["lifecycle_number"], 1)
        self.assertEqual(prev_lc["doj"], "2024-07-15")
        self.assertEqual(prev_lc["dol"], "2025-01-15")
        self.assertEqual(prev_lc["status"], "SEPARATED")
        self.assertEqual(len(prev_lc["invoice_history"]), 1)
        self.assertEqual(len(prev_lc["payment_ledger"]), 1)
        self.assertEqual(len(prev_lc["validation_history"]), 1)

        # 4. Scoping check: Validate a new invoice for Bob's new lifecycle
        # Bob claims ₹1200 for Joining in the new lifecycle (date: 2025-07-01)
        headers_inv = ["Ticket Number", "Employee Name", "Billing Stage", "Claimed Amount"]
        data_inv = [
            ["TKT_REHIRE_2", "Bob Rehire", "Joining", 1200.0]
        ]
        excel_inv_bytes = self._create_excel_bytes(data_inv, headers_inv)

        ImportService.import_invoice_workbook(
            db=self.db,
            file_content=excel_inv_bytes,
            file_name="inv_new.xlsx",
            invoice_number_override="INV-NEW-99"
        )

        validation_res = ValidationService.validate_invoice(self.db, "INV-NEW-99")
        self.assertEqual(validation_res["fraud_count"], 0)
        self.assertEqual(validation_res["error_count"], 0)

        # Check that Bob's joining payout is approved because history is scoped to the current lifecycle
        new_inv_rec = self.db.query(InvoiceRecord).filter(
            InvoiceRecord.invoice_number == "INV-NEW-99",
            InvoiceRecord.trainee_id == "T_REHIRE_01"
        ).one()
        self.assertEqual(new_inv_rec.status, "VALIDATED")
        self.assertEqual(new_inv_rec.approved_joining_amount, 1200.0)

    def test_new_reports_generation(self):
        # Seed a trainee and invoice record
        trainee = Trainee(
            id="T_REP_002",
            name="John Report",
            ticket_number="TKT_REP_02",
            doj=datetime.date(2025, 1, 1),
            scheme="B.Tech",
            status="ACTIVE"
        )
        self.db.add(trainee)
        self.db.commit()

        inv = InvoiceRecord(
            trainee_id="T_REP_002",
            invoice_number="INV-REP-101",
            invoice_date=datetime.date(2025, 2, 1),
            billed_joining_amount=1200.0,
            billed_total_amount=1200.0,
            approved_joining_amount=1200.0,
            approved_total_amount=1200.0,
            status="APPROVED",
            file_name="inv_rep_101.xlsx",
            extra_data={"ticket_number": "TKT_REP_02"}
        )
        self.db.add(inv)
        self.db.commit()

        # Seed payment ledger
        ledger = PaymentLedger(
            trainee_id="T_REP_002",
            invoice_number="INV-REP-101",
            payment_type="JOINING",
            amount_paid=1200.0,
            payment_date=datetime.date(2025, 2, 5)
        )
        self.db.add(ledger)
        self.db.commit()

        # 1. Corrected Invoice report
        corrected_bytes = ReportService.generate_corrected_invoice_excel(self.db, "INV-REP-101")
        wb = openpyxl.load_workbook(io.BytesIO(corrected_bytes), read_only=True, data_only=True)
        ws = wb.active
        self.assertEqual(ws.title, "Corrected Invoice")
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        self.assertIn("Original Value (₹)", headers)
        self.assertIn("Approved Value (₹)", headers)
        self.assertIn("Rejected Value (₹)", headers)
        
        # 2. Vendor Payment Summary report
        vendor_sum_bytes = ReportService.generate_vendor_payment_summary_excel(self.db, "INV-REP-101")
        wb_v = openpyxl.load_workbook(io.BytesIO(vendor_sum_bytes), read_only=True, data_only=True)
        ws_v = wb_v.active
        self.assertEqual(ws_v.title, "Vendor Payment Summary")
        rows_v = list(ws_v.iter_rows(values_only=True))
        headers_v = rows_v[0]
        self.assertIn("Cumulative Total Disbursed (₹)", headers_v)
        self.assertIn("Remaining Payout Limit (₹)", headers_v)

    def test_dashboard_stats_endpoint_aggregation(self):
        # Clear existing data in memory
        self.db.query(Trainee).delete()
        self.db.query(InvoiceRecord).delete()
        self.db.commit()

        # Seed Trainees
        # Today DOJ, active
        t1 = Trainee(id="T_DB_1", name="Active 1", doj=datetime.date.today(), scheme="B.Tech", status="ACTIVE")
        # Today DOL, early exit (DOJ is today, so same month and under 30 days)
        t2 = Trainee(
            id="T_DB_2",
            name="Early Exit",
            doj=datetime.date.today(),
            dol=datetime.date.today(),
            scheme="NAPS",
            status="SEPARATED"
        )
        # Blocked
        t3 = Trainee(id="T_DB_3", name="Blocked 1", doj=datetime.date(2025, 1, 1), scheme="M.Tech", status="BLOCKED")
        self.db.add_all([t1, t2, t3])
        
        # Seed separation records
        from backend.app.models.models import SeparationRecord
        sep = SeparationRecord(
            trainee_id="T_DB_2",
            dol=datetime.date.today(),
            reason="Early Exit",
            file_name="sep.xlsx",
            extra_data={"sheet": "NAPS", "tenure": 15}
        )
        self.db.add(sep)
        self.db.commit()

        # Call get_dashboard_stats endpoint directly
        from backend.app.api.endpoints import get_dashboard_stats
        stats_res = get_dashboard_stats(self.db)

        self.assertEqual(stats_res["total_trainees"], 3)
        self.assertEqual(stats_res["active_trainees"], 1)
        self.assertEqual(stats_res["separated_trainees"], 1)
        self.assertEqual(stats_res["blocked_trainees"], 1)
        self.assertEqual(stats_res["btech_count"], 1)
        self.assertEqual(stats_res["naps_count"], 1)
        self.assertEqual(stats_res["mtech_count"], 1)
        self.assertEqual(stats_res["early_separations"], 1)
        self.assertEqual(stats_res["joining_this_month"], 2) # t1 and t2 joined this month (today and 15 days ago)
        self.assertEqual(stats_res["separations_this_month"], 1) # t2 separated today

