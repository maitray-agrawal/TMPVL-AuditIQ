import unittest
import datetime
import io
import openpyxl
from unittest.mock import MagicMock, patch
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from backend.app.core.db import Base
from backend.app.models.models import Trainee, InvoiceRecord
from backend.app.services.import_service import ImportService
from backend.app.services.workbook_parser import WorkbookParser

class TestInvoiceRefactoring(unittest.TestCase):
    def setUp(self):
        # Set up memory SQLite database
        self.engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(self.engine)
        Session = sessionmaker(bind=self.engine)
        self.db = Session()

        # Seed standard trainees for testing
        self.trainee_1 = Trainee(
            id="T_REC_001",
            name="Alice Trainee",
            ticket_number="TKT_REC_99",
            doj=datetime.date(2026, 6, 1),
            scheme="NAPS",
            status="ACTIVE"
        )
        self.trainee_2 = Trainee(
            id="T_REC_002",
            name="Bob Trainee",
            ticket_number="TKT_REC_100",
            doj=datetime.date(2025, 12, 1),
            scheme="NAPS",
            status="ACTIVE"
        )
        self.db.add(self.trainee_1)
        self.db.add(self.trainee_2)
        self.db.commit()

    def tearDown(self):
        self.db.close()
        Base.metadata.drop_all(self.engine)

    def _create_excel_bytes(self, rows, headers, sheet_name="Invoice", sheet_state="visible") -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.sheet_state = sheet_state
        
        ws.append(headers)
        for r in rows:
            ws.append(r)
            
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    def test_parse_pair_garment_counts(self):
        """Test textual representation parsing of garment counts."""
        # 1. Exact matches: jeans and shirts specified
        res = ImportService.parse_pair("2 Pair Jeans & 2 Pair T Shirt")
        self.assertEqual(res["jeans_count"], 2.0)
        self.assertEqual(res["shirt_count"], 2.0)
        self.assertEqual(res["pair_count"], 2.0)

        # 2. Generic match
        res2 = ImportService.parse_pair("2 Pair")
        self.assertEqual(res2["jeans_count"], 2.0)
        self.assertEqual(res2["shirt_count"], 2.0)
        self.assertEqual(res2["pair_count"], 2.0)

        # 3. Numeric match
        res3 = ImportService.parse_pair("1.5")
        self.assertEqual(res3["jeans_count"], 1.5)
        self.assertEqual(res3["shirt_count"], 1.5)
        self.assertEqual(res3["pair_count"], 1.5)

        # 4. Partial matches (one garment specified)
        res4 = ImportService.parse_pair("1 Pair Jeans")
        self.assertEqual(res4["jeans_count"], 1.0)
        self.assertEqual(res4["shirt_count"], 0.0)
        self.assertEqual(res4["pair_count"], 1.0)

        # 5. Invalid / blank input
        res5 = ImportService.parse_pair(None)
        self.assertEqual(res5["jeans_count"], 0.0)
        self.assertEqual(res5["shirt_count"], 0.0)
        self.assertEqual(res5["pair_count"], 0.0)

    def test_excel_import_synonyms_matching(self):
        """Test Excel import with synonyms mapped to appropriate model attributes."""
        headers = [
            "T No", "Candidates Name", "Date Of Joining", 
            "Batch No", "Kit Pair", "Bill Amount", "Distribution Date", "Page No"
        ]
        rows = [
            ["TKT_REC_99", "Alice Trainee", "2026-06-01", "Batch A", "2 Pair Jeans & 2 Pair T Shirt", 1200.0, "2026-07-01", 1]
        ]
        excel_bytes = self._create_excel_bytes(rows, headers)

        res = ImportService.import_invoice_workbook(
            db=self.db,
            file_content=excel_bytes,
            file_name="synonyms_invoice.xlsx"
        )

        self.assertEqual(res["inserted_records"], 1)
        self.assertEqual(res["errors"], [])

        # Retrieve record
        rec = self.db.query(InvoiceRecord).first()
        self.assertIsNotNone(rec)
        self.assertEqual(rec.trainee_id, "T_REC_001")
        self.assertEqual(rec.billed_name, "Alice Trainee")
        self.assertEqual(rec.billed_joining_amount, 1200.0)
        self.assertEqual(rec.extra_data["batch"], "Batch A")
        self.assertEqual(rec.extra_data["page_number"], "1")
        self.assertEqual(rec.extra_data["shirt_quantity"], 2.0)
        self.assertEqual(rec.extra_data["jean_quantity"], 2.0)

    def test_excel_date_stage_apportionment(self):
        """Test stage apportionment based on joining date vs distribution date."""
        # 1. Under 180 days tenure (Joining payout)
        headers = ["Ticket No", "Candidate Name", "DOJ", "Batch", "Uniform Pair", "Billing Amount", "Distribution Date"]
        rows = [
            # 30 days tenure -> should be mapped to billed_joining_amount
            ["TKT_REC_99", "Alice Trainee", "2026-06-01", "Batch A", "2 Pair", 1200.0, "2026-07-01"],
            # 212 days tenure -> should be mapped to billed_180_days_amount
            ["TKT_REC_100", "Bob Trainee", "2025-12-01", "Batch B", "2 Pair", 600.0, "2026-07-01"]
        ]
        excel_bytes = self._create_excel_bytes(rows, headers)

        res = ImportService.import_invoice_workbook(
            db=self.db,
            file_content=excel_bytes,
            file_name="date_stage_invoice.xlsx"
        )
        self.assertEqual(res["inserted_records"], 2)

        records = self.db.query(InvoiceRecord).all()
        rec_alice = [r for r in records if r.trainee_id == "T_REC_001"][0]
        rec_bob = [r for r in records if r.trainee_id == "T_REC_002"][0]

        # Alice: tenure < 180 days
        self.assertEqual(rec_alice.billed_joining_amount, 1200.0)
        self.assertEqual(rec_alice.billed_180_days_amount, 0.0)

        # Bob: tenure >= 180 days
        self.assertEqual(rec_bob.billed_joining_amount, 0.0)
        self.assertEqual(rec_bob.billed_180_days_amount, 600.0)

    def test_excel_legacy_billing_stage(self):
        """Test fallback stage parsing using the legacy Billing Stage / columns."""
        # Scenario 1: Using Billing Stage column
        headers = ["Ticket", "Employee Name", "Billing Stage", "Amount"]
        rows = [
            ["TKT_REC_99", "Alice Trainee", "Joining", 1200.0],
            ["TKT_REC_100", "Bob Trainee", "180 Days", 600.0]
        ]
        excel_bytes = self._create_excel_bytes(rows, headers)

        res = ImportService.import_invoice_workbook(
            db=self.db,
            file_content=excel_bytes,
            file_name="legacy_stage_invoice.xlsx"
        )
        self.assertEqual(res["inserted_records"], 2)

        records = self.db.query(InvoiceRecord).all()
        rec_alice = [r for r in records if r.trainee_id == "T_REC_001"][0]
        rec_bob = [r for r in records if r.trainee_id == "T_REC_002"][0]

        self.assertEqual(rec_alice.billed_joining_amount, 1200.0)
        self.assertEqual(rec_bob.billed_180_days_amount, 600.0)

    @patch("pdfplumber.open")
    def test_pdf_import_synonym_matching(self, mock_pdf_open):
        """Test PDF table extraction and synonym matching ingestion."""
        mock_pdf = MagicMock()
        mock_page = MagicMock()
        
        # PDF page extracts tables with synonyms headers
        mock_page.extract_tables.return_value = [
            [
                ["T No", "Candidates Name", "Date Of Joining", "Batch No", "Pair", "Amount", "Distribution Date", "Page No"],
                ["TKT_REC_99", "Alice Trainee", "2026-06-01", "Batch PDF", "2 Pair Jeans & 2 Pair T Shirt", "1200", "2026-07-01", "4"]
            ]
        ]
        mock_pdf.pages = [mock_page]
        mock_pdf_open.return_value.__enter__.return_value = mock_pdf

        res = ImportService.import_invoice_workbook(
            db=self.db,
            file_content=b"dummy pdf data",
            file_name="test_invoice.pdf"
        )

        self.assertEqual(res["inserted_records"], 1)
        self.assertEqual(res["errors"], [])

        # Check DB values
        rec = self.db.query(InvoiceRecord).first()
        self.assertIsNotNone(rec)
        self.assertEqual(rec.trainee_id, "T_REC_001")
        self.assertEqual(rec.billed_name, "Alice Trainee")
        self.assertEqual(rec.billed_joining_amount, 1200.0)
        self.assertEqual(rec.extra_data["batch"], "Batch PDF")
        self.assertEqual(rec.extra_data["page_number"], "4")

    def test_visible_sheets_iteration(self):
        """Test visible sheets processing and hidden sheet avoidance."""
        # Create Excel with one visible Invoice sheet, one hidden Invoice sheet
        wb = openpyxl.Workbook()
        ws_vis = wb.active
        ws_vis.title = "Visible Invoice"
        ws_vis.append(["Ticket No", "Candidate Name", "DOJ", "Batch", "Uniform Pair", "Billing Amount", "Distribution Date"])
        ws_vis.append(["TKT_REC_99", "Alice Trainee", "2026-06-01", "Batch A", "2 Pair", 1200.0, "2026-07-01"])

        ws_hid = wb.create_sheet(title="Hidden Invoice")
        ws_hid.sheet_state = "hidden"
        ws_hid.append(["Ticket No", "Candidate Name", "DOJ", "Batch", "Uniform Pair", "Billing Amount", "Distribution Date"])
        ws_hid.append(["TKT_REC_100", "Bob Trainee", "2025-12-01", "Batch B", "2 Pair", 600.0, "2026-07-01"])

        out = io.BytesIO()
        wb.save(out)
        excel_bytes = out.getvalue()

        res = ImportService.import_invoice_workbook(
            db=self.db,
            file_content=excel_bytes,
            file_name="multi_sheets.xlsx"
        )

        # Alice is in the visible sheet -> should be imported
        # Bob is in the hidden sheet -> should be ignored
        self.assertEqual(res["inserted_records"], 1)
        records = self.db.query(InvoiceRecord).all()
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].trainee_id, "T_REC_001")

    def test_workbook_robustness_handling(self):
        """Test parser robustness with formula errors, empty rows, and missing values."""
        headers = ["Ticket No", "Candidate Name", "DOJ", "Batch", "Uniform Pair", "Billing Amount", "Distribution Date"]
        rows = [
            # Valid row
            ["TKT_REC_99", "Alice Trainee", "2026-06-01", "Batch A", "2 Pair", 1200.0, "2026-07-01"],
            # Empty row
            [None, None, None, None, None, None, None],
            # Row with formula error in date
            ["TKT_REC_100", "Bob Trainee", "#VALUE!", "Batch B", "2 Pair", 600.0, "2026-07-01"]
        ]
        excel_bytes = self._create_excel_bytes(rows, headers)

        res = ImportService.import_invoice_workbook(
            db=self.db,
            file_content=excel_bytes,
            file_name="robustness_invoice.xlsx"
        )

        # 1 valid row should succeed. The empty row is skipped.
        # The row with a formula error should succeed but default to invoice date for stage apportionment.
        self.assertEqual(res["inserted_records"], 2)
        records = self.db.query(InvoiceRecord).all()
        self.assertEqual(len(records), 2)
