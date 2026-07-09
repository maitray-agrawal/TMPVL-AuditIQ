import unittest
import datetime
import io
import openpyxl
from backend.app.services.workbook_parser import WorkbookParser

class TestWorkbookParser(unittest.TestCase):
    def _create_mock_workbook(self) -> bytes:
        """Helper to construct an in-memory Excel workbook with various test sheets."""
        wb = openpyxl.Workbook()
        
        # 1. BDC Sheet (visible)
        ws_bdc = wb.active
        ws_bdc.title = "BDC Trainees"
        
        # Add some headers and merged cells
        ws_bdc.append(["Trainee ID", "Trainee Name", "Date of Joining", "Aadhaar Card", "Shop Floor"])
        ws_bdc.append(["T001", "Alice Smith", "2026-01-15", "1234-5678-9012", "Welding"])
        ws_bdc.append(["T002", "Bob Johnson", 46020.0, "9876 5432 1098", "Assembly"])  # 46020.0 Excel date is 2025-12-30
        
        # Add a merged cell: merging Shop Floor for the next two rows
        ws_bdc.merge_cells("E3:E4")
        ws_bdc.cell(row=3, column=5, value="Paint Shop")
        ws_bdc.cell(row=4, column=1, value="T003")
        ws_bdc.cell(row=4, column=2, value="Charlie Brown")
        ws_bdc.cell(row=4, column=3, value="10/02/2026")
        ws_bdc.cell(row=4, column=4, value="111122223333")
        
        # 2. Separation Sheet (visible)
        ws_sep = wb.create_sheet(title="Separations")
        ws_sep.append(["Personnel No", "Date of Leaving", "Reason"])
        ws_sep.append(["T001", "2026-02-15", "Resigned"])
        # Add a formula error row
        ws_sep.append(["T002", "#REF!", "Terminated"])
        
        # 3. Invoice Sheet (visible)
        ws_inv = wb.create_sheet(title="January Invoice")
        ws_inv.append(["Ticket", "Name", "Joining Payment", "180 Days Payment", "Shirt Qty", "Jeans Qty"])
        ws_inv.append(["TKT100", "Alice Smith", 1200.0, 0.0, 3, 2])
        # Add a blank row to make sure it's skipped
        ws_inv.append([None, None, None, None, None, None])
        
        # 4. Hidden Sheet
        ws_hidden = wb.create_sheet(title="Hidden Data")
        ws_hidden.sheet_state = "hidden"
        ws_hidden.append(["Trainee ID", "Trainee Name", "Date of Joining"])
        ws_hidden.append(["T004", "Hidden Trainee", "2026-01-01"])

        # 5. Unknown Sheet
        ws_unknown = wb.create_sheet(title="Metadata")
        ws_unknown.append(["Config Name", "Value"])
        ws_unknown.append(["Batch Size", 2000])

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    def test_workbook_parser_detection_and_normalization(self):
        excel_bytes = self._create_mock_workbook()
        parsed = WorkbookParser.parse_workbook(excel_bytes, "mock_test.xlsx")
        
        stats = parsed["stats"]
        self.assertEqual(stats["workbook_name"], "mock_test.xlsx")
        # Total sheets: 5
        self.assertEqual(stats["number_of_sheets"], 5)
        # Processed sheets: BDC Trainees, Separations, January Invoice
        self.assertIn("BDC Trainees", stats["sheets_processed"])
        self.assertIn("Separations", stats["sheets_processed"])
        self.assertIn("January Invoice", stats["sheets_processed"])
        # Skipped sheets: Hidden Data (hidden)
        self.assertIn("Hidden Data", stats["skipped_sheets"])
        self.assertIn("Metadata", stats["sheets_processed"])
        
        sheets_map = {s["sheet_name"]: s for s in parsed["sheets"]}
        
        # Verify Generic sheet parsing
        metadata_sheet = sheets_map["Metadata"]
        self.assertEqual(metadata_sheet["sheet_type"], "Unknown")
        self.assertEqual(metadata_sheet["valid_rows"], 1)
        
        # Verify BDC sheet parsing
        bdc_sheet = sheets_map["BDC Trainees"]
        self.assertEqual(bdc_sheet["sheet_type"], "Employee Master")
        self.assertEqual(bdc_sheet["valid_rows"], 3)  # Row 2, Row 3, Row 4
        
        rows = bdc_sheet["rows"]
        # Row 1 (index 0): T001
        self.assertEqual(rows[0]["traineeid"], "T001")
        self.assertEqual(rows[0]["traineename"], "Alice Smith")
        self.assertEqual(rows[0]["dateofjoining"], datetime.date(2026, 1, 15))
        self.assertEqual(rows[0]["aadhaarcard"], "1234-5678-9012")
        self.assertEqual(rows[0]["shopfloor"], "Welding")
        
        # Row 2 (index 1): T002 with Excel date
        self.assertEqual(rows[1]["traineeid"], "T002")
        self.assertEqual(rows[1]["dateofjoining"], datetime.date(2025, 12, 29))
        
        # Row 3 (index 2): T003 with merged cell value "Paint Shop"
        self.assertEqual(rows[2]["traineeid"], "T003")
        self.assertEqual(rows[2]["shopfloor"], "Paint Shop")  # Merged cell lookup test passed!
        self.assertEqual(rows[2]["dateofjoining"], datetime.date(2026, 2, 10))

        # Verify Separation sheet parsing
        sep_sheet = sheets_map["Separations"]
        self.assertEqual(sep_sheet["sheet_type"], "Separation")
        self.assertEqual(sep_sheet["valid_rows"], 2)
        
        sep_rows = sep_sheet["rows"]
        self.assertEqual(sep_rows[0]["personnelno"], "T001")
        self.assertEqual(sep_rows[0]["dateofleaving"], datetime.date(2026, 2, 15))
        
        # Formula error row normalization
        self.assertEqual(sep_rows[1]["personnelno"], "T002")
        self.assertIsNone(sep_rows[1]["dateofleaving"])  # #REF! resolved to None

        # Verify Invoice sheet parsing
        inv_sheet = sheets_map["January Invoice"]
        self.assertEqual(inv_sheet["sheet_type"], "Invoice")
        self.assertEqual(inv_sheet["valid_rows"], 1)  # Blank row skipped!
        self.assertGreaterEqual(inv_sheet["blank_rows"], 1)
        
        inv_rows = inv_sheet["rows"]
        self.assertEqual(inv_rows[0]["ticket"], "TKT100")
        self.assertEqual(inv_rows[0]["joiningpayment"], 1200.0)
        self.assertEqual(inv_rows[0]["shirtqty"], 3.0)

if __name__ == "__main__":
    unittest.main()
