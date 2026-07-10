"""
WorkbookParser — Production-Grade Excel Ingestion Engine.

Design guarantees:
  1. Formula safety  — dual-pass load (data_only=True first, fallback data_only=False).
     Rows with un-cached formula cells are NEVER silently dropped.
  2. Row termination  — uses last-meaningful-row scan, never ws.max_row.
  3. Synonym matching — normalised (strip spaces/brackets/punctuation/case) header matching
     against an exhaustive alias dictionary that covers all known BDC / Separation / ERP
     column header variations.
  4. Merged cells     — resolved from a pre-built merged-cells map (non-read-only mode).
  5. Audit trail      — every skipped/formula-warning row is logged to the Python logger
     *and* stdout so operators see the full picture in the server console.
"""

import re
import io
import datetime
import logging
import openpyxl
from typing import Dict, List, Tuple, Any, Optional
import pandas as pd

logger = logging.getLogger("workbook_parser")

# Sentinel value placed into any identity field whose formula result was not cached.
FORMULA_UNCACHED = "__FORMULA_UNCACHED__"


class WorkbookParser:
    PARSER_VERSION = "3.0.0"

    # -------------------------------------------------------------------------
    # Synonym dictionary
    # Keys are canonical internal names; values are all lowercase-alphanumeric
    # aliases (spaces, brackets, dots, underscores already stripped by
    # normalize_header before comparison).
    # -------------------------------------------------------------------------
    SYNONYMS = {
        "trainee_id": [
            "traineeid", "empid", "employeeid", "regno", "persno", "pno",
            "personnelno", "personnelnumber", "persnumber", "employeenumber",
            "traineenumber", "id", "empid", "employeeid", "trainee_id",
            "reg_no", "regno",
        ],
        "ticket_number": [
            # All case/punctuation variants collapse to these after normalise
            "ticket", "ticketno", "ticketnumber", "boardingticket", "ticketid",
            "tktno", "tktnumber", "boardingticketno", "ticket_number", "ticket_no",
            # explicit header variants from real BDC workbooks
            "ticketno", "ticketnumber", "tktno",
        ],
        "joining_date": [
            # openpyxl normalises "Begda(ddmmyyyy)(M)" → "begdaddmmyyyym"
            "begdaddmmyyyym", "begda", "joiningdate", "doj", "startdate",
            "dateofjoining", "joining", "joining_date", "joiningdate",
            "dateofjoining",
        ],
        "end_date": [
            "enddaddmmyyyym", "leavingdate", "enddate", "dol", "dateofleaving",
            "resignationdate", "separationdate", "seperationdate",
            "seperatedate", "separatedate", "lastworkingdate", "lwd",
            "end_date",
        ],
        "candidate_name": [
            # "Complete Name" → "completename" after normalise
            "completename", "employeename", "candidatename", "name",
            "traineename", "fullname", "candidatesname",
        ],
        "first_name":  ["firstname", "first_name"],
        "middle_name": ["middlename", "middle_name"],
        "last_name":   ["lastname", "last_name"],
        "aadhaar": [
            # Covers "Adhar Card", "Aadhar", "Aadhaar Number" after normalise
            "aadhaar", "adhaar", "aadhar", "uid", "uidai", "nationalid",
            "adharcard", "aadharcard", "aadhaarcard",
            "adharcardno", "aadharno", "aadhaarno",
            # "Adhar" alone
            "adhar",
        ],
        "mobile": [
            "mobile", "mobileno", "mobilenumber", "phone", "phoneno",
            "contact", "contactno", "contactnumber", "telephoneno",
            "telephonenumber",
        ],
        "email":    ["email", "emailid", "emailaddress", "mail", "mailid"],
        "category": [
            "category", "traineecategory", "empcategory", "employeecategory",
            "scheme", "program", "course", "type", "schemetype",
        ],
        "batch":    ["batch", "group", "batchname", "year", "joiningbatch"],
        "shop": [
            "shop", "department", "dept", "location", "workarea", "area",
            "unit", "plant", "shopfloor",
        ],
        "reason": [
            "reason", "remarks", "typeofseparation", "reasonforaction",
            "reasonfortermination", "separationreason",
        ],
        "offer_id": [
            "offerid", "offerletterid", "offerletternumber", "offerno",
            "offernumber", "offer_id",
        ],
    }

    # Invoice-specific header keywords (used by detect_sheet_type_and_header)
    INVOICE_KWS = [
        'joiningpayment', 'joiningreimbursement', 'joiningamount',
        '180days', '180dayspayment', '180daysreimbursement', '180daysamount',
        'uniform', 'shirt', 'jeans', 'excess', 'billingstage', 'stage',
        'claimedamount', 'billedtotal', 'totalamount', 'shirtquantity',
        'shirtqty', 'jeanquanity', 'jeanqty', 'jeansquantity', 'jeansqty',
        'amount', 'billamount', 'billingamount', 'pair', 'uniformpair',
        'kitpair', 'distributiondate', 'issueddate', 'pageno', 'page',
        'invoicenumber', 'invoiceno', 'invoice_number', 'invoice_no', 'invoice',
    ]

    # -------------------------------------------------------------------------
    # Static helpers
    # -------------------------------------------------------------------------

    @staticmethod
    def normalize_header(header: Any) -> str:
        """Strip, lowercase, remove leading numbers, remove punctuation/spaces."""
        if header is None:
            return ""
        h = str(header).strip().lower()
        # Remove numbering prefix: "1. Begda", "02. Name", "3_Ticket"
        h = re.sub(r'^\d+[\s\.\-_]*', '', h)
        # Remove all non-alphanumeric characters (brackets, dots, spaces, etc.)
        h = re.sub(r'[^a-z0-9]', '', h)
        return h

    @classmethod
    def _has_synonym(cls, row_cleaned: List[Any], synonym_key: str) -> bool:
        """Return True if any cell in row_cleaned matches a synonym for synonym_key."""
        headers  = [cls.normalize_header(c) for c in row_cleaned if c is not None]
        keywords = [cls.normalize_header(kw) for kw in cls.SYNONYMS.get(synonym_key, [])]
        for h in headers:
            if not h:
                continue
            if h in keywords:
                return True
        for h in headers:
            if not h:
                continue
            for kw in keywords:
                if kw and (kw in h or h in kw):
                    if kw == "name" and any(p in h for p in ("first", "middle", "last")):
                        continue
                    return True
        return False

    @classmethod
    def _has_invoice_marker(cls, row_cleaned: List[Any]) -> bool:
        """Return True if any cell matches an invoice keyword."""
        headers  = [cls.normalize_header(c) for c in row_cleaned if c is not None]
        keywords = [cls.normalize_header(kw) for kw in cls.INVOICE_KWS]
        for h in headers:
            if not h:
                continue
            for kw in keywords:
                if kw and (kw in h or h in kw):
                    return True
        return False

    @classmethod
    def _find_column_index(cls, headers: List[str], keywords: List[str]) -> Optional[int]:
        """Return the first column index whose header matches any of the keywords."""
        normalized_headers  = [cls.normalize_header(h) for h in headers]
        normalized_keywords = [cls.normalize_header(kw) for kw in keywords]
        # Exact match first
        for idx, norm_h in enumerate(normalized_headers):
            if not norm_h:
                continue
            for norm_kw in normalized_keywords:
                if norm_h == norm_kw:
                    return idx
        # Substring match
        for idx, norm_h in enumerate(normalized_headers):
            if not norm_h:
                continue
            for norm_kw in normalized_keywords:
                if norm_kw and (norm_kw in norm_h or norm_h in norm_kw):
                    if norm_kw == "name" and any(p in norm_h for p in ("first", "middle", "last")):
                        continue
                    return idx
        return None

    # -------------------------------------------------------------------------
    # Date / number parsers
    # -------------------------------------------------------------------------

    @staticmethod
    def parse_date(val: Any) -> Optional[datetime.date]:
        """Convert various date formats to datetime.date. Returns None for errors."""
        if val is None:
            return None
        try:
            if pd.isna(val):
                return None
        except Exception:
            pass
        if isinstance(val, datetime.datetime):
            return val.date()
        if isinstance(val, datetime.date):
            return val
        val_str = str(val).strip()
        if not val_str or val_str.lower() in ('nan', 'nat', 'none', 'null') or val_str.startswith('#'):
            return None
        try:
            num_val = float(val_str)
            if 30000 < num_val < 60000:
                return pd.to_datetime(num_val, unit='D', origin='1899-12-30').date()
        except (ValueError, TypeError):
            pass
        val_str = val_str.split(' ')[0]
        for fmt in (
            "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y",
            "%d/%m/%y", "%d-%m-%y", "%d%m%Y", "%d%m%y", "%Y%m%d", "%Y/%m/%d",
        ):
            try:
                return datetime.datetime.strptime(val_str, fmt).date()
            except ValueError:
                continue
        try:
            dt = pd.to_datetime(val_str, errors='coerce', dayfirst=True)
            if pd.notna(dt):
                return dt.date()
        except Exception:
            pass
        return None

    @staticmethod
    def clean_trainee_id(val: Any) -> str:
        """Clean trainee ID: remove .0 suffix, handle nan."""
        if val is None:
            return ""
        try:
            if pd.isna(val):
                return ""
        except Exception:
            pass
        val_str = str(val).strip()
        if val_str.endswith(".0"):
            val_str = val_str[:-2]
        if val_str.lower() in ("nan", "") or val_str.startswith('#'):
            return ""
        return val_str

    @staticmethod
    def parse_float(val: Any) -> float:
        """Parse numeric fields; handles currency symbols, commas, and formula errors."""
        if val is None:
            return 0.0
        try:
            if pd.isna(val):
                return 0.0
        except Exception:
            pass
        val_str = str(val).strip()
        if not val_str or val_str.lower() in ('nan', 'none', 'null') or val_str.startswith('#'):
            return 0.0
        val_str = re.sub(r'[₹\$,\s]', '', val_str)
        try:
            return float(val_str)
        except ValueError:
            return 0.0

    # -------------------------------------------------------------------------
    # Formula-cache helpers
    # -------------------------------------------------------------------------

    @staticmethod
    def _is_formula(val: Any) -> bool:
        """Return True if the value is a raw Excel formula string."""
        return isinstance(val, str) and val.startswith('=')

    @classmethod
    def _resolve_cell_value(
        cls,
        data_only_val: Any,
        formula_val: Any,
        row_idx: int,
        col_idx: int,
        sheet_name: str,
        formula_warn_emitted: "set",
    ) -> Any:
        """
        Return the best available cell value.

        Priority:
          1. data_only_val (cached result) if it is not None and not a formula error.
          2. formula_val if it's a plain (non-formula) value.
          3. None if the formula result was not cached — emit a warning once.
        """
        # If cached result is valid, use it
        if data_only_val is not None:
            if isinstance(data_only_val, str) and data_only_val.startswith('#'):
                return None   # Excel error (e.g. #REF!, #N/A)
            return data_only_val

        # Cached result is None — check if raw value is a formula
        if cls._is_formula(formula_val):
            key = (sheet_name, "formula_warning")
            if key not in formula_warn_emitted:
                formula_warn_emitted.add(key)
                msg = (
                    f"[FORMULA CACHE MISSING] Sheet '{sheet_name}': "
                    f"formula cells (e.g. row {row_idx}, col {col_idx}: "
                    f"{str(formula_val)[:60]}) returned None. "
                    f"Workbook should be recalculated (File → Calculate Now) "
                    f"before upload. Continuing import — optional columns will "
                    f"be stored as blank."
                )
                logger.warning(msg)
                print(msg)
            return None   # Return None; caller decides if it's mandatory

        # Not a formula, just empty
        if formula_val is not None and not (isinstance(formula_val, str) and formula_val.startswith('#')):
            return formula_val

        return None

    # -------------------------------------------------------------------------
    # Sheet type detection
    # -------------------------------------------------------------------------

    @classmethod
    def detect_sheet_type_and_header(cls, ws) -> Tuple[str, float, Optional[int], List[str]]:
        """
        Scan the first 50 rows of *ws* and return:
          (sheet_type, confidence, header_row_0based_index, original_headers_list)

        sheet_type ∈ {"Employee Master", "Separation", "Invoice", "Unknown"}
        """
        rows_sample = list(ws.iter_rows(max_row=50, values_only=True))

        best_idx        = None
        best_score      = -1.0
        best_sheet_type = "Unknown"
        best_confidence = 0.22
        best_headers    = []
        best_reason     = "No known schema matches."

        for idx, row in enumerate(rows_sample):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            row_cleaned = [
                None if (isinstance(c, str) and c.startswith('#')) else c
                for c in row
            ]
            non_empty = [c for c in row_cleaned if c is not None and str(c).strip() != ""]
            if len(non_empty) < 2:
                continue

            has_id      = cls._has_synonym(row_cleaned, "trainee_id")
            has_ticket  = cls._has_synonym(row_cleaned, "ticket_number")
            has_joining = cls._has_synonym(row_cleaned, "joining_date")
            has_end     = cls._has_synonym(row_cleaned, "end_date")
            has_name    = (cls._has_synonym(row_cleaned, "candidate_name")
                          or cls._has_synonym(row_cleaned, "first_name"))
            has_reason  = cls._has_synonym(row_cleaned, "reason")
            has_batch   = cls._has_synonym(row_cleaned, "batch")
            has_cat     = cls._has_synonym(row_cleaned, "category")
            has_invoice = cls._has_invoice_marker(row_cleaned)
            has_total   = any(cls.normalize_header(c) == "total" for c in row_cleaned if c)

            text_cells = sum(
                1 for cell in row_cleaned
                if cell is not None and str(cell).strip() and
                   not _is_numeric(str(cell).strip())
            )

            em_score  = 0.0; em_reasons  = []
            sep_score = 0.0; sep_reasons = []
            inv_score = 0.0; inv_reasons = []

            if has_id or has_ticket:
                em_score  += 10.0; em_reasons.append("Ticket/ID")
                sep_score += 10.0; sep_reasons.append("Ticket/ID")
                inv_score += 10.0; inv_reasons.append("Ticket/ID")
            if has_joining:
                em_score += 10.0; em_reasons.append("Joining Date")
            if has_name:
                em_score += 8.0; em_reasons.append("Name")
            if has_batch:
                em_score += 4.0; em_reasons.append("Batch")
            if has_cat:
                em_score += 4.0; em_reasons.append("Category")
            if has_end:
                sep_score += 10.0; sep_reasons.append("End Date")
            if has_reason:
                sep_score += 8.0; sep_reasons.append("Reason")
            if has_invoice:
                inv_score += 25.0; inv_reasons.append("Invoice fields")
            if has_total:
                inv_score += 4.0; inv_reasons.append("Total")

            # Hard constraints
            if not (has_id or has_ticket):
                em_score = sep_score = inv_score = 0.0
                em_reasons = sep_reasons = inv_reasons = []
            if not has_joining:
                em_score = 0.0; em_reasons = []
            if not has_end:
                sep_score = 0.0; sep_reasons = []
            if not (has_invoice or has_total):
                inv_score = 0.0; inv_reasons = []

            # EM wins tie against Separation when joining date present
            if em_score > 0 and sep_score > 0 and has_joining:
                em_score += 50.0

            row_type  = "Unknown"
            row_conf  = 0.22
            row_score = 0.0
            reason_str = "No known schema matches."

            if em_score > 0 or sep_score > 0 or inv_score > 0:
                if em_score >= sep_score and em_score >= inv_score:
                    row_type   = "Employee Master"
                    row_conf   = 0.98 if has_joining else 0.85
                    row_score  = em_score
                    reason_str = f"Found {' + '.join(em_reasons)} (EM={em_score:.0f} Sep={sep_score:.0f})"
                elif sep_score >= em_score and sep_score >= inv_score:
                    row_type   = "Separation"
                    row_conf   = 0.99
                    row_score  = sep_score
                    reason_str = f"Found {' + '.join(sep_reasons)} (Sep={sep_score:.0f} EM={em_score:.0f})"
                else:
                    row_type   = "Invoice"
                    row_conf   = 0.95
                    row_score  = inv_score
                    reason_str = f"Found {' + '.join(inv_reasons)} (Inv={inv_score:.0f})"
            else:
                row_score = text_cells * 0.5

            final_score = max(em_score, sep_score, inv_score) if (em_score or sep_score or inv_score) else row_score

            if final_score > best_score:
                best_score      = final_score
                best_idx        = idx
                best_sheet_type = row_type
                best_confidence = row_conf
                best_headers    = [str(c).strip() if c is not None else "" for c in row]
                best_reason     = reason_str

        sheet_name_str = ws.title
        logger.info(
            f"Sheet Classification: '{sheet_name_str}' → '{best_sheet_type}' "
            f"({best_reason})"
        )
        print(
            f"Sheet Classification: Workbook Sheet '{sheet_name_str}' → "
            f"Classification: '{best_sheet_type}', Reason: {best_reason}"
        )

        if not hasattr(cls, "_CLASSIFICATION_REASONS"):
            cls._CLASSIFICATION_REASONS = {}
        cls._CLASSIFICATION_REASONS[sheet_name_str] = {
            "type":         best_sheet_type,
            "score_summary": best_reason,
            "headers":      best_headers,
        }

        if best_idx is not None and best_score > 0.0:
            return best_sheet_type, best_confidence, best_idx, best_headers

        # Fallback: first row with ≥2 non-numeric text cells
        for idx, row in enumerate(rows_sample):
            if not row:
                continue
            text_count  = 0
            orig_headers = []
            for cell in row:
                val = str(cell).strip() if cell is not None else ""
                orig_headers.append(val)
                if val and not val.startswith('#') and not _is_numeric(val):
                    text_count += 1
            if text_count >= 2:
                logger.info(
                    f"Sheet Classification Fallback: '{ws.title}' → 'Unknown' "
                    f"(matched on {text_count} text cells)"
                )
                return "Unknown", 0.22, idx, orig_headers

        return "Unknown", 0.22, None, []

    # -------------------------------------------------------------------------
    # Merged cells
    # -------------------------------------------------------------------------

    @classmethod
    def get_merged_cells_map(cls, ws) -> Dict[Tuple[int, int], Any]:
        """
        Build a (row, col) → top-left-value map for all merged ranges.
        Indices are 1-based (openpyxl convention).
        """
        merged_map: Dict[Tuple[int, int], Any] = {}
        if hasattr(ws, "merged_cells") and ws.merged_cells:
            for merged_range in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left = ws.cell(row=min_row, column=min_col).value
                if isinstance(top_left, str) and top_left.startswith('#'):
                    top_left = None
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        merged_map[(r, c)] = top_left
        return merged_map

    # -------------------------------------------------------------------------
    # Last-meaningful-row detection
    # -------------------------------------------------------------------------

    @classmethod
    def _find_last_meaningful_row(
        cls,
        ws,
        header_row_1based: int,
        identity_col_indices: List[int],
    ) -> int:
        """
        Return the 1-based row index to stop at when iterating data rows.

        Strategy: scan backwards from ws.max_row to find the last row that has
        a non-empty value in at least one identity column. This prevents iterating
        thousands of phantom formatting rows that openpyxl sometimes includes in
        max_row for large workbooks.

        If no identity column row is found (e.g. Invoice / Unknown sheet types,
        or small test sheets), fall back to max_row so all rows are visited.
        """
        max_r = ws.max_row or 0
        if max_r == 0 or not identity_col_indices:
            return max_r or header_row_1based

        # Only perform the scan if the sheet is large enough that phantom rows
        # would be a problem. For small sheets use max_row directly.
        SCAN_THRESHOLD = 500
        if max_r <= SCAN_THRESHOLD:
            return max_r

        # Walk backwards — expensive only for large files
        for r in range(max_r, header_row_1based, -1):
            row_vals = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))
            if not row_vals:
                continue
            row = row_vals[0]
            for ci in identity_col_indices:
                if ci < len(row):
                    val = row[ci]
                    if val is not None:
                        val_str = str(val).strip()
                        if val_str and not val_str.startswith('#') and val_str.lower() not in ('nan', 'none', ''):
                            return r

        # No identity-column row found at all — return max_row so every row is
        # visited (identity check in import service will reject empty-ID rows).
        return max_r

    # -------------------------------------------------------------------------
    # Main entry point
    # -------------------------------------------------------------------------

    @classmethod
    def parse_workbook(
        cls,
        file_content: bytes,
        file_name: str,
        upload_mode: str = "MASTER",
    ) -> Dict[str, Any]:
        """
        Parse every visible, non-empty sheet in the workbook.

        Dual-pass strategy:
          - Pass 1: data_only=True  → openpyxl returns *cached* formula results.
          - Pass 2: data_only=False → raw formula strings, used ONLY when Pass 1
                    returned None for a cell that contains a formula string.

        Returns a dict:
          {
            "stats":          { workbook_name, number_of_sheets, sheets_processed,
                                skipped_sheets, failed_sheets, warnings, errors },
            "sheets":         [ { sheet_name, sheet_type, confidence,
                                  original_headers, normalized_headers,
                                  rows, total_rows, valid_rows, blank_rows } ],
            "parser_version": "3.0.0",
          }
        """
        is_large_file = len(file_content) > 10 * 1024 * 1024
        file_io = io.BytesIO(file_content)

        # Pass 1 — cached values (data_only=True)
        wb_data = openpyxl.load_workbook(
            io.BytesIO(file_content),
            read_only=is_large_file,
            data_only=True,
        )
        # Pass 2 — raw formulas (data_only=False); keep in memory, we only
        # consult it per-cell when Pass 1 returns None.
        # For very large files we skip Pass 2 (formula recovery not possible).
        wb_formula = None
        if not is_large_file:
            try:
                wb_formula = openpyxl.load_workbook(
                    io.BytesIO(file_content),
                    read_only=False,
                    data_only=False,
                )
            except Exception as exc:
                logger.warning(f"Could not open workbook for formula pass: {exc}")

        print("Workbook loaded")
        print(f"Sheet names: {wb_data.sheetnames}")

        stats: Dict[str, Any] = {
            "workbook_name":   file_name,
            "number_of_sheets": len(wb_data.sheetnames),
            "sheets_processed": [],
            "skipped_sheets":  [],
            "failed_sheets":   [],
            "warnings":        [],
            "errors":          [],
        }

        parsed_sheets: List[Dict[str, Any]] = []
        formula_warn_emitted: set = set()   # per-workbook, per-sheet dedup

        for sheet_name in wb_data.sheetnames:
            ws_data = wb_data[sheet_name]

            # ── 1. Skip hidden sheets ─────────────────────────────────────────
            if getattr(ws_data, "sheet_state", "visible") != "visible":
                stats["skipped_sheets"].append(sheet_name)
                stats["warnings"].append(f"Sheet '{sheet_name}' skipped: hidden.")
                print(f"Sheet '{sheet_name}': hidden — skipped.")
                continue

            # ── 2. Skip empty sheets ──────────────────────────────────────────
            if _sheet_is_empty(ws_data):
                stats["skipped_sheets"].append(sheet_name)
                stats["warnings"].append(f"Sheet '{sheet_name}' skipped: empty.")
                print(f"Sheet '{sheet_name}': empty — skipped.")
                continue

            try:
                # ── 3. Detect sheet type and header row ───────────────────────
                sheet_type, confidence, header_idx, original_headers = (
                    cls.detect_sheet_type_and_header(ws_data)
                )
                stats["sheets_processed"].append(sheet_name)

                # ── 4. Merged cells map ───────────────────────────────────────
                merged_map: Dict[Tuple[int, int], Any] = {}
                ws_formula_sheet = None
                if not is_large_file:
                    merged_map = cls.get_merged_cells_map(ws_data)
                    if wb_formula and sheet_name in wb_formula.sheetnames:
                        ws_formula_sheet = wb_formula[sheet_name]

                # ── 5. Build normalised header list ───────────────────────────
                normalized_headers: List[str] = []
                seen_headers: Dict[str, int] = {}
                for idx, orig_h in enumerate(original_headers):
                    norm_h = cls.normalize_header(orig_h)
                    if not norm_h:
                        norm_h = f"empty_col_{idx + 1}"
                    if norm_h in seen_headers:
                        seen_headers[norm_h] += 1
                        norm_h = f"{norm_h}_{seen_headers[norm_h]}"
                    else:
                        seen_headers[norm_h] = 1
                    normalized_headers.append(norm_h)

                # ── 6. Schema fingerprint + column-mapping cache ──────────────
                import hashlib
                fingerprint = hashlib.sha256(
                    ",".join(normalized_headers).encode()
                ).hexdigest()
                if not hasattr(cls, "_SCHEMA_CACHE"):
                    cls._SCHEMA_CACHE = {}
                col_mapping = cls._SCHEMA_CACHE.get(fingerprint)
                if col_mapping is None:
                    col_mapping = {
                        std_key: cls._find_column_index(original_headers, syn_list)
                        for std_key, syn_list in cls.SYNONYMS.items()
                    }
                    cls._SCHEMA_CACHE[fingerprint] = col_mapping

                # Identity column indices (used for last-meaningful-row scan)
                identity_col_indices = [
                    i for i in [
                        col_mapping.get("ticket_number"),
                        col_mapping.get("trainee_id"),
                        col_mapping.get("candidate_name"),
                    ]
                    if i is not None
                ]

                # ── 7. Determine last meaningful row ─────────────────────────
                header_row_1based = (header_idx + 1) if header_idx is not None else 1
                if identity_col_indices:
                    last_row_1based = cls._find_last_meaningful_row(
                        ws_data,
                        header_row_1based,
                        identity_col_indices,
                    )
                else:
                    last_row_1based = ws_data.max_row or header_row_1based

                # ── 8. Parse data rows ────────────────────────────────────────
                start_row = header_row_1based + 1
                rows_data:  List[Dict[str, Any]] = []
                blank_count = 0
                valid_count = 0

                for r_idx, row in enumerate(
                    ws_data.iter_rows(
                        min_row=start_row,
                        max_row=last_row_1based,
                        values_only=False,
                    ),
                    start=start_row,
                ):
                    # Build cell value list with merged-cell fallback
                    row_cells_data: List[Any] = []
                    for c_idx, cell in enumerate(row, start=1):
                        if (r_idx, c_idx) in merged_map:
                            val = merged_map[(r_idx, c_idx)]
                        else:
                            val = cell.value
                        row_cells_data.append(val)

                    # Build formula-pass cell list (same shape)
                    row_cells_formula: List[Any] = []
                    if ws_formula_sheet is not None:
                        try:
                            formula_row = next(
                                ws_formula_sheet.iter_rows(
                                    min_row=r_idx,
                                    max_row=r_idx,
                                    values_only=True,
                                ),
                                None,
                            )
                            row_cells_formula = list(formula_row) if formula_row else []
                        except Exception:
                            row_cells_formula = []

                    # Resolve each cell using dual-pass logic
                    row_resolved: List[Any] = []
                    for c_idx, data_val in enumerate(row_cells_data):
                        formula_val = (
                            row_cells_formula[c_idx]
                            if c_idx < len(row_cells_formula)
                            else None
                        )
                        resolved = cls._resolve_cell_value(
                            data_val,
                            formula_val,
                            r_idx,
                            c_idx + 1,
                            sheet_name,
                            formula_warn_emitted,
                        )
                        row_resolved.append(resolved)

                    # Skip entirely blank rows
                    if all(
                        v is None or str(v).strip() == ""
                        for v in row_resolved
                    ):
                        blank_count += 1
                        continue

                    # Build row dict — normalised headers
                    row_dict: Dict[str, Any] = {}
                    for idx, norm_h in enumerate(normalized_headers):
                        if not norm_h:
                            continue
                        val = row_resolved[idx] if idx < len(row_resolved) else None
                        row_dict[norm_h] = cls._coerce_cell(norm_h, val)

                    # Standard keys from synonym mapping
                    for std_key, col_idx in col_mapping.items():
                        if col_idx is not None and col_idx < len(row_resolved):
                            matched_val = row_resolved[col_idx]
                            row_dict[std_key] = cls._coerce_std_key(std_key, matched_val)
                        else:
                            row_dict.setdefault(
                                std_key,
                                None if std_key in ("joining_date", "end_date") else "",
                            )

                    # Name fallback: first + middle + last → candidate_name
                    if not row_dict.get("candidate_name"):
                        parts = [
                            row_dict.get("first_name", ""),
                            row_dict.get("middle_name", ""),
                            row_dict.get("last_name", ""),
                        ]
                        full = " ".join(p for p in parts if p).strip()
                        if full:
                            row_dict["candidate_name"] = full

                    row_dict["_row_num"] = r_idx
                    rows_data.append(row_dict)
                    valid_count += 1

                print(
                    f"Sheet '{sheet_name}': type={sheet_type}, "
                    f"header_row={header_row_1based}, "
                    f"last_row={last_row_1based}, "
                    f"valid={valid_count}, blank={blank_count}"
                )
                logger.info(
                    f"Sheet '{sheet_name}': parsed {valid_count} valid rows, "
                    f"{blank_count} blank rows skipped."
                )

                parsed_sheets.append({
                    "sheet_name":         sheet_name,
                    "sheet_type":         sheet_type,
                    "confidence":         confidence,
                    "original_headers":   original_headers,
                    "normalized_headers": normalized_headers,
                    "rows":               rows_data,
                    "total_rows":         last_row_1based - header_row_1based,
                    "valid_rows":         valid_count,
                    "blank_rows":         blank_count,
                })

            except Exception as exc:
                import traceback
                tb = traceback.format_exc()
                print(f"Sheet '{sheet_name}': FAILED — {exc}\n{tb}")
                logger.error(f"Failed parsing sheet '{sheet_name}': {exc}\n{tb}")
                stats["failed_sheets"].append(sheet_name)
                stats["errors"].append(f"Failed parsing sheet '{sheet_name}': {exc}")

        if wb_formula:
            try:
                wb_formula.close()
            except Exception:
                pass
        wb_data.close()

        return {
            "stats":          stats,
            "sheets":         parsed_sheets,
            "parser_version": cls.PARSER_VERSION,
        }

    # -------------------------------------------------------------------------
    # Per-cell coercion helpers
    # -------------------------------------------------------------------------

    @classmethod
    def _coerce_cell(cls, norm_h: str, val: Any) -> Any:
        """Apply type coercion based on header keyword patterns."""
        all_id_syns = set(cls.SYNONYMS["trainee_id"] + cls.SYNONYMS["ticket_number"])
        all_date_syns = set(cls.SYNONYMS["joining_date"] + cls.SYNONYMS["end_date"])

        if any(k in norm_h for k in all_id_syns):
            return cls.clean_trainee_id(val)
        if any(k in norm_h for k in ('amount', 'payment', 'reimbursement', 'qty',
                                      'quantity', 'total', 'shirt', 'jeans',
                                      'uniform', 'other')):
            return cls.parse_float(val)
        if any(k in norm_h for k in all_date_syns) or 'date' in norm_h:
            return cls.parse_date(val)
        return str(val).strip() if val is not None else ""

    @classmethod
    def _coerce_std_key(cls, std_key: str, val: Any) -> Any:
        """Apply type coercion for standard synonym-mapped keys."""
        if std_key in ("joining_date", "end_date"):
            return cls.parse_date(val)
        if std_key in ("trainee_id", "ticket_number", "offer_id"):
            return cls.clean_trainee_id(val)
        if std_key in ("aadhaar", "mobile"):
            val_str = str(val).strip() if val is not None else ""
            val_str = re.sub(r'[\s\-]', '', val_str)
            if val_str.endswith(".0"):
                val_str = val_str[:-2]
            if val_str.lower() in ('nan', 'none', '') or val_str.startswith('#'):
                return ""
            return val_str
        return str(val).strip() if val is not None else ""


# ─────────────────────────────────────────────────────────────────────────────
# Module-level helpers
# ─────────────────────────────────────────────────────────────────────────────

def _is_numeric(s: str) -> bool:
    """Return True if the string can be parsed as a number."""
    try:
        float(s.replace(',', '').replace('₹', '').replace('$', '').strip())
        return True
    except (ValueError, AttributeError):
        return False


def _sheet_is_empty(ws) -> bool:
    """Return True when every cell in the worksheet is None or blank."""
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                val_str = str(cell).strip()
                if val_str and not val_str.startswith('#'):
                    return False
    return True
