import openpyxl
import datetime
import time
from openpyxl.styles import Font

from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Optional
from sqlalchemy.orm import Session
from sqlalchemy import func, case
from backend.app.models.models import InvoiceRecord, ValidationResult, Trainee, PaymentLedger, AuditLog
from backend.app.repositories.repositories import AuditLogRepository

class ReportService:
    # Reusable style objects to minimize memory overhead per cells
    HEADER_FONT = Font(name="Calibri", size=11, bold=False)

    @staticmethod
    def _calculate_identical_column_width(sheet_name: str, col_idx: int, total_rows: int) -> float:
        """Reproduce the original buggy column width calculation to keep output identical.
        
        Original formula: max_len = max(len(str(val or "")) for val in col)
        where `val` is a Cell object, so str(val) evaluates to:
        "<Cell 'SheetName'.ColumnLetterRowNumber>"
        """
        col_letter = get_column_letter(col_idx)
        # The longest coordinate representation will be at the last row
        last_cell_str = f"<Cell '{sheet_name}'.{col_letter}{total_rows}>"
        max_len = len(last_cell_str)
        return float(max(max_len + 3, 12))

    @classmethod
    def _create_workbook_bytes(cls, sheet_name: str, headers: list, rows: list) -> bytes:
        """Helper to create and save a memory-efficient write-only workbook with autofitted widths."""
        output = BytesIO()
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet(title=sheet_name)

        # Set column widths based on identical coordinate calculation
        total_rows = len(rows) + 1  # 1 for header
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = cls._calculate_identical_column_width(sheet_name, col_idx, total_rows)

        # Write styled header cells using reused header font style
        header_cells = []
        for h in headers:
            cell = WriteOnlyCell(ws, value=h)
            cell.font = cls.HEADER_FONT
            header_cells.append(cell)
        ws.append(header_cells)

        # Append data rows
        for row in rows:
            ws.append(row)

        wb.save(output)
        wb.close()
        return output.getvalue()

    @classmethod
    def generate_approved_invoice_excel(cls, db: Session, invoice_number: str) -> bytes:
        """Generates an Excel workbook containing approved items for a given invoice."""
        start_time = time.time()
        # Query specific columns only with join to avoid loading full ORM objects and N+1 query loop
        records = db.query(
            InvoiceRecord.trainee_id,
            InvoiceRecord.billed_name,
            InvoiceRecord.approved_joining_amount,
            InvoiceRecord.approved_180_days_amount,
            InvoiceRecord.approved_total_amount,
            InvoiceRecord.billed_joining_amount,
            InvoiceRecord.billed_180_days_amount,
            InvoiceRecord.billed_other_amount,
            Trainee.doj,
            Trainee.dol,
            Trainee.scheme
        ).outerjoin(
            Trainee, InvoiceRecord.trainee_id == Trainee.id
        ).filter(
            InvoiceRecord.invoice_number == invoice_number,
            InvoiceRecord.status == "APPROVED"
        ).all()

        rows = []
        for r in records:
            # Only include rows that have actual payouts approved
            if r.approved_joining_amount > 0.0 or r.approved_180_days_amount > 0.0:
                # Resolve DOJ / DOL / Scheme
                doj = r.doj.strftime("%Y-%m-%d") if r.doj else ""
                dol = r.dol.strftime("%Y-%m-%d") if r.dol else ""
                scheme = r.scheme or ""

                rows.append([
                    r.trainee_id or "",
                    r.billed_name or "",
                    scheme,
                    doj,
                    dol,
                    r.billed_joining_amount,
                    r.billed_180_days_amount,
                    r.approved_joining_amount,
                    r.approved_180_days_amount,
                    r.approved_total_amount,
                    r.billed_other_amount
                ])

        headers = [
            "Trainee ID", "Trainee Name", "Scheme", "DOJ", "DOL",
            "Billed Joining (₹)", "Billed 180-Days (₹)", "Approved Joining (₹)",
            "Approved 180-Days (₹)", "Approved Total (₹)", "Billed Other/Ignored (₹)"
        ]

        duration = time.time() - start_time
        # Log to system audit trail
        AuditLogRepository.add_log(
            db=db,
            action="GENERATE_APPROVED_INVOICE_REPORT",
            module="REPORTS",
            details=f"Generated Approved Invoice report for Invoice: {invoice_number}.",
            operator="Admin",
            workbook=f"approved_invoice_{invoice_number}.xlsx",
            duration=duration,
            invoice_number=invoice_number
        )

        return cls._create_workbook_bytes("Approved Payouts", headers, rows)

    @classmethod
    def generate_exception_report_excel(cls, db: Session, invoice_number: Optional[str] = None, format_version: str = "legacy") -> bytes:
        """Generates an Excel workbook listing all validation warnings and errors."""
        start_time = time.time()
        if format_version == "reconciliation":
            query = db.query(InvoiceRecord)
            if invoice_number:
                query = query.filter(InvoiceRecord.invoice_number == invoice_number)
            
            # Filter modified, rejected, or flagged records
            flagged_ids = db.query(ValidationResult.invoice_record_id).filter(
                ValidationResult.status.in_(["WARNING", "ERROR", "FRAUD"])
            ).scalar_subquery()
            
            records = query.filter(
                (InvoiceRecord.status == "EXCEPTION") | 
                (InvoiceRecord.approved_total_amount < InvoiceRecord.billed_total_amount) |
                (InvoiceRecord.id.in_(flagged_ids))
            ).all()

            rows = []
            for record in records:
                ticket = ""
                if record.trainee:
                    ticket = record.trainee.ticket_number or ""
                if not ticket and record.extra_data and isinstance(record.extra_data, dict):
                    ticket = record.extra_data.get("ticket_number") or ""
                    
                name = record.billed_name or (record.trainee.name if record.trainee else "N/A")
                vendor_claim = record.billed_total_amount
                approved_amount = record.approved_total_amount
                rejected_amount = max(0.0, vendor_claim - approved_amount)
                
                # Fetch validation results for reasons
                res = db.query(ValidationResult.message).filter(
                    ValidationResult.invoice_record_id == record.id
                ).all()
                failure_reasons = "; ".join(r[0] for r in res) if res else "N/A"
                
                rows.append([
                    ticket or "N/A",
                    name,
                    vendor_claim,
                    approved_amount,
                    rejected_amount,
                    failure_reasons
                ])
                
            headers = [
                "Ticket", "Name", "Vendor Claim (₹)", "Approved Amount (₹)", "Rejected Amount (₹)", "Failure Reasons"
            ]

            duration = time.time() - start_time
            AuditLogRepository.add_log(
                db=db,
                action="GENERATE_EXCEPTION_REPORT",
                module="REPORTS",
                details=f"Generated Exception report (reconciliation format) for Invoice: {invoice_number or 'All'}.",
                operator="Admin",
                workbook=f"exceptions_{invoice_number}.xlsx" if invoice_number else "exceptions_report_all.xlsx",
                duration=duration,
                invoice_number=invoice_number
            )

            return cls._create_workbook_bytes("Exceptions", headers, rows)

        # Legacy / default formatting
        query = db.query(
            InvoiceRecord.invoice_number,
            ValidationResult.trainee_id,
            Trainee.name.label("trainee_name"),
            ValidationResult.rule_name,
            ValidationResult.status,
            ValidationResult.message,
            ValidationResult.reason_code,
            ValidationResult.recommended_action,
            ValidationResult.created_at
        ).outerjoin(
            InvoiceRecord, ValidationResult.invoice_record_id == InvoiceRecord.id
        ).outerjoin(
            Trainee, ValidationResult.trainee_id == Trainee.id
        ).filter(
            ValidationResult.status.in_(["WARNING", "ERROR"])
        )
        
        if invoice_number:
            query = query.filter(InvoiceRecord.invoice_number == invoice_number)
            
        results = query.all()

        rows = []
        for r in results:
            inv_no = r.invoice_number if r.invoice_number else "N/A"
            trainee_name = r.trainee_name if r.trainee_name else "N/A"
            timestamp = r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else ""
            
            rows.append([
                inv_no,
                r.trainee_id or "N/A",
                trainee_name,
                r.rule_name,
                r.status,
                r.reason_code or "N/A",
                r.message,
                r.recommended_action or "N/A",
                timestamp
            ])

        headers = [
            "Invoice Number", "Trainee ID", "Trainee Name", "Rule Violated",
            "Severity", "Reason Code", "Details", "Recommended Action", "Timestamp"
        ]

        duration = time.time() - start_time
        # Log to system audit trail
        AuditLogRepository.add_log(
            db=db,
            action="GENERATE_EXCEPTION_REPORT",
            module="REPORTS",
            details=f"Generated Exception report for Invoice: {invoice_number or 'All'}.",
            operator="Admin",
            workbook=f"exceptions_{invoice_number}.xlsx" if invoice_number else "exceptions_report_all.xlsx",
            duration=duration,
            invoice_number=invoice_number
        )

        return cls._create_workbook_bytes("Exceptions", headers, rows)

    @classmethod
    def generate_fraud_report_excel(cls, db: Session, invoice_number: Optional[str] = None) -> bytes:
        """Generates an Excel listing of all validation fraud results."""
        start_time = time.time()
        # Query specific columns only with join to avoid N+1 queries
        query = db.query(
            InvoiceRecord.invoice_number,
            ValidationResult.trainee_id,
            Trainee.name.label("trainee_name"),
            ValidationResult.rule_name,
            ValidationResult.status,
            ValidationResult.message,
            ValidationResult.reason_code,
            ValidationResult.recommended_action,
            ValidationResult.created_at
        ).outerjoin(
            InvoiceRecord, ValidationResult.invoice_record_id == InvoiceRecord.id
        ).outerjoin(
            Trainee, ValidationResult.trainee_id == Trainee.id
        ).filter(
            ValidationResult.status.in_(["FRAUD", "ERROR"])
        )

        if invoice_number:
            query = query.filter(InvoiceRecord.invoice_number == invoice_number)

        results = query.all()

        rows = []
        for r in results:
            inv_no = r.invoice_number if r.invoice_number else "N/A"
            trainee_name = r.trainee_name if r.trainee_name else "N/A"
            logged_at = r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else ""

            rows.append([
                inv_no,
                r.trainee_id or "N/A",
                trainee_name,
                r.rule_name,
                "FRAUD / CRITICAL" if r.status == "FRAUD" else r.status,
                r.reason_code or "N/A",
                r.message,
                r.recommended_action or "N/A",
                logged_at
            ])

        headers = [
            "Invoice Number", "Trainee ID", "Trainee Name", "Fraud Rule",
            "Severity", "Reason Code", "Incident Details", "Recommended Action", "Logged At"
        ]

        duration = time.time() - start_time
        # Log to system audit trail
        AuditLogRepository.add_log(
            db=db,
            action="GENERATE_FRAUD_REPORT",
            module="REPORTS",
            details=f"Generated Fraud report{f' for Invoice: {invoice_number}' if invoice_number else ''}.",
            operator="Admin",
            workbook=f"fraud_report_{invoice_number}.xlsx" if invoice_number else "fraud_incidents_report.xlsx",
            duration=duration,
            invoice_number=invoice_number
        )

        return cls._create_workbook_bytes("Fraud Incidents", headers, rows)

    @classmethod
    def generate_payment_summary_excel(cls, db: Session, invoice_number: Optional[str] = None, format_version: str = "legacy") -> bytes:
        """Generates an Excel of all trainees and their complete payments summary."""
        start_time = time.time()
        if format_version == "reconciliation":
            query = db.query(InvoiceRecord)
            if invoice_number:
                query = query.filter(InvoiceRecord.invoice_number == invoice_number)
            records = query.all()
            
            invoice_total = sum(r.billed_total_amount for r in records)
            approved_total = sum(r.approved_total_amount for r in records)
            rejected_total = max(0.0, invoice_total - approved_total)
            money_saved = rejected_total
            
            num_approved = sum(1 for r in records if r.approved_total_amount > 0.0)
            num_rejected = sum(1 for r in records if r.approved_total_amount == 0.0)
            
            rows = [[
                invoice_total,
                approved_total,
                rejected_total,
                money_saved,
                num_approved,
                num_rejected
            ]]
            
            headers = [
                "Invoice Total (₹)",
                "Approved Total (₹)",
                "Rejected Total (₹)",
                "Money Saved (₹)",
                "Number Approved",
                "Number Rejected"
            ]
            
            duration = time.time() - start_time
            AuditLogRepository.add_log(
                db=db,
                action="GENERATE_PAYMENT_SUMMARY_REPORT",
                module="REPORTS",
                details=f"Generated Payment Summary report (reconciliation format) for Invoice: {invoice_number or 'All'}.",
                operator="Admin",
                workbook=f"payment_summary_{invoice_number}.xlsx" if invoice_number else "payment_summary_ledger.xlsx",
                duration=duration,
                invoice_number=invoice_number
            )
            
            return cls._create_workbook_bytes("Payment Summary", headers, rows)

        # Legacy / default formatting
        joining_sub = db.query(
            PaymentLedger.trainee_id,
            func.sum(PaymentLedger.amount_paid).label("joining_paid")
        ).filter(
            PaymentLedger.payment_type == "JOINING"
        ).group_by(PaymentLedger.trainee_id).subquery()

        days180_sub = db.query(
            PaymentLedger.trainee_id,
            func.sum(PaymentLedger.amount_paid).label("days180_paid")
        ).filter(
            PaymentLedger.payment_type == "180_DAYS"
        ).group_by(PaymentLedger.trainee_id).subquery()

        trainees = db.query(
            Trainee.id,
            Trainee.name,
            Trainee.scheme,
            Trainee.doj,
            Trainee.dol,
            Trainee.status,
            Trainee.blocked_reason,
            func.coalesce(joining_sub.c.joining_paid, 0.0).label("joining_paid"),
            func.coalesce(days180_sub.c.days180_paid, 0.0).label("days180_paid")
        ).outerjoin(
            joining_sub, Trainee.id == joining_sub.c.trainee_id
        ).outerjoin(
            days180_sub, Trainee.id == days180_sub.c.trainee_id
        ).all()

        rows = []
        for t in trainees:
            j_paid = t.joining_paid
            d180_paid = t.days180_paid
            total = j_paid + d180_paid
            
            doj = t.doj.strftime("%Y-%m-%d") if t.doj else ""
            dol = t.dol.strftime("%Y-%m-%d") if t.dol else ""

            rows.append([
                t.id,
                t.name,
                t.scheme,
                doj,
                dol,
                t.status,
                t.blocked_reason or "",
                j_paid,
                d180_paid,
                total,
                max(0.0, 1800.0 - total)
            ])

        headers = [
            "Trainee ID", "Name", "Scheme", "DOJ", "DOL", "Status",
            "Blocked Reason", "Joining Paid (₹)", "180-Days Paid (₹)",
            "Total Disbursed (₹)", "Remaining Payout Limit (₹)"
        ]

        duration = time.time() - start_time
        # Log to system audit trail
        AuditLogRepository.add_log(
            db=db,
            action="GENERATE_PAYMENT_SUMMARY_REPORT",
            module="REPORTS",
            details="Generated Payment Summary report.",
            operator="Admin",
            workbook=f"payment_summary_{invoice_number}.xlsx" if invoice_number else "payment_summary_ledger.xlsx",
            duration=duration,
            invoice_number=invoice_number
        )

        return cls._create_workbook_bytes("Payment Ledgers", headers, rows)

    @classmethod
    def generate_audit_report_excel(cls, db: Session) -> bytes:
        """Generates an Excel export of the entire system audit trail."""
        start_time = time.time()
        # Query specific columns only to avoid loading full ORM objects
        logs = db.query(
            AuditLog.timestamp,
            AuditLog.action,
            AuditLog.module,
            AuditLog.details
        ).order_by(AuditLog.timestamp.desc()).all()

        rows = []
        for l in logs:
            timestamp = l.timestamp.strftime("%Y-%m-%d %H:%M:%S") if l.timestamp else ""
            rows.append([
                timestamp,
                l.action,
                l.module,
                l.details
            ])

        headers = ["Timestamp (UTC)", "Action Type", "System Module", "Logged Details"]

        duration = time.time() - start_time
        # Log to system audit trail
        AuditLogRepository.add_log(
            db=db,
            action="GENERATE_AUDIT_REPORT",
            module="REPORTS",
            details="Generated Audit Report.",
            operator="Admin",
            workbook="system_audit_logs.xlsx",
            duration=duration
        )

        return cls._create_workbook_bytes("Audit Trails", headers, rows)

    @classmethod
    def generate_rejected_invoice_excel(cls, db: Session, invoice_number: str) -> bytes:
        """Generates an Excel listing every invoice row that was fully rejected (approved_total == 0)."""
        start_time = time.time()
        # Fetch all rejected records for the invoice with a join to Trainee for scheme/doj/dol
        records = db.query(
            InvoiceRecord.id,
            InvoiceRecord.trainee_id,
            InvoiceRecord.billed_name,
            InvoiceRecord.billed_joining_amount,
            InvoiceRecord.billed_180_days_amount,
            InvoiceRecord.billed_total_amount,
            InvoiceRecord.approved_total_amount,
            InvoiceRecord.status,
            Trainee.scheme,
            Trainee.doj,
            Trainee.dol,
        ).outerjoin(
            Trainee, InvoiceRecord.trainee_id == Trainee.id
        ).filter(
            InvoiceRecord.invoice_number == invoice_number,
            InvoiceRecord.approved_total_amount == 0.0,
            InvoiceRecord.status.in_(["EXCEPTION", "VALIDATED", "REJECTED"])
        ).all()

        # Batch-fetch all validation results for these record IDs to avoid N+1
        record_ids = [r.id for r in records]
        vr_rows = []
        if record_ids:
            vr_rows = db.query(
                ValidationResult.invoice_record_id,
                ValidationResult.rule_name,
                ValidationResult.status,
                ValidationResult.message,
                ValidationResult.reason_code,
            ).filter(
                ValidationResult.invoice_record_id.in_(record_ids)
            ).all()

        # Group validation results by record id
        vr_by_record: dict = {}
        for vr in vr_rows:
            vr_by_record.setdefault(vr.invoice_record_id, []).append(vr)

        rows = []
        for r in records:
            doj = r.doj.strftime("%Y-%m-%d") if r.doj else ""
            dol = r.dol.strftime("%Y-%m-%d") if r.dol else ""
            scheme = r.scheme or ""

            vrs = vr_by_record.get(r.id, [])
            # Collect all unique rejection reasons and reason codes
            reasons = "; ".join(v.message for v in vrs) if vrs else "No validation run"
            reason_codes = "; ".join(v.reason_code for v in vrs if v.reason_code) if vrs else "N/A"
            severity = "FRAUD" if any(v.status == "FRAUD" for v in vrs) else (
                "ERROR" if any(v.status == "ERROR" for v in vrs) else "WARNING"
            ) if vrs else "N/A"

            rows.append([
                r.trainee_id or "N/A",
                r.billed_name or "N/A",
                scheme,
                doj,
                dol,
                r.billed_joining_amount,
                r.billed_180_days_amount,
                r.billed_total_amount,
                r.approved_total_amount,
                severity,
                reason_codes,
                reasons,
            ])

        headers = [
            "Trainee ID", "Billed Name", "Scheme", "DOJ", "DOL",
            "Billed Joining (₹)", "Billed 180-Days (₹)", "Billed Total (₹)",
            "Approved Total (₹)", "Severity", "Reason Code(s)", "Rejection Reason(s)"
        ]

        duration = time.time() - start_time
        AuditLogRepository.add_log(
            db=db,
            action="GENERATE_REJECTED_INVOICE_REPORT",
            module="REPORTS",
            details=f"Generated Rejected Invoice report for Invoice: {invoice_number}.",
            operator="Admin",
            workbook=f"rejected_invoice_{invoice_number}.xlsx",
            duration=duration,
            invoice_number=invoice_number
        )

        return cls._create_workbook_bytes("Rejected Invoice", headers, rows)

    @classmethod
    def generate_corrected_invoice_excel(cls, db: Session, invoice_number: str) -> bytes:
        """Generates a Corrected Invoice Excel sheet comparing original billed amounts against approved amounts."""
        start_time = time.time()
        # Query all records for the given invoice number
        records = db.query(
            InvoiceRecord.id,
            InvoiceRecord.trainee_id,
            InvoiceRecord.billed_name,
            InvoiceRecord.billed_joining_amount,
            InvoiceRecord.billed_180_days_amount,
            InvoiceRecord.billed_other_amount,
            InvoiceRecord.billed_total_amount,
            InvoiceRecord.approved_joining_amount,
            InvoiceRecord.approved_180_days_amount,
            InvoiceRecord.approved_total_amount,
            InvoiceRecord.extra_data
        ).filter(
            InvoiceRecord.invoice_number == invoice_number
        ).all()

        record_ids = [r.id for r in records]
        
        # Batch-load all validation results for this invoice to avoid N+1
        vr_map = {}
        if record_ids:
            vr_rows = db.query(
                ValidationResult.invoice_record_id,
                ValidationResult.message
            ).filter(
                ValidationResult.invoice_record_id.in_(record_ids)
            ).all()
            for v_id, msg in vr_rows:
                vr_map.setdefault(v_id, []).append(msg)

        rows = []
        for r in records:
            ticket = ""
            if r.extra_data and isinstance(r.extra_data, dict):
                ticket = r.extra_data.get("ticket_number") or ""
            
            # Concatenate reasons or default to "Passed / No Correction"
            reasons = vr_map.get(r.id, [])
            reason_str = "; ".join(reasons) if reasons else "Passed / No Correction"

            rejected_val = max(0.0, r.billed_total_amount - r.approved_total_amount)

            rows.append([
                r.trainee_id or "N/A",
                r.billed_name or "N/A",
                ticket or "N/A",
                r.billed_joining_amount,
                r.billed_180_days_amount,
                r.billed_other_amount,
                r.billed_total_amount,
                r.approved_total_amount,
                rejected_val,
                reason_str
            ])

        headers = [
            "Trainee ID", "Billed Name", "Ticket Number", 
            "Billed Joining (₹)", "Billed 180-Days (₹)", "Billed Other/Kit (₹)",
            "Original Value (₹)", "Approved Value (₹)", "Rejected Value (₹)", "Reason"
        ]

        duration = time.time() - start_time
        AuditLogRepository.add_log(
            db=db,
            action="GENERATE_CORRECTED_INVOICE_REPORT",
            module="REPORTS",
            details=f"Generated Corrected Invoice report for Invoice: {invoice_number}.",
            operator="Admin",
            workbook=f"corrected_invoice_{invoice_number}.xlsx",
            duration=duration,
            invoice_number=invoice_number
        )

        return cls._create_workbook_bytes("Corrected Invoice", headers, rows)

    @classmethod
    def generate_vendor_payment_summary_excel(cls, db: Session, invoice_number: str) -> bytes:
        """Generates a detailed per-trainee payment summary for the given invoice."""
        start_time = time.time()
        # Query all records for the given invoice number joined with Trainee details
        records = db.query(
            InvoiceRecord.trainee_id,
            InvoiceRecord.billed_name,
            InvoiceRecord.approved_joining_amount,
            InvoiceRecord.approved_180_days_amount,
            Trainee.scheme,
            Trainee.doj,
            Trainee.dol,
            Trainee.status
        ).outerjoin(
            Trainee, InvoiceRecord.trainee_id == Trainee.id
        ).filter(
            InvoiceRecord.invoice_number == invoice_number
        ).all()

        trainee_ids = [r.trainee_id for r in records if r.trainee_id]

        # Batch-load all payment ledger entries for these trainees to avoid N+1
        ledger_map = {}
        if trainee_ids:
            leds = db.query(
                PaymentLedger.trainee_id,
                PaymentLedger.payment_type,
                PaymentLedger.amount_paid
            ).filter(
                PaymentLedger.trainee_id.in_(trainee_ids)
            ).all()
            for l in leds:
                ledger_map.setdefault(l.trainee_id, []).append(l)

        rows = []
        for r in records:
            t_id = r.trainee_id
            scheme = r.scheme or "N/A"
            doj_str = r.doj.strftime("%Y-%m-%d") if r.doj else "N/A"
            dol_str = r.dol.strftime("%Y-%m-%d") if r.dol else "N/A"
            status = r.status or "PENDING"

            # Calculate historical payments from ledger
            t_ledgers = ledger_map.get(t_id, [])
            hist_joining = sum(l.amount_paid for l in t_ledgers if l.payment_type == "JOINING")
            hist_180 = sum(l.amount_paid for l in t_ledgers if l.payment_type == "180_DAYS")

            cum_disbursed = hist_joining + hist_180
            remaining_limit = max(0.0, 1800.0 - cum_disbursed)

            rows.append([
                t_id or "N/A",
                r.billed_name or "N/A",
                scheme,
                doj_str,
                dol_str,
                status,
                r.approved_joining_amount,
                r.approved_180_days_amount,
                hist_joining,
                hist_180,
                cum_disbursed,
                remaining_limit
            ])

        headers = [
            "Trainee ID", "Trainee Name", "Scheme", "DOJ", "DOL", "Status",
            "Current Approved Joining (₹)", "Current Approved 180-Days (₹)",
            "Historical Joining Paid (₹)", "Historical 180-Days Paid (₹)",
            "Cumulative Total Disbursed (₹)", "Remaining Payout Limit (₹)"
        ]

        duration = time.time() - start_time
        AuditLogRepository.add_log(
            db=db,
            action="GENERATE_VENDOR_PAYMENT_SUMMARY",
            module="REPORTS",
            details=f"Generated Vendor Payment Summary report for Invoice: {invoice_number}.",
            operator="Admin",
            workbook=f"vendor_payment_summary_{invoice_number}.xlsx",
            duration=duration,
            invoice_number=invoice_number
        )

        return cls._create_workbook_bytes("Vendor Payment Summary", headers, rows)

    @classmethod
    def generate_finance_summary_excel(
        cls, db: Session, invoice_number: Optional[str] = None
    ) -> bytes:
        """Generates a Finance Summary Excel with per-invoice aggregated totals and counts."""
        start_time = time.time()
        # Pull all invoice records (filtered or all)
        query = db.query(
            InvoiceRecord.invoice_number,
            InvoiceRecord.invoice_date,
            InvoiceRecord.billed_total_amount,
            InvoiceRecord.approved_total_amount,
            InvoiceRecord.status,
        )
        if invoice_number:
            query = query.filter(InvoiceRecord.invoice_number == invoice_number)
        all_records = query.all()

        if not all_records:
            duration = time.time() - start_time
            AuditLogRepository.add_log(
                db=db,
                action="GENERATE_FINANCE_SUMMARY_REPORT",
                module="REPORTS",
                details=f"Generated Finance Summary report (no data) for Invoice: {invoice_number or 'All'}.",
                operator="Admin",
                workbook=f"finance_summary_{invoice_number}.xlsx" if invoice_number else "finance_summary_all.xlsx",
                duration=duration,
                invoice_number=invoice_number
            )
            return cls._create_workbook_bytes("Finance Summary", [
                "Invoice Number", "Invoice Date",
                "Total Rows", "Billed Total (₹)", "Approved Total (₹)",
                "Rejected Total (₹)", "Pending Total (₹)", "Money Saved (₹)",
                "Approved Count", "Rejected Count", "Exception Count",
                "Fraud Count", "Error Count", "Warning Count"
            ], [])

        # Group by invoice_number in memory
        invoice_groups: dict = {}
        for r in all_records:
            inv = r.invoice_number
            if inv not in invoice_groups:
                invoice_groups[inv] = {
                    "invoice_date": r.invoice_date,
                    "records": []
                }
            invoice_groups[inv]["records"].append(r)

        # Batch-fetch all validation results grouped by invoice_number
        inv_numbers = list(invoice_groups.keys())
        vr_data = db.query(
            InvoiceRecord.invoice_number,
            ValidationResult.status,
        ).join(
            ValidationResult, ValidationResult.invoice_record_id == InvoiceRecord.id
        ).filter(
            InvoiceRecord.invoice_number.in_(inv_numbers)
        ).all()

        # Build per-invoice validation counts
        vr_counts: dict = {}
        for row in vr_data:
            inv = row.invoice_number
            if inv not in vr_counts:
                vr_counts[inv] = {"FRAUD": 0, "ERROR": 0, "WARNING": 0}
            if row.status in vr_counts[inv]:
                vr_counts[inv][row.status] += 1

        rows = []
        for inv_num, grp in sorted(invoice_groups.items()):
            records = grp["records"]
            inv_date = grp["invoice_date"]
            date_str = inv_date.strftime("%Y-%m-%d") if inv_date else ""

            billed_total = sum(r.billed_total_amount for r in records)
            approved_total = sum(r.approved_total_amount for r in records)
            rejected_total = max(0.0, billed_total - approved_total)
            pending_total = sum(
                r.billed_total_amount for r in records if r.status == "PENDING"
            )
            money_saved = rejected_total

            total_rows = len(records)
            approved_count = sum(1 for r in records if r.approved_total_amount > 0.0)
            rejected_count = sum(
                1 for r in records
                if r.approved_total_amount == 0.0 and r.status != "PENDING"
            )
            exception_count = sum(1 for r in records if r.status == "EXCEPTION")

            vc = vr_counts.get(inv_num, {"FRAUD": 0, "ERROR": 0, "WARNING": 0})

            rows.append([
                inv_num,
                date_str,
                total_rows,
                round(billed_total, 2),
                round(approved_total, 2),
                round(rejected_total, 2),
                round(pending_total, 2),
                round(money_saved, 2),
                approved_count,
                rejected_count,
                exception_count,
                vc["FRAUD"],
                vc["ERROR"],
                vc["WARNING"],
            ])

        headers = [
            "Invoice Number", "Invoice Date",
            "Total Rows", "Billed Total (₹)", "Approved Total (₹)",
            "Rejected Total (₹)", "Pending Total (₹)", "Money Saved (₹)",
            "Approved Count", "Rejected Count", "Exception Count",
            "Fraud Count", "Error Count", "Warning Count"
        ]

        AuditLogRepository.add_log(
            db=db,
            action="GENERATE_FINANCE_SUMMARY_REPORT",
            module="REPORTS",
            details=f"Generated Finance Summary report for Invoice: {invoice_number or 'All'}."
        )

        return cls._create_workbook_bytes("Finance Summary", headers, rows)

    @classmethod
    def _create_multisheet_workbook_bytes(cls, sheets_data: list) -> bytes:
        """Helper to create and save a memory-efficient workbook with multiple sheets.
        sheets_data is a list of dicts: [{"title": str, "headers": list, "rows": list}]
        """
        output = BytesIO()
        wb = openpyxl.Workbook(write_only=True)
        for sdata in sheets_data:
            sheet_title = sdata["title"]
            headers = sdata["headers"]
            rows = sdata["rows"]
            ws = wb.create_sheet(title=sheet_title)
            
            # Set column widths based on identical coordinate calculation
            total_rows = len(rows) + 1  # 1 for header
            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = cls._calculate_identical_column_width(sheet_title, col_idx, total_rows)
            
            # Write styled header cells using reused header font style
            header_cells = []
            for h in headers:
                cell = WriteOnlyCell(ws, value=h)
                cell.font = cls.HEADER_FONT
                header_cells.append(cell)
            ws.append(header_cells)
            
            # Append data rows
            for row in rows:
                ws.append(row)
                
        wb.save(output)
        wb.close()
        return output.getvalue()

    @classmethod
    def get_finance_analytics_data(cls, db: Session) -> dict:
        """Retrieves and computes the 11 analytical metrics using optimized database queries."""
        import datetime

        # Define subqueries first to reuse them
        ledger_sub = db.query(
            PaymentLedger.trainee_id,
            func.sum(PaymentLedger.amount_paid).label("paid")
        ).group_by(PaymentLedger.trainee_id).subquery()

        joining_sub = db.query(
            PaymentLedger.trainee_id,
            func.sum(PaymentLedger.amount_paid).label("joining_paid")
        ).filter(PaymentLedger.payment_type == "JOINING").group_by(PaymentLedger.trainee_id).subquery()

        days180_sub = db.query(
            PaymentLedger.trainee_id,
            func.sum(PaymentLedger.amount_paid).label("days180_paid")
        ).filter(PaymentLedger.payment_type == "180_DAYS").group_by(PaymentLedger.trainee_id).subquery()

        # 1. Monthly Spend
        monthly_invoice_data = db.query(
            func.strftime('%Y-%m', InvoiceRecord.invoice_date).label('month'),
            func.sum(InvoiceRecord.billed_total_amount).label('billed'),
            func.sum(InvoiceRecord.approved_total_amount).label('approved')
        ).group_by('month').order_by('month').all()

        monthly_ledger_data = db.query(
            func.strftime('%Y-%m', PaymentLedger.payment_date).label('month'),
            func.sum(PaymentLedger.amount_paid).label('disbursed')
        ).group_by('month').all()

        months_dict = {}
        for m, billed, approved in monthly_invoice_data:
            if not m:
                continue
            months_dict[m] = {
                "month": m,
                "billed": billed or 0.0,
                "approved": approved or 0.0,
                "rejected": max(0.0, (billed or 0.0) - (approved or 0.0)),
                "disbursed": 0.0
            }
        for m, disbursed in monthly_ledger_data:
            if not m:
                continue
            if m not in months_dict:
                months_dict[m] = {"month": m, "billed": 0.0, "approved": 0.0, "rejected": 0.0, "disbursed": 0.0}
            months_dict[m]["disbursed"] = disbursed or 0.0
        
        monthly_spend = sorted(months_dict.values(), key=lambda x: x["month"])

        # 2. Vendor Wise Spend
        vendor_raw = db.query(
            InvoiceRecord.file_name.label('vendor_file'),
            func.sum(InvoiceRecord.billed_total_amount).label('billed'),
            func.sum(InvoiceRecord.approved_total_amount).label('approved')
        ).group_by(InvoiceRecord.file_name).order_by(InvoiceRecord.file_name).all()

        vendor_spend = []
        for file_name, billed, approved in vendor_raw:
            vendor_name = file_name.split('.')[0] if file_name else "Unknown Vendor"
            vendor_spend.append({
                "vendor": vendor_name,
                "file_name": file_name or "N/A",
                "billed": billed or 0.0,
                "approved": approved or 0.0,
                "rejected": max(0.0, (billed or 0.0) - (approved or 0.0))
            })

        # 3. Category Wise Spend
        category_raw = db.query(
            func.coalesce(Trainee.scheme, "Unmatched / Unknown").label('scheme'),
            func.sum(InvoiceRecord.billed_total_amount).label('billed'),
            func.sum(InvoiceRecord.approved_total_amount).label('approved')
        ).outerjoin(
            Trainee, InvoiceRecord.trainee_id == Trainee.id
        ).group_by(Trainee.scheme).order_by(Trainee.scheme).all()

        category_spend = []
        for scheme, billed, approved in category_raw:
            category_spend.append({
                "category": scheme,
                "billed": billed or 0.0,
                "approved": approved or 0.0,
                "rejected": max(0.0, (billed or 0.0) - (approved or 0.0))
            })

        # 4. Joining Payment Summary
        joining_raw = db.query(
            func.count(InvoiceRecord.id).label('total_claims'),
            func.sum(case((InvoiceRecord.approved_joining_amount > 0.0, 1), else_=0)).label('approved_claims'),
            func.sum(InvoiceRecord.billed_joining_amount).label('billed'),
            func.sum(InvoiceRecord.approved_joining_amount).label('approved')
        ).filter(InvoiceRecord.billed_joining_amount > 0.0).first()


        joining_summary = {
            "total_claims": joining_raw.total_claims if joining_raw and joining_raw.total_claims else 0,
            "approved_claims": joining_raw.approved_claims if joining_raw and joining_raw.approved_claims else 0,
            "billed": joining_raw.billed if joining_raw and joining_raw.billed else 0.0,
            "approved": joining_raw.approved if joining_raw and joining_raw.approved else 0.0,
            "rejected": max(0.0, ((joining_raw.billed or 0.0) - (joining_raw.approved or 0.0)) if joining_raw else 0.0)
        }

        # 5. 180 Day Payment Summary
        days180_raw = db.query(
            func.count(InvoiceRecord.id).label('total_claims'),
            func.sum(case((InvoiceRecord.approved_180_days_amount > 0.0, 1), else_=0)).label('approved_claims'),
            func.sum(InvoiceRecord.billed_180_days_amount).label('billed'),
            func.sum(InvoiceRecord.approved_180_days_amount).label('approved')
        ).filter(InvoiceRecord.billed_180_days_amount > 0.0).first()


        days180_summary = {
            "total_claims": days180_raw.total_claims if days180_raw and days180_raw.total_claims else 0,
            "approved_claims": days180_raw.approved_claims if days180_raw and days180_raw.approved_claims else 0,
            "billed": days180_raw.billed if days180_raw and days180_raw.billed else 0.0,
            "approved": days180_raw.approved if days180_raw and days180_raw.approved else 0.0,
            "rejected": max(0.0, ((days180_raw.billed or 0.0) - (days180_raw.approved or 0.0)) if days180_raw else 0.0)
        }

        # 6. Rejected Amount Summary (Violations rule breakdown)
        rejections_raw = db.query(
            ValidationResult.rule_name.label('rule_name'),
            ValidationResult.status.label('status'),
            func.count(ValidationResult.id).label('incidents')
        ).group_by(ValidationResult.rule_name, ValidationResult.status).all()

        rejected_summary = []
        for rule_name, status, incidents in rejections_raw:
            rejected_summary.append({
                "rule_name": rule_name,
                "status": status,
                "incidents": incidents or 0
            })

        # 7. Fraud Savings
        fraud_record_ids = db.query(ValidationResult.invoice_record_id).filter(
            ValidationResult.status == 'FRAUD'
        ).distinct().scalar_subquery()

        fraud_raw = db.query(
            func.sum(InvoiceRecord.billed_total_amount - InvoiceRecord.approved_total_amount).label('savings'),
            func.count(InvoiceRecord.id).label('count')
        ).filter(InvoiceRecord.id.in_(fraud_record_ids)).first()

        fraud_savings = {
            "savings": fraud_raw.savings if fraud_raw and fraud_raw.savings else 0.0,
            "count": fraud_raw.count if fraud_raw and fraud_raw.count else 0
        }

        # 8. Blocked Employee Savings
        blocked_record_ids = db.query(ValidationResult.invoice_record_id).filter(
            (ValidationResult.reason_code == "EMPLOYEE_BLOCKED") | (ValidationResult.rule_name.like("%Blocked%"))
        ).distinct().scalar_subquery()

        blocked_raw = db.query(
            func.sum(InvoiceRecord.billed_total_amount - InvoiceRecord.approved_total_amount).label('savings'),
            func.count(InvoiceRecord.id).label('count')
        ).filter(
            (InvoiceRecord.id.in_(blocked_record_ids)) |
            (InvoiceRecord.trainee_id.in_(db.query(Trainee.id).filter(Trainee.status == "BLOCKED").scalar_subquery()))
        ).first()

        blocked_employee_savings = {
            "savings": blocked_raw.savings if blocked_raw and blocked_raw.savings else 0.0,
            "count": blocked_raw.count if blocked_raw and blocked_raw.count else 0
        }

        # 9. Kit Savings
        kit_other_savings = db.query(
            func.sum(InvoiceRecord.billed_other_amount).label('savings')
        ).scalar() or 0.0

        kit_limit_record_ids = db.query(ValidationResult.invoice_record_id).filter(
            ValidationResult.reason_code.in_(["KIT_EXCESS_THRESHOLD", "KIT_QTY_MISMATCH"])
        ).distinct().scalar_subquery()

        kit_limit_savings = db.query(
            func.sum(InvoiceRecord.billed_total_amount - InvoiceRecord.billed_other_amount - InvoiceRecord.approved_total_amount)
        ).filter(InvoiceRecord.id.in_(kit_limit_record_ids)).scalar() or 0.0

        total_kit_savings = kit_other_savings + max(0.0, kit_limit_savings)

        kit_count = db.query(func.count(InvoiceRecord.id)).filter(
            (InvoiceRecord.billed_other_amount > 0.0) | (InvoiceRecord.id.in_(kit_limit_record_ids))
        ).scalar() or 0

        kit_savings = {
            "savings": total_kit_savings,
            "count": kit_count
        }

        # 10. Remaining Liability
        remaining_liability = db.query(
            func.sum(1800.0 - func.coalesce(ledger_sub.c.paid, 0.0))
        ).select_from(Trainee).outerjoin(
            ledger_sub, Trainee.id == ledger_sub.c.trainee_id
        ).filter(
            Trainee.status == "ACTIVE"
        ).scalar() or 0.0

        # 11. Forecast Timeline
        active_trainees = db.query(
            Trainee.id,
            Trainee.doj,
            func.coalesce(joining_sub.c.joining_paid, 0.0).label("j_paid"),
            func.coalesce(days180_sub.c.days180_paid, 0.0).label("d180_paid")
        ).outerjoin(
            joining_sub, Trainee.id == joining_sub.c.trainee_id
        ).outerjoin(
            days180_sub, Trainee.id == days180_sub.c.trainee_id
        ).filter(
            Trainee.status == "ACTIVE"
        ).all()

        today = datetime.date.today()
        forecast_dict = {}
        for t_id, doj, j_paid, d180_paid in active_trainees:
            if not doj:
                continue
            # Joining forecast
            if j_paid < 1200.0:
                j_amount = 1200.0 - j_paid
                j_date = max(doj, today)
                j_month = j_date.strftime("%Y-%m")
                forecast_dict.setdefault(j_month, {"month": j_month, "joining": 0.0, "days180": 0.0, "total": 0.0})
                forecast_dict[j_month]["joining"] += j_amount
                
            # 180-Day forecast
            if d180_paid < 600.0:
                d180_amount = 600.0 - d180_paid
                eligible_date = doj + datetime.timedelta(days=180)
                d180_date = max(eligible_date, today)
                d180_month = d180_date.strftime("%Y-%m")
                forecast_dict.setdefault(d180_month, {"month": d180_month, "joining": 0.0, "days180": 0.0, "total": 0.0})
                forecast_dict[d180_month]["days180"] += d180_amount

        for m in forecast_dict:
            forecast_dict[m]["total"] = forecast_dict[m]["joining"] + forecast_dict[m]["days180"]

        forecast = sorted(forecast_dict.values(), key=lambda x: x["month"])

        # General summary stats
        total_trainees = db.query(func.count(Trainee.id)).scalar() or 0
        active_trainees_count = db.query(func.count(Trainee.id)).filter(Trainee.status == "ACTIVE").scalar() or 0
        blocked_trainees_count = db.query(func.count(Trainee.id)).filter(Trainee.status == "BLOCKED").scalar() or 0
        separated_trainees_count = db.query(func.count(Trainee.id)).filter(Trainee.status == "SEPARATED").scalar() or 0
        
        overall_billed = db.query(func.sum(InvoiceRecord.billed_total_amount)).scalar() or 0.0
        overall_approved = db.query(func.sum(InvoiceRecord.approved_total_amount)).scalar() or 0.0
        overall_rejected = max(0.0, overall_billed - overall_approved)
        overall_paid = db.query(func.sum(PaymentLedger.amount_paid)).scalar() or 0.0

        return {
            "overall_summary": {
                "total_trainees": total_trainees,
                "active_trainees": active_trainees_count,
                "blocked_trainees": blocked_trainees_count,
                "separated_trainees": separated_trainees_count,
                "total_billed": overall_billed,
                "total_approved": overall_approved,
                "total_rejected": overall_rejected,
                "total_paid": overall_paid
            },
            "monthly_spend": monthly_spend,
            "vendor_spend": vendor_spend,
            "category_spend": category_spend,
            "joining_summary": joining_summary,
            "days180_summary": days180_summary,
            "rejected_summary": rejected_summary,
            "fraud_savings": fraud_savings,
            "blocked_employee_savings": blocked_employee_savings,
            "kit_savings": kit_savings,
            "remaining_liability": remaining_liability,
            "forecast": forecast
        }

    @classmethod
    def generate_finance_analytics_excel(cls, db: Session) -> bytes:
        """Generates a consolidated multi-sheet Excel workbook of all finance analytics."""
        data = cls.get_finance_analytics_data(db)
        
        sheets_data = []
        
        # 1. Executive Summary
        o = data["overall_summary"]
        summary_rows = [
            ["Trainee Lifecycle Stats", ""],
            ["Total Trainees", o["total_trainees"]],
            ["Active Trainees", o["active_trainees"]],
            ["Blocked Trainees", o["blocked_trainees"]],
            ["Separated Trainees", o["separated_trainees"]],
            ["", ""],
            ["Invoice & Disbursed Totals", ""],
            ["Total Billed Payouts (₹)", o["total_billed"]],
            ["Total Approved Payouts (₹)", o["total_approved"]],
            ["Total Rejected Payouts (₹)", o["total_rejected"]],
            ["Total Disbursed Payouts (₹)", o["total_paid"]],
            ["", ""],
            ["Savings & Liability Summary", ""],
            ["Fraud Savings (₹)", data["fraud_savings"]["savings"]],
            ["Blocked Employee Savings (₹)", data["blocked_employee_savings"]["savings"]],
            ["Kit Savings (₹)", data["kit_savings"]["savings"]],
            ["Remaining Active Trainee Liability (₹)", data["remaining_liability"]]
        ]
        sheets_data.append({
            "title": "Executive Summary",
            "headers": ["Metric Name", "Value"],
            "rows": summary_rows
        })
        
        # 2. Monthly Spend
        monthly_rows = []
        for r in data["monthly_spend"]:
            monthly_rows.append([r["month"], r["billed"], r["approved"], r["rejected"], r["disbursed"]])
        sheets_data.append({
            "title": "Monthly Spend Summary",
            "headers": ["Month", "Billed Amount (₹)", "Approved Amount (₹)", "Rejected Amount (₹)", "Disbursed Amount (₹)"],
            "rows": monthly_rows
        })
        
        # 3. Vendor Spend
        vendor_rows = []
        for r in data["vendor_spend"]:
            vendor_rows.append([r["vendor"], r["file_name"], r["billed"], r["approved"], r["rejected"]])
        sheets_data.append({
            "title": "Vendor Spend Summary",
            "headers": ["Vendor / Invoice Source", "Workbook File Name", "Billed Amount (₹)", "Approved Amount (₹)", "Rejected Amount (₹)"],
            "rows": vendor_rows
        })
        
        # 4. Category Spend
        cat_rows = []
        for r in data["category_spend"]:
            cat_rows.append([r["category"], r["billed"], r["approved"], r["rejected"]])
        sheets_data.append({
            "title": "Category Spend Summary",
            "headers": ["Category / Scheme", "Billed Amount (₹)", "Approved Amount (₹)", "Rejected Amount (₹)"],
            "rows": cat_rows
        })
        
        # 5. Joining Payment Summary
        j = data["joining_summary"]
        joining_rows = [
            ["Total Joining Claims Count", j["total_claims"]],
            ["Approved Joining Claims Count", j["approved_claims"]],
            ["Billed Joining Amount (₹)", j["billed"]],
            ["Approved Joining Amount (₹)", j["approved"]],
            ["Rejected Joining Amount (₹)", j["rejected"]]
        ]
        sheets_data.append({
            "title": "Joining Payout Summary",
            "headers": ["Metric Description", "Value"],
            "rows": joining_rows
        })
        
        # 6. 180 Day Payment Summary
        d = data["days180_summary"]
        days180_rows = [
            ["Total 180-Day Claims Count", d["total_claims"]],
            ["Approved 180-Day Claims Count", d["approved_claims"]],
            ["Billed 180-Day Amount (₹)", d["billed"]],
            ["Approved 180-Day Amount (₹)", d["approved"]],
            ["Rejected 180-Day Amount (₹)", d["rejected"]]
        ]
        sheets_data.append({
            "title": "180-Day Payout Summary",
            "headers": ["Metric Description", "Value"],
            "rows": days180_rows
        })
        
        # 7. Rejected Amount Summary (Rules rejections count)
        rej_rows = []
        for r in data["rejected_summary"]:
            rej_rows.append([r["rule_name"], r["status"], r["incidents"]])
        sheets_data.append({
            "title": "Validation Rejections Count",
            "headers": ["Rule Violated", "Severity Level", "Incident Count"],
            "rows": rej_rows
        })
        
        # 8. Forecast Timeline
        forecast_rows = []
        for r in data["forecast"]:
            forecast_rows.append([r["month"], r["joining"], r["days180"], r["total"]])
        sheets_data.append({
            "title": "Liability Forecast Timeline",
            "headers": ["Projected Month", "Projected Joining (₹)", "Projected 180-Day (₹)", "Projected Monthly Liability (₹)"],
            "rows": forecast_rows
        })
        
        AuditLogRepository.add_log(
            db=db,
            action="GENERATE_FINANCE_ANALYTICS_EXCEL",
            module="REPORTS",
            details="Generated consolidated Finance Analytics report (Excel)."
        )
        
        return cls._create_multisheet_workbook_bytes(sheets_data)

    @classmethod
    def generate_finance_analytics_csv(cls, db: Session, metric: str = "summary") -> bytes:
        """Generates a CSV byte stream for a specific finance metric."""
        import csv
        from io import StringIO
        
        data = cls.get_finance_analytics_data(db)
        
        output = StringIO()
        writer = csv.writer(output)
        
        if metric == "monthly_spend":
            headers = ["Month", "Billed Amount (₹)", "Approved Amount (₹)", "Rejected Amount (₹)", "Disbursed Amount (₹)"]
            writer.writerow(headers)
            for row in data["monthly_spend"]:
                writer.writerow([row["month"], row["billed"], row["approved"], row["rejected"], row["disbursed"]])
                
        elif metric == "vendor_spend":
            headers = ["Vendor Name", "Source File Name", "Billed Amount (₹)", "Approved Amount (₹)", "Rejected Amount (₹)"]
            writer.writerow(headers)
            for row in data["vendor_spend"]:
                writer.writerow([row["vendor"], row["file_name"], row["billed"], row["approved"], row["rejected"]])
                
        elif metric == "category_spend":
            headers = ["Category/Scheme", "Billed Amount (₹)", "Approved Amount (₹)", "Rejected Amount (₹)"]
            writer.writerow(headers)
            for row in data["category_spend"]:
                writer.writerow([row["category"], row["billed"], row["approved"], row["rejected"]])
                
        elif metric == "joining_summary":
            headers = ["Metric", "Value"]
            writer.writerow(headers)
            s = data["joining_summary"]
            writer.writerow(["Total Joining Claims Count", s["total_claims"]])
            writer.writerow(["Approved Joining Claims Count", s["approved_claims"]])
            writer.writerow(["Billed Joining Amount (₹)", s["billed"]])
            writer.writerow(["Approved Joining Amount (₹)", s["approved"]])
            writer.writerow(["Rejected Joining Amount (₹)", s["rejected"]])
            
        elif metric == "days180_summary":
            headers = ["Metric", "Value"]
            writer.writerow(headers)
            s = data["days180_summary"]
            writer.writerow(["Total 180-Day Claims Count", s["total_claims"]])
            writer.writerow(["Approved 180-Day Claims Count", s["approved_claims"]])
            writer.writerow(["Billed 180-Day Amount (₹)", s["billed"]])
            writer.writerow(["Approved 180-Day Amount (₹)", s["approved"]])
            writer.writerow(["Rejected 180-Day Amount (₹)", s["rejected"]])
            
        elif metric == "rejected_summary":
            headers = ["Rule Name", "Severity Status", "Incident Count"]
            writer.writerow(headers)
            for row in data["rejected_summary"]:
                writer.writerow([row["rule_name"], row["status"], row["incidents"]])
                
        elif metric == "forecast":
            headers = ["Month", "Projected Joining (₹)", "Projected 180-Day (₹)", "Projected Total (₹)"]
            writer.writerow(headers)
            for row in data["forecast"]:
                writer.writerow([row["month"], row["joining"], row["days180"], row["total"]])
                
        else: # "summary" or default
            headers = ["Finance Metric / Payout Savings", "Count", "Value (₹)"]
            writer.writerow(headers)
            
            # Overall overview
            o = data["overall_summary"]
            writer.writerow(["Total Trainees", o["total_trainees"], ""])
            writer.writerow(["Active Trainees", o["active_trainees"], ""])
            writer.writerow(["Blocked Trainees", o["blocked_trainees"], ""])
            writer.writerow(["Separated Trainees", o["separated_trainees"], ""])
            writer.writerow(["Total Billed Amount", "", o["total_billed"]])
            writer.writerow(["Total Approved Amount", "", o["total_approved"]])
            writer.writerow(["Total Rejected Amount", "", o["total_rejected"]])
            writer.writerow(["Total Disbursed Amount (Ledger)", "", o["total_paid"]])
            
            # Savings & liability
            writer.writerow(["Fraud Savings", data["fraud_savings"]["count"], data["fraud_savings"]["savings"]])
            writer.writerow(["Blocked Employee Savings", data["blocked_employee_savings"]["count"], data["blocked_employee_savings"]["savings"]])
            writer.writerow(["Kit Savings", data["kit_savings"]["count"], data["kit_savings"]["savings"]])
            writer.writerow(["Remaining Liability (Active)", "", data["remaining_liability"]])
            
        AuditLogRepository.add_log(
            db=db,
            action="GENERATE_FINANCE_ANALYTICS_CSV",
            module="REPORTS",
            details=f"Generated Finance Analytics report (CSV, metric: {metric})."
        )
        
        return output.getvalue().encode('utf-8')

    @classmethod
    def generate_finance_analytics_pdf(cls, db: Session) -> bytes:
        """Generates a highly styled, professional PDF finance analytics report."""
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        data = cls.get_finance_analytics_data(db)
        
        output = BytesIO()
        # Create a document template
        doc = SimpleDocTemplate(
            output,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        styles = getSampleStyleSheet()
        
        # Define clean, curated color palette (modern slate/blue/neutral theme)
        primary_color = colors.HexColor("#1A365D")  # Deep Blue
        secondary_color = colors.HexColor("#2B6CB0")  # Vibrant Slate Blue
        text_dark = colors.HexColor("#2D3748")  # Charcoal
        bg_light = colors.HexColor("#F7FAFC")  # Warm white/light gray
        border_color = colors.HexColor("#E2E8F0")  # Slate border
        
        # Modify existing styles to avoid conflicts
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=22,
            textColor=primary_color,
            spaceAfter=8,
            alignment=1  # Centered
        )
        
        subtitle_style = ParagraphStyle(
            'ReportSubtitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=11,
            textColor=secondary_color,
            spaceAfter=25,
            alignment=1  # Centered
        )
        
        h1_style = ParagraphStyle(
            'Heading1_Custom',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=14,
            textColor=primary_color,
            spaceBefore=12,
            spaceAfter=8,
            keepWithNext=True
        )
        
        body_style = ParagraphStyle(
            'Body_Custom',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9.0,
            textColor=text_dark,
            spaceAfter=5
        )
        
        table_text = ParagraphStyle(
            'TableText',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8.0,
            textColor=text_dark
        )
        
        table_header_text = ParagraphStyle(
            'TableHeaderText',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8.0,
            textColor=colors.white
        )

        story = []
        
        # =========================================================================
        # PAGE 1: COVER PAGE & EXECUTIVE SUMMARY
        # =========================================================================
        story.append(Spacer(1, 30))
        story.append(Paragraph("TMPVL Billing Audit & Fraud Detection", title_style))
        story.append(Paragraph("CONSOLIDATED FINANCE ANALYTICS & SAVINGS REPORT", subtitle_style))
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("Executive Financial Summary", h1_style))
        
        # Create a beautiful key-value grid for Executive Summary metrics
        o = data["overall_summary"]
        summary_table_data = [
            [
                Paragraph("<b>Total Trainees:</b>", body_style), Paragraph(str(o["total_trainees"]), body_style),
                Paragraph("<b>Total Billed:</b>", body_style), Paragraph(f"₹{o['total_billed']:,.2f}", body_style)
            ],
            [
                Paragraph("<b>Active Trainees:</b>", body_style), Paragraph(str(o["active_trainees"]), body_style),
                Paragraph("<b>Total Approved:</b>", body_style), Paragraph(f"₹{o['total_approved']:,.2f}", body_style)
            ],
            [
                Paragraph("<b>Blocked Trainees:</b>", body_style), Paragraph(str(o["blocked_trainees"]), body_style),
                Paragraph("<b>Total Rejected:</b>", body_style), Paragraph(f"₹{o['total_rejected']:,.2f}", body_style)
            ],
            [
                Paragraph("<b>Separated Trainees:</b>", body_style), Paragraph(str(o["separated_trainees"]), body_style),
                Paragraph("<b>Total Disbursed:</b>", body_style), Paragraph(f"₹{o['total_paid']:,.2f}", body_style)
            ]
        ]
        
        summary_table = Table(summary_table_data, colWidths=[120, 100, 120, 160])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), bg_light),
            ('BOX', (0,0), (-1,-1), 1, border_color),
            ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING', (0,0), (-1,-1), 10),
            ('RIGHTPADDING', (0,0), (-1,-1), 10),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("Savings & Liability Profile", h1_style))
        
        savings_table_data = [
            [Paragraph("Savings Category", table_header_text), Paragraph("Incidents Count", table_header_text), Paragraph("Total Savings (₹)", table_header_text)],
            [Paragraph("Fraud Payout Savings", table_text), Paragraph(str(data["fraud_savings"]["count"]), table_text), Paragraph(f"₹{data['fraud_savings']['savings']:,.2f}", table_text)],
            [Paragraph("Blocked Employee Payout Savings", table_text), Paragraph(str(data["blocked_employee_savings"]["count"]), table_text), Paragraph(f"₹{data['blocked_employee_savings']['savings']:,.2f}", table_text)],
            [Paragraph("Kit/Uniform Payout Savings", table_text), Paragraph(str(data["kit_savings"]["count"]), table_text), Paragraph(f"₹{data['kit_savings']['savings']:,.2f}", table_text)],
            [Paragraph("<b>Remaining Active Trainee Liability</b>", table_text), Paragraph("-", table_text), Paragraph(f"<b>₹{data['remaining_liability']:,.2f}</b>", table_text)],
        ]
        
        savings_table = Table(savings_table_data, colWidths=[240, 100, 160])
        savings_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), primary_color),
            ('BOX', (0,0), (-1,-1), 1, border_color),
            ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, bg_light]),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING', (0,0), (-1,-1), 10),
        ]))
        story.append(savings_table)
        
        # Add footer or generation timestamp to the first page
        story.append(Spacer(1, 20))
        story.append(Paragraph(f"<i>Report generated on: {datetime.date.today().strftime('%Y-%m-%d')} | Data Source: Central Audit SQLite DB</i>", body_style))
        
        # =========================================================================
        # PAGE 2: MONTHLY SPEND & VENDOR SPEND
        # =========================================================================
        story.append(PageBreak())
        
        story.append(Paragraph("Monthly Spend Trend Summary", h1_style))
        
        monthly_table_headers = [
            Paragraph("Month", table_header_text),
            Paragraph("Billed Total (₹)", table_header_text),
            Paragraph("Approved Total (₹)", table_header_text),
            Paragraph("Rejected Total (₹)", table_header_text),
            Paragraph("Disbursed Total (₹)", table_header_text)
        ]
        monthly_table_data = [monthly_table_headers]
        for m in data["monthly_spend"]:
            monthly_table_data.append([
                Paragraph(m["month"], table_text),
                Paragraph(f"₹{m['billed']:,.2f}", table_text),
                Paragraph(f"₹{m['approved']:,.2f}", table_text),
                Paragraph(f"₹{m['rejected']:,.2f}", table_text),
                Paragraph(f"₹{m['disbursed']:,.2f}", table_text),
            ])
            
        monthly_table = Table(monthly_table_data, colWidths=[80, 105, 105, 105, 105])
        monthly_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), secondary_color),
            ('BOX', (0,0), (-1,-1), 1, border_color),
            ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, bg_light]),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(monthly_table)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("Vendor-Wise Payout Spend", h1_style))
        
        vendor_table_headers = [
            Paragraph("Vendor / Invoice Source", table_header_text),
            Paragraph("Billed Claim (₹)", table_header_text),
            Paragraph("Approved Claim (₹)", table_header_text),
            Paragraph("Rejected Claim (₹)", table_header_text)
        ]
        vendor_table_data = [vendor_table_headers]
        for v in data["vendor_spend"]:
            vendor_table_data.append([
                Paragraph(f"<b>{v['vendor']}</b><br/><font color='#718096' size='7'>{v['file_name']}</font>", table_text),
                Paragraph(f"₹{v['billed']:,.2f}", table_text),
                Paragraph(f"₹{v['approved']:,.2f}", table_text),
                Paragraph(f"₹{v['rejected']:,.2f}", table_text)
            ])
            
        vendor_table = Table(vendor_table_data, colWidths=[200, 100, 100, 100])
        vendor_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), secondary_color),
            ('BOX', (0,0), (-1,-1), 1, border_color),
            ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, bg_light]),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(vendor_table)
        
        # =========================================================================
        # PAGE 3: CATEGORY SPEND & PAYOUT STAGES
        # =========================================================================
        story.append(PageBreak())
        
        story.append(Paragraph("Category (Scheme) Wise Spend", h1_style))
        
        cat_table_headers = [
            Paragraph("Scheme / Course", table_header_text),
            Paragraph("Billed Total (₹)", table_header_text),
            Paragraph("Approved Total (₹)", table_header_text),
            Paragraph("Rejected Total (₹)", table_header_text)
        ]
        cat_table_data = [cat_table_headers]
        for c in data["category_spend"]:
            cat_table_data.append([
                Paragraph(c["category"], table_text),
                Paragraph(f"₹{c['billed']:,.2f}", table_text),
                Paragraph(f"₹{c['approved']:,.2f}", table_text),
                Paragraph(f"₹{c['rejected']:,.2f}", table_text)
            ])
            
        cat_table = Table(cat_table_data, colWidths=[170, 110, 110, 110])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), secondary_color),
            ('BOX', (0,0), (-1,-1), 1, border_color),
            ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, bg_light]),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(cat_table)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("Payout Stage Summaries", h1_style))
        
        j = data["joining_summary"]
        d = data["days180_summary"]
        stage_table_data = [
            [
                Paragraph("Payout Type", table_header_text),
                Paragraph("Total Claims", table_header_text),
                Paragraph("Approved Claims", table_header_text),
                Paragraph("Billed Total (₹)", table_header_text),
                Paragraph("Approved Total (₹)", table_header_text),
                Paragraph("Rejected Total (₹)", table_header_text)
            ],
            [
                Paragraph("<b>Joining Payment</b>", table_text),
                Paragraph(str(j["total_claims"]), table_text),
                Paragraph(str(j["approved_claims"]), table_text),
                Paragraph(f"₹{j['billed']:,.2f}", table_text),
                Paragraph(f"₹{j['approved']:,.2f}", table_text),
                Paragraph(f"₹{j['rejected']:,.2f}", table_text)
            ],
            [
                Paragraph("<b>180-Day Payment</b>", table_text),
                Paragraph(str(d["total_claims"]), table_text),
                Paragraph(str(d["approved_claims"]), table_text),
                Paragraph(f"₹{d['billed']:,.2f}", table_text),
                Paragraph(f"₹{d['approved']:,.2f}", table_text),
                Paragraph(f"₹{d['rejected']:,.2f}", table_text)
            ]
        ]
        
        stage_table = Table(stage_table_data, colWidths=[110, 70, 80, 80, 80, 80])
        stage_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), secondary_color),
            ('BOX', (0,0), (-1,-1), 1, border_color),
            ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, bg_light]),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(stage_table)
        
        # =========================================================================
        # PAGE 4: REJECTIONS BREAKDOWN & FORECAST
        # =========================================================================
        story.append(PageBreak())
        
        story.append(Paragraph("Validation Rule Violation & Rejections Breakdown", h1_style))
        
        rej_table_headers = [
            Paragraph("Validation Rule Name", table_header_text),
            Paragraph("Severity Level", table_header_text),
            Paragraph("Violations Incident Count", table_header_text)
        ]
        rej_table_data = [rej_table_headers]
        for r in data["rejected_summary"]:
            status_text = r["status"]
            if status_text == "FRAUD":
                status_formatted = Paragraph("<font color='#E53E3E'><b>FRAUD</b></font>", table_text)
            elif status_text == "ERROR":
                status_formatted = Paragraph("<font color='#DD6B20'><b>ERROR</b></font>", table_text)
            else:
                status_formatted = Paragraph("<font color='#D69E2E'><b>WARNING</b></font>", table_text)
                
            rej_table_data.append([
                Paragraph(r["rule_name"], table_text),
                status_formatted,
                Paragraph(str(r["incidents"]), table_text)
            ])
            
        if len(rej_table_data) == 1:
            rej_table_data.append([Paragraph("No rejection incidents recorded.", table_text), Paragraph("-", table_text), Paragraph("0", table_text)])
            
        rej_table = Table(rej_table_data, colWidths=[240, 110, 150])
        rej_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), secondary_color),
            ('BOX', (0,0), (-1,-1), 1, border_color),
            ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, bg_light]),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(rej_table)
        story.append(Spacer(1, 10))
        
        story.append(Paragraph("Future Payout Forecast Summary", h1_style))
        story.append(Paragraph("Expected joining and 180-day liability timelines for current active trainees.", body_style))
        story.append(Spacer(1, 5))
        
        forecast_headers = [
            Paragraph("Projected Month", table_header_text),
            Paragraph("Projected Joining (₹)", table_header_text),
            Paragraph("Projected 180-Day (₹)", table_header_text),
            Paragraph("Monthly Projected Liability (₹)", table_header_text)
        ]
        forecast_table_data = [forecast_headers]
        for f in data["forecast"][:12]:  # Limit to 12 months for readable PDF representation
            forecast_table_data.append([
                Paragraph(f["month"], table_text),
                Paragraph(f"₹{f['joining']:,.2f}", table_text),
                Paragraph(f"₹{f['days180']:,.2f}", table_text),
                Paragraph(f"<b>₹{f['total']:,.2f}</b>", table_text)
            ])
            
        if len(forecast_table_data) == 1:
            forecast_table_data.append([Paragraph("No upcoming liability projected.", table_text), Paragraph("-", table_text), Paragraph("-", table_text), Paragraph("₹0.00", table_text)])
            
        forecast_table = Table(forecast_table_data, colWidths=[120, 120, 120, 140])
        forecast_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), secondary_color),
            ('BOX', (0,0), (-1,-1), 1, border_color),
            ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, bg_light]),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(forecast_table)
        
        doc.build(story)
        
        AuditLogRepository.add_log(
            db=db,
            action="GENERATE_FINANCE_ANALYTICS_PDF",
            module="REPORTS",
            details="Generated consolidated Finance Analytics report (PDF)."
        )
        
        return output.getvalue()

