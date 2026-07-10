import re
import os
import datetime
from typing import Dict, List, Tuple, Any, Optional
from sqlalchemy import func
from sqlalchemy.orm import Session
from backend.app.repositories.repositories import TraineeRepository, InvoiceRepository, AuditLogRepository
from backend.app.models.models import Trainee, InvoiceRecord, SeparationRecord, BDCRecord, PaymentLedger, ValidationResult, UploadHistory, TraineeLifecycle, Invoice, InvoiceItem
from backend.app.services.workbook_parser import WorkbookParser
from backend.app.core.json_util import make_json_serializable

class ImportService:
    @staticmethod
    def get_column_value(row: Dict[str, Any], keywords: List[str]) -> Any:
        """Helper to extract normalized value from parsed row using list of alias keywords."""
        for kw in keywords:
            norm_kw = WorkbookParser.normalize_header(kw)
            if norm_kw in row:
                return row[norm_kw]
        return None

    @staticmethod
    def normalize_header(header: str) -> str:
        return WorkbookParser.normalize_header(header)

    @classmethod
    def _serialize_trainee_state(cls, trainee: Optional[Trainee]) -> Dict[str, Any]:
        if not trainee:
            return {}
        return {
            "id": trainee.id,
            "name": trainee.name,
            "doj": trainee.doj.strftime("%Y-%m-%d") if trainee.doj else None,
            "dol": trainee.dol.strftime("%Y-%m-%d") if trainee.dol else None,
            "scheme": trainee.scheme,
            "status": trainee.status,
            "blocked_reason": trainee.blocked_reason,
            "aadhaar": trainee.aadhaar,
            "ticket_number": trainee.ticket_number,
            "category": trainee.category,
            "batch": trainee.batch,
            "shop": trainee.shop,
        }

    @classmethod
    def _parse_sheet_month(cls, sheet_name: str) -> str:
        """Parse a month name or range from sheet name (e.g. 'April 26' -> 'April 2026')"""
        name = sheet_name.strip()
        month_pattern = re.compile(
            r'\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\b[- ]?(\d{2,4})',
            re.IGNORECASE
        )
        matches = list(month_pattern.finditer(name))
        if not matches:
            return name
            
        months_map_full = {
            "jan": "January", "feb": "February", "mar": "March", "apr": "April",
            "may": "May", "jun": "June", "jul": "July", "aug": "August",
            "sep": "September", "oct": "October", "nov": "November", "dec": "December"
        }
        
        def format_match(m):
            mon_raw = m.group(1).lower()[:3]
            mon = months_map_full.get(mon_raw, m.group(1).capitalize())
            yr = m.group(2)
            if len(yr) == 2:
                yr = "20" + yr
            return f"{mon} {yr}"
            
        if len(matches) == 1:
            return format_match(matches[0])
        elif len(matches) >= 2:
            return f"{format_match(matches[0])} to {format_match(matches[-1])}"
        return name

    @classmethod
    def derive_category_and_scheme(cls, sheet_name: str, row_dict: Dict[str, Any], file_name: str = "") -> Tuple[str, str]:
        if not hasattr(cls, "_DERIVE_CACHE"):
            cls._DERIVE_CACHE = {}

        category_syns = ["category", "traineecategory", "empcategory", "employeecategory"]
        program_syns = ["scheme", "program", "course", "type", "schemetype", "programtype", "program/scheme"]
        raw_cat = cls.get_column_value(row_dict, category_syns)
        raw_prog = cls.get_column_value(row_dict, program_syns)

        cache_key = (sheet_name, raw_cat, raw_prog, file_name)
        if cache_key in cls._DERIVE_CACHE:
            return cls._DERIVE_CACHE[cache_key]

        # Helper to normalize category and scheme names
        def normalize_res(val_str: str) -> Optional[Tuple[str, str]]:
            val_clean = str(val_str).strip().upper()
            # Remove separation or master keywords to get clean category name
            val_clean = re.sub(r'[\-_]?(SEPARATION|MASTER|SHEET|BDC|EMPLOYEE|RECORDS|DATA|UPLOAD|SYNC)\b', '', val_clean).strip()
            if not val_clean or re.match(r'^(SHEET|PAGE|BOOK|EXCEL|UNKNOWN|TRAINEE|TRAINEES|EMPLOYEE|EMPLOYEES|MASTER|SEPARATION|SEPARATIONS|BDC|DATA)\d*$', val_clean):
                return None
            
            # Map standard schemes
            if "B.TECH" in val_clean or "BTECH" in val_clean:
                return "B.TECH", "B.Tech"
            elif "M.TECH" in val_clean or "MTECH" in val_clean:
                return "M.TECH", "M.Tech"
            elif "NAPS" in val_clean:
                return "NAPS", "NAPS"
            elif "NATS" in val_clean:
                return "NATS", "NATS"
            elif "ITI" in val_clean:
                return "ITI", "ITI"
            else:
                # Return uppercase category and Title Case scheme
                return val_clean, val_clean.title()

        # Priority 1: Sheet Name
        sheet_res = normalize_res(sheet_name)
        if sheet_res:
            cls._DERIVE_CACHE[cache_key] = sheet_res
            return sheet_res

        # Priority 2: Category Column
        if raw_cat:
            cat_res = normalize_res(str(raw_cat))
            if cat_res:
                cls._DERIVE_CACHE[cache_key] = cat_res
                return cat_res

        # Priority 3: Program Type / Scheme Column
        if raw_prog:
            prog_res = normalize_res(str(raw_prog))
            if prog_res:
                cls._DERIVE_CACHE[cache_key] = prog_res
                return prog_res

        # Priority 4: Workbook metadata (file name)
        if file_name:
            base_name = file_name.rsplit('.', 1)[0]
            file_res = normalize_res(base_name)
            if file_res:
                cls._DERIVE_CACHE[cache_key] = file_res
                return file_res

        # Priority 5: Unknown
        cls._DERIVE_CACHE[cache_key] = ("UNKNOWN", "Unknown")
        return "UNKNOWN", "Unknown"

    @classmethod
    def _process_employee_master_sheet(cls, db: Session, sheet: Dict[str, Any], file_name: str, BATCH_SIZE: int, state: Dict[str, Any]) -> None:
        import time
        import uuid
        import logging
        import re
        logger = logging.getLogger("workbook_parser")
        
        sheet_name = sheet["sheet_name"]
        sheet_inserted = 0
        sheet_updated = 0
        sheet_incomplete_inserted = 0
        sheet_incomplete_updated = 0
        sheet_duplicate_updated = 0
        sheet_identity_conflict = 0   # Missing mandatory identity (Ticket)
        sheet_aadhaar_conflict = 0    # Aadhaar-based fraud / manual review
        sheet_blank = 0
        sheet_failed = 0
        
        sheet_start_time = time.time()
        sheet_before_state = {}
        sheet_after_state = {}

        # 1. Automatic category and scheme detection
        first_row = sheet["rows"][0] if sheet["rows"] else {}
        sheet_category, sheet_scheme = cls.derive_category_and_scheme(sheet_name, first_row, file_name)

        has_offer_id_col = False
        for orig_h in sheet["original_headers"]:
            norm_h = WorkbookParser.normalize_header(orig_h)
            if norm_h in WorkbookParser.SYNONYMS["offer_id"]:
                has_offer_id_col = True
                break
        
        if not has_offer_id_col:
            state["warnings"].append(f"Sheet '{sheet_name}': Missing optional/historical column 'offer_id'.")

        # Check if columns exist in the sheet headers
        has_ticket_col = False
        has_aadhaar_col = False
        for orig_h in sheet["original_headers"]:
            norm_h = WorkbookParser.normalize_header(orig_h)
            if norm_h in WorkbookParser.SYNONYMS["ticket_number"]:
                has_ticket_col = True
            if norm_h in WorkbookParser.SYNONYMS["aadhaar"]:
                has_aadhaar_col = True

        # 1. Preload all trainees and lifecycles once per sheet to avoid N+1 bottlenecks
        t_dup_start = time.time()
        all_trainees = db.query(Trainee).all()
        existing_by_id = {t.id: t for t in all_trainees}
        existing_by_ticket = {t.ticket_number: t for t in all_trainees if t.ticket_number}
        existing_by_aadhaar = {t.aadhaar: t for t in all_trainees if t.aadhaar}

        all_lcs = db.query(TraineeLifecycle.trainee_id, TraineeLifecycle.joining_date, TraineeLifecycle.leaving_date).all()
        existing_lcs = {(lc_id, lc_j, lc_l) for lc_id, lc_j, lc_l in all_lcs}
        dup_resolution_duration = time.time() - t_dup_start
        state.setdefault("dup_resolution_time", 0.0)
        state["dup_resolution_time"] += dup_resolution_duration

        seen_trainee_ids = set()
        seen_tickets = set()
        all_synonyms = set()
        for syns in WorkbookParser.SYNONYMS.values():
            all_synonyms.update(syns)

        def process_batch(rows_to_process):
            nonlocal sheet_inserted, sheet_updated, sheet_incomplete_inserted, sheet_incomplete_updated, sheet_duplicate_updated, sheet_identity_conflict, sheet_aadhaar_conflict, sheet_blank, sheet_failed
            nonlocal existing_by_id, existing_by_ticket, existing_by_aadhaar, existing_lcs
            
            cleaned_rows = []
            
            for row in rows_to_process:
                row_num = row["_row_num"]
                state["rows_processed"] += 1
                
                # Ignore blank rows: if all keys except '_row_num' are blank/None/empty
                is_blank = True
                for k, v in row.items():
                    if k != "_row_num" and v is not None and str(v).strip() != "":
                        is_blank = False
                        break
                if is_blank:
                    sheet_blank += 1
                    continue

                r_id = str(row.get("trainee_id") or "").strip()
                r_ticket = str(row.get("ticket_number") or "").strip()
                r_name = str(row.get("candidate_name") or "").strip()
                r_doj = row.get("joining_date")
                r_offer_id = str(row.get("offer_id") or "").strip()
                r_aadhaar = str(row.get("aadhaar") or "").strip()
                r_aadhaar = r_aadhaar.replace(" ", "").replace("-", "") if r_aadhaar else ""
                r_mobile = str(row.get("mobile") or "").strip()
                r_email = str(row.get("email") or "").strip()
                r_batch = str(row.get("batch") or "").strip()
                r_shop = str(row.get("shop") or "").strip()
                
                # Clean nan strings
                if r_id.lower() == 'nan': r_id = ""
                if r_ticket.lower() == 'nan': r_ticket = ""
                if r_name.lower() == 'nan': r_name = ""
                if r_offer_id.lower() == 'nan': r_offer_id = ""
                if r_aadhaar.lower() == 'nan': r_aadhaar = ""
                if r_mobile.lower() == 'nan': r_mobile = ""
                if r_email.lower() == 'nan': r_email = ""
                if r_batch.lower() == 'nan': r_batch = ""
                if r_shop.lower() == 'nan': r_shop = ""

                # Category and scheme are resolved per row using the fast cached resolver
                r_category, r_scheme = cls.derive_category_and_scheme(sheet_name, row, file_name)

                cleaned_rows.append({
                    "row_num": row_num,
                    "id": r_id,
                    "ticket": r_ticket,
                    "name": r_name,
                    "doj": r_doj,
                    "offer_id": r_offer_id,
                    "aadhaar": r_aadhaar,
                    "mobile": r_mobile,
                    "email": r_email,
                    "batch": r_batch,
                    "shop": r_shop,
                    "category": r_category,
                    "scheme": r_scheme,
                    "raw_row": row
                })

            if not cleaned_rows:
                return

            for item in cleaned_rows:
                row_num = item["row_num"]
                try:
                    r_id = item["id"]
                    r_ticket = item["ticket"]
                    r_name = item["name"]
                    r_doj = item["doj"]
                    r_offer_id = item["offer_id"]
                    r_aadhaar = item["aadhaar"]
                    r_mobile = item["mobile"]
                    r_email = item["email"]
                    r_batch = item["batch"]
                    r_shop = item["shop"]
                    r_category = item["category"]
                    r_scheme = item["scheme"]
                    raw_row = item["raw_row"]

                    # 1. Missing Identity Check — Ticket Number is primary identity.
                    #    Aadhaar alone is not sufficient to reject a row; it is used
                    #    only for lifecycle / fraud checks (Step 3 below).
                    identity_missing = False
                    if has_ticket_col and not r_ticket:
                        # Ticket column exists but this row has no value → reject
                        identity_missing = True
                    elif not has_ticket_col and not r_id:
                        # No ticket col and no trainee_id → reject
                        identity_missing = True

                    if identity_missing:
                        sheet_identity_conflict += 1
                        state["skip_count"] += 1
                        print(f"Sheet : {sheet_name} | Excel Row : {row_num} | Ticket : {r_ticket or 'MISSING'} | Identity Incomplete (no Ticket No.) | Status : IDENTITY_INCOMPLETE")
                        logger.warning(f"Sheet : {sheet_name} | Excel Row : {row_num} | Ticket : {r_ticket or 'MISSING'} | Identity Incomplete (no Ticket No.) | Status : IDENTITY_INCOMPLETE")
                        continue

                    # 2. Identity Resolution — Ticket Number has priority
                    trainee = None
                    if r_ticket in existing_by_ticket:
                        trainee = existing_by_ticket[r_ticket]
                    elif r_id in existing_by_id:
                        trainee = existing_by_id[r_id]

                    # 3. Aadhaar-based lifecycle / fraud detection
                    #
                    #  Case A: Same Aadhaar, same Ticket → normal update (handled by Step 2 already)
                    #  Case B: Same Aadhaar, NEW Ticket, previous lifecycle SEPARATED/BLOCKED
                    #          → this is a RE-HIRE, NOT fraud — allow via rehire logic below
                    #  Case C: Same Aadhaar, NEW Ticket, previous lifecycle ACTIVE
                    #          → FRAUD: Multiple Active Tickets under same Aadhaar
                    #  Case D: Same Aadhaar, different person (Aadhaar already owned by a
                    #          completely different matched trainee) → manual review
                    if r_aadhaar:
                        aadhaar_owner = existing_by_aadhaar.get(r_aadhaar)
                        if aadhaar_owner:
                            if trainee and trainee.id == aadhaar_owner.id:
                                # Case A — same person, same ticket → normal update, no action here
                                pass
                            elif trainee and trainee.id != aadhaar_owner.id:
                                # Case D — Ticket resolves to one person but Aadhaar to another
                                sheet_aadhaar_conflict += 1
                                state["skip_count"] += 1
                                msg = (f"Sheet : {sheet_name} | Excel Row : {row_num} | Ticket : {r_ticket} | "
                                       f"FRAUD/MANUAL-REVIEW: Aadhaar {r_aadhaar} already belongs to a different "
                                       f"employee (ID: {aadhaar_owner.id}, Ticket: {aadhaar_owner.ticket_number})")
                                print(msg)
                                logger.warning(msg)
                                continue
                            elif not trainee:
                                # Aadhaar found but ticket didn't match → check lifecycle
                                prev_status = aadhaar_owner.status
                                if prev_status in ("SEPARATED", "BLOCKED"):
                                    # Case B — re-hire under new ticket; proceed normally
                                    # The rehire logic in Step 5 will archive the old lifecycle
                                    trainee = aadhaar_owner
                                    msg = (f"Sheet : {sheet_name} | Excel Row : {row_num} | Ticket : {r_ticket} | "
                                           f"RE-HIRE detected: Aadhaar {r_aadhaar} was previously {prev_status} "
                                           f"under ticket {aadhaar_owner.ticket_number}. Creating new lifecycle.")
                                    print(msg)
                                    logger.info(msg)
                                elif prev_status == "ACTIVE":
                                    # Case C — same Aadhaar, two different active tickets → fraud
                                    sheet_aadhaar_conflict += 1
                                    state["skip_count"] += 1
                                    msg = (f"Sheet : {sheet_name} | Excel Row : {row_num} | Ticket : {r_ticket} | "
                                           f"FRAUD: Aadhaar {r_aadhaar} is ACTIVE under ticket "
                                           f"{aadhaar_owner.ticket_number}. Cannot create duplicate active record.")
                                    print(msg)
                                    logger.warning(msg)
                                    continue
                                else:
                                    # Unknown status — allow as re-hire with a warning
                                    trainee = aadhaar_owner
                                    msg = (f"Sheet : {sheet_name} | Excel Row : {row_num} | Ticket : {r_ticket} | "
                                           f"WARNING: Aadhaar {r_aadhaar} resolved to existing employee "
                                           f"(status={prev_status}). Treating as update.")
                                    print(msg)
                                    logger.warning(msg)

                    # 4. Duplicate Sheet Detection
                    is_duplicate_updated = False
                    if r_ticket in seen_tickets:
                        is_duplicate_updated = True
                    seen_tickets.add(r_ticket)
                    if r_id:
                        seen_trainee_ids.add(r_id)

                    resolved_id = trainee.id if trainee else (r_id if r_id else r_ticket)
                    
                    if "processed_ids" not in state:
                        state["processed_ids"] = set()
                    state["processed_ids"].add(resolved_id)

                    # 5. Missing Data Quality Checks
                    missing_fields = []
                    if not r_name:
                        r_name = "MISSING"
                        missing_fields.append("Name")
                    if not r_doj:
                        r_doj = None
                        missing_fields.append("DOJ")
                    if not r_offer_id:
                        r_offer_id = "MISSING"
                        missing_fields.append("Offer ID")
                    if not r_mobile:
                        r_mobile = "MISSING"
                        missing_fields.append("Mobile")
                    if not r_email:
                        r_email = "MISSING"
                        missing_fields.append("Email")
                    if not r_shop:
                        r_shop = "MISSING"
                        missing_fields.append("Shop")
                    if not r_category or r_category == "UNKNOWN":
                        missing_fields.append("Category")
                        if not r_category:
                            r_category = "UNKNOWN"

                    validation_status = "INCOMPLETE" if missing_fields else "COMPLETE"
                    validation_metadata = {
                        "validation_status": validation_status,
                        "missing_fields": missing_fields
                    }

                    # Determine if any fields actually changed
                    changed = False
                    if trainee:
                        if (trainee.name != r_name or
                            trainee.doj != r_doj or
                            trainee.category != r_category or
                            trainee.scheme != r_scheme or
                            trainee.batch != (r_batch or None) or
                            trainee.shop != (r_shop or None) or
                            (r_aadhaar and trainee.aadhaar != r_aadhaar) or
                            (r_ticket and trainee.ticket_number != r_ticket) or
                            trainee.offer_id != (r_offer_id or None) or
                            trainee.mobile != (r_mobile or None) or
                            trainee.email != (r_email or None)):
                            changed = True

                    # Determine reconciliation category
                    if is_duplicate_updated:
                        sheet_duplicate_updated += 1
                        state["updated_count"] += 1
                        print("Duplicate Ticket")
                        print("Existing Updated")
                        print("Duplicate Ticket")
                        print("Updated Existing Employee")
                        logger.warning(f"Duplicate Ticket | Existing Updated | Row: {row_num}")
                    else:
                        if trainee:
                            if changed:
                                state["updated_count"] += 1
                                if missing_fields:
                                    sheet_incomplete_updated += 1
                                else:
                                    sheet_updated += 1
                            else:
                                state["skip_count"] += 1
                                if missing_fields:
                                    sheet_incomplete_updated += 1
                                else:
                                    sheet_updated += 1
                        else:
                            state["created_count"] += 1
                            if missing_fields:
                                sheet_incomplete_inserted += 1
                            else:
                                sheet_inserted += 1

                    # Log incomplete rows as WARNING
                    if missing_fields:
                        missing_str = ", ".join(missing_fields)
                        print(f"Sheet : {sheet_name} | Excel Row : {row_num} | Ticket : {r_ticket} | Imported | Status : INCOMPLETE | Missing: {missing_str}")
                        logger.warning(f"Sheet : {sheet_name} | Excel Row : {row_num} | Ticket : {r_ticket} | Imported | Status : INCOMPLETE | Missing: {missing_str}")

                    # Extract unknown columns
                    unknown_data = {}
                    for k, val in raw_row.items():
                        if k not in WorkbookParser.SYNONYMS and k not in all_synonyms and not k.startswith('_'):
                            unknown_data[k] = val

                    if trainee:
                        if trainee.id not in sheet_before_state:
                            sheet_before_state[trainee.id] = cls._serialize_trainee_state(trainee)
                        
                        # Rehire check
                        is_rehire = False
                        if trainee.status in ("SEPARATED", "BLOCKED"):
                            if trainee.dol and r_doj and r_doj > trainee.dol and r_doj != trainee.doj:
                                is_rehire = True

                        if is_rehire:
                            # Archive previous lifecycle to TraineeLifecycle
                            prev_doj = trainee.doj
                            prev_invoices = []
                            prev_ledger = []
                            prev_validations = []
                            reason = "Separated"

                            if prev_doj:
                                invs = db.query(InvoiceRecord).filter(
                                    InvoiceRecord.trainee_id == trainee.id,
                                    InvoiceRecord.invoice_date >= prev_doj
                                ).all()
                                prev_invoices = [
                                    {
                                        "invoice_number": inv.invoice_number,
                                        "invoice_date": inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else None,
                                        "billed_joining_amount": inv.billed_joining_amount,
                                        "billed_180_days_amount": inv.billed_180_days_amount,
                                        "billed_other_amount": inv.billed_other_amount,
                                        "billed_total_amount": inv.billed_total_amount,
                                        "approved_joining_amount": inv.approved_joining_amount,
                                        "approved_180_days_amount": inv.approved_180_days_amount,
                                        "approved_total_amount": inv.approved_total_amount,
                                        "status": inv.status,
                                        "file_name": inv.file_name
                                    }
                                    for inv in invs
                                ]

                                leds = db.query(PaymentLedger).filter(
                                    PaymentLedger.trainee_id == trainee.id,
                                    PaymentLedger.payment_date >= prev_doj
                                ).all()
                                prev_ledger = [
                                    {
                                        "invoice_number": led.invoice_number,
                                        "payment_type": led.payment_type,
                                        "amount_paid": led.amount_paid,
                                        "payment_date": led.payment_date.strftime("%Y-%m-%d") if led.payment_date else None
                                    }
                                    for led in leds
                                ]

                                vals = db.query(ValidationResult).filter(
                                    ValidationResult.trainee_id == trainee.id,
                                    ValidationResult.created_at >= datetime.datetime(prev_doj.year, prev_doj.month, prev_doj.day)
                                ).all()
                                prev_validations = [
                                    {
                                        "rule_name": val.rule_name,
                                        "status": val.status,
                                        "message": val.message,
                                        "reason_code": val.reason_code,
                                        "recommended_action": val.recommended_action
                                    }
                                    for val in vals
                                ]

                                if trainee.status == "BLOCKED":
                                    reason = trainee.blocked_reason or "Blocked"
                                elif trainee.status == "SEPARATED":
                                    latest_sep = db.query(SeparationRecord).filter(
                                        SeparationRecord.trainee_id == trainee.id
                                    ).order_by(SeparationRecord.dol.desc()).first()
                                    if latest_sep:
                                        reason = latest_sep.reason or "Separated"

                            # Guard: only create a lifecycle archive if separation upload
                            # hasn't already written one for the same date range.
                            has_existing_lc = (trainee.id, trainee.doj, trainee.dol) in existing_lcs
                            if not has_existing_lc:
                                old_lc = TraineeLifecycle(
                                    trainee_id=trainee.id,
                                    joining_date=trainee.doj,
                                    leaving_date=trainee.dol,
                                    status=trainee.status,
                                    upload_id=state.get("upload_id"),
                                    extra_data={
                                        "invoice_history": prev_invoices,
                                        "payment_ledger": prev_ledger,
                                        "validation_history": prev_validations,
                                        "reason": reason,
                                        "category": trainee.category,
                                        "batch": trainee.batch,
                                        "shop": trainee.shop,
                                        "bdc_workbook": trainee.current_workbook,
                                        "bdc_sheet": trainee.current_sheet
                                    }
                                )
                                db.add(old_lc)
                                db.flush()
                                existing_lcs.add((trainee.id, trainee.doj, trainee.dol))
                            state["rows_rehired"] += 1

                            trainee.dol = None
                            trainee.status = "ACTIVE"
                            trainee.blocked_reason = None

                        # Update existing fields
                        trainee.name = r_name
                        if r_doj is not None:
                            trainee.doj = r_doj
                        trainee.category = r_category
                        trainee.scheme = r_scheme
                        trainee.batch = r_batch or None
                        trainee.shop = r_shop or None
                        if r_aadhaar:
                            trainee.aadhaar = r_aadhaar
                        if r_ticket:
                            trainee.ticket_number = r_ticket
                        trainee.offer_id = r_offer_id or None
                        trainee.mobile = r_mobile or None
                        trainee.email = r_email or None

                        ext_data = trainee.extra_data or {}
                        ext_data["offer_id"] = r_offer_id
                        ext_data["mobile"] = r_mobile
                        ext_data["email"] = r_email
                        ext_data["validation_metadata"] = validation_metadata
                        for k, val in unknown_data.items():
                            ext_data[k] = val
                        trainee.extra_data = make_json_serializable(ext_data)

                        trainee.current_workbook = file_name
                        trainee.current_sheet = sheet_name

                        if trainee.status not in ("SEPARATED", "BLOCKED"):
                            trainee.status = "ACTIVE"
                        
                        sheet_after_state[trainee.id] = cls._serialize_trainee_state(trainee)
                    else:
                        # Create new Trainee
                        ext_data = {
                            "offer_id": r_offer_id,
                            "mobile": r_mobile,
                            "email": r_email,
                            "validation_metadata": validation_metadata
                        }
                        for k, val in unknown_data.items():
                            ext_data[k] = val

                        new_trainee = Trainee(
                            id=resolved_id,
                            name=r_name,
                            doj=r_doj,
                            scheme=r_scheme,
                            category=r_category,
                            batch=r_batch or None,
                            shop=r_shop or None,
                            aadhaar=r_aadhaar or None,
                            ticket_number=r_ticket or None,
                            status="ACTIVE",
                            current_workbook=file_name,
                            current_sheet=sheet_name,
                            offer_id=r_offer_id,
                            mobile=r_mobile,
                            email=r_email,
                            extra_data=ext_data
                        )
                        db.add(new_trainee)
                        existing_by_id[resolved_id] = new_trainee
                        if r_ticket:
                            existing_by_ticket[r_ticket] = new_trainee
                        if r_aadhaar:
                            existing_by_aadhaar[r_aadhaar] = new_trainee
                        
                        sheet_after_state[resolved_id] = cls._serialize_trainee_state(new_trainee)

                    # Add BDCRecord
                    bdc_rec = BDCRecord(
                        trainee_id=resolved_id,
                        doj=r_doj,
                        scheme=r_scheme,
                        file_name=file_name,
                        extra_data={
                            "sheet_name": sheet_name,
                            "category": r_category,
                            "aadhaar": r_aadhaar,
                            "ticket_number": r_ticket,
                            "batch": r_batch,
                            "shop": r_shop,
                            "offer_id": r_offer_id,
                            "mobile": r_mobile,
                            "email": r_email,
                            **unknown_data
                        }
                    )
                    db.add(bdc_rec)
                    state["success_count"] += 1
                    
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    sheet_failed += 1
                    state["error_count"] += 1
                    state["errors"].append(f"Sheet '{sheet_name}', Row {row_num}: Database operation failed: {str(e)}")

        # End of process_batch

        # Run process_batch in batches of 500, commit and print progress
        BATCH_SIZE = 500
        total_rows = len(sheet["rows"])
        batch_num = 1
        for i in range(0, total_rows, BATCH_SIZE):
            process_batch(sheet["rows"][i:i+BATCH_SIZE])
            print(f"Processing {min(i + BATCH_SIZE, total_rows)} / {total_rows}")
            db.flush()
            db.commit()
            print(f"Commit Batch {batch_num}")
            batch_num += 1

        sheet_duration = time.time() - sheet_start_time
        db_insert_duration = max(0.0, sheet_duration - dup_resolution_duration)
        state.setdefault("db_insert_time", 0.0)
        state["db_insert_time"] += db_insert_duration
        
        parsed_rows = len(sheet["rows"]) + sheet["blank_rows"]
        sheet_incomplete = sheet_incomplete_inserted + sheet_incomplete_updated
        total_accounted = (sheet_inserted + sheet_updated + sheet_incomplete + 
                           sheet_duplicate_updated + sheet_identity_conflict +
                           sheet_aadhaar_conflict +
                           sheet["blank_rows"] + sheet_failed)
        reconcile_status = "PASS" if parsed_rows == total_accounted else "FAIL"

        reconcile_msg = (
            f"\n===================================\n"
            f"Sheet : {sheet_name}\n"
            f"Excel Rows          : {sheet.get('total_rows', parsed_rows)}\n"
            f"Parsed Rows         : {parsed_rows}\n"
            f"Inserted            : {sheet_inserted}\n"
            f"Updated             : {sheet_updated}\n"
            f"Incomplete          : {sheet_incomplete}\n"
            f"Duplicate Updated   : {sheet_duplicate_updated}\n"
            f"Missing Ticket (ID) : {sheet_identity_conflict}\n"
            f"Aadhaar Conflicts   : {sheet_aadhaar_conflict}\n"
            f"Blank Rows          : {sheet['blank_rows']}\n"
            f"Failed Rows         : {sheet_failed}\n"
            f"--------------------------------\n"
            f"TOTAL ACCOUNTED\n"
            f"Expected            : {parsed_rows}\n"
            f"Actual              : {total_accounted}\n"
            f"PASS / FAIL         : {reconcile_status}\n"
            f"===================================\n"
        )
        print(reconcile_msg)
        logger.info(reconcile_msg)

        # Audit logging
        AuditLogRepository.add_log(
            db=db,
            action="IMPORT_BDC_SHEET",
            module="BDC_UPLOAD",
            details=f"Workbook: {file_name} | Sheet: {sheet_name} | Reconcile: {reconcile_status} (Parsed: {parsed_rows}, Accounted: {total_accounted})",
            operator="Admin",
            workbook=file_name,
            sheet=sheet_name,
            rows_count=parsed_rows,
            duration=sheet_duration,
            inserted=sheet_inserted + sheet_incomplete_inserted,
            updated=sheet_updated + sheet_duplicate_updated + sheet_incomplete_updated,
            failed=sheet_failed,
            warnings=sheet_incomplete + sheet_identity_conflict,
            errors=sheet_failed,
            before_state={"trainees": sheet_before_state} if sheet_before_state else None,
            after_state={"trainees": sheet_after_state} if sheet_after_state else None,
            upload_id=state.get("upload_id"),
            commit=False
        )

        state["sheet_summaries"][sheet_name] = {
            "sheet_name": sheet_name,
            "sheet_type": "Employee Master",
            "scheme": sheet_scheme,
            "rows_read": parsed_rows,
            "inserted": sheet_inserted + sheet_incomplete_inserted,
            "updated": sheet_updated + sheet_duplicate_updated + sheet_incomplete_updated,
            "skipped": sheet_blank + sheet_identity_conflict,
            "aadhaar_conflicts": sheet_aadhaar_conflict,
            "warnings": sheet_incomplete + sheet_identity_conflict + sheet_aadhaar_conflict,
            "errors": sheet_failed
        }

    @classmethod
    def _process_separation_sheet(cls, db: Session, sheet: Dict[str, Any], file_name: str, BATCH_SIZE: int, state: Dict[str, Any]) -> None:
        import time
        import uuid
        import logging
        import datetime
        logger = logging.getLogger("workbook_parser")
        
        sheet_name = sheet["sheet_name"]
        sheet_inserted = 0
        sheet_updated = 0
        sheet_incomplete = 0
        sheet_duplicate_updated = 0
        sheet_identity_conflict = 0
        sheet_blank = 0
        sheet_failed = 0
        sheet_separated = 0
        sheet_blocked = 0
        
        sheet_start_time = time.time()
        sheet_before_state = {}
        sheet_after_state = {}

        sheet_month = cls._parse_sheet_month(sheet_name)
        seen_row_trainees = set()

        first_row = sheet["rows"][0] if sheet["rows"] else {}
        _, sheet_scheme = cls.derive_category_and_scheme(sheet_name, first_row, file_name)

        # Preload trainees to avoid DB lookup N+1 bottleneck
        t_dup_start = time.time()
        all_trainees = db.query(Trainee).all()
        existing_trainees = {}
        for t in all_trainees:
            existing_trainees[t.id] = t
            if t.ticket_number:
                existing_trainees[t.ticket_number] = t

        # Preload separation records
        all_seps = db.query(SeparationRecord.trainee_id, SeparationRecord.dol).all()
        existing_seps = {(sep_id, sep_dol) for sep_id, sep_dol in all_seps}
        dup_resolution_duration = time.time() - t_dup_start
        state.setdefault("dup_resolution_time", 0.0)
        state["dup_resolution_time"] += dup_resolution_duration

        def process_batch(rows_to_process):
            nonlocal sheet_inserted, sheet_updated, sheet_incomplete, sheet_duplicate_updated, sheet_identity_conflict, sheet_blank, sheet_failed, sheet_separated, sheet_blocked
            nonlocal existing_trainees, existing_seps
            
            cleaned_rows = []
            
            for row in rows_to_process:
                row_num = row["_row_num"]
                state["rows_processed"] += 1
                
                # Ignore blank rows: if all keys except '_row_num' are blank/None/empty
                is_blank = True
                for k, v in row.items():
                    if k != "_row_num" and v is not None and str(v).strip() != "":
                        is_blank = False
                        break
                if is_blank:
                    sheet_blank += 1
                    continue

                trainee_id = row.get("trainee_id") or row.get("ticket_number") or ""
                trainee_id = str(trainee_id).strip()
                if trainee_id.lower() == 'nan': trainee_id = ""
                
                dol = row.get("end_date")
                reason = str(row.get("reason") or "").strip() or "Separated"
                if reason.lower() == 'nan': reason = "Separated"

                cleaned_rows.append({
                    "row_num": row_num,
                    "id": trainee_id,
                    "dol": dol,
                    "reason": reason
                })

            if not cleaned_rows:
                return

            for item in cleaned_rows:
                row_num = item["row_num"]
                try:
                    trainee_id = item["id"]
                    dol = item["dol"]
                    reason = item["reason"]

                    # 1. Missing Identity Check: DO NOT insert a fake employee if identity is missing
                    if not trainee_id:
                        sheet_identity_conflict += 1
                        state["skip_count"] += 1
                        print(f"Sheet : {sheet_name} | Excel Row : {row_num} | Ticket : MISSING | Identity Incomplete | Status : IDENTITY_INCOMPLETE")
                        logger.warning(f"Sheet : {sheet_name} | Excel Row : {row_num} | Ticket : MISSING | Identity Incomplete | Status : IDENTITY_INCOMPLETE")
                        continue

                    # 2. Duplicate Sheet Ingestion Check
                    if (trainee_id, dol) in seen_row_trainees:
                        sheet_duplicate_updated += 1
                        state["skip_count"] += 1
                        print("Duplicate Ticket")
                        print("Existing Updated")
                        logger.warning(f"Duplicate separation entry within sheet: Row {row_num}")
                        continue
                    seen_row_trainees.add((trainee_id, dol))

                    missing_fields = []
                    if not dol:
                        dol = None
                        missing_fields.append("DOL")

                    # 3. Trainee Resolution / Create if missing
                    trainee = existing_trainees.get(trainee_id)
                    is_new = False
                    if not trainee:
                        # Create trainee as INCOMPLETE/SEPARATED
                        is_new = True
                        validation_metadata = {
                            "validation_status": "INCOMPLETE",
                            "missing_fields": ["Name", "DOJ"] + missing_fields
                        }
                        trainee = Trainee(
                            id=trainee_id,
                            name="MISSING",
                            doj=None,
                            dol=dol,
                            scheme=sheet_scheme,
                            status="SEPARATED",
                            ticket_number=trainee_id if trainee_id.isdigit() else None,
                            extra_data={"validation_metadata": validation_metadata}
                        )
                        db.add(trainee)
                        db.flush()
                        existing_trainees[trainee_id] = trainee
                        existing_trainees[trainee.ticket_number] = trainee
                        sheet_incomplete += 1
                        state["created_count"] += 1
                    else:
                        # Existing trainee: newer separation check
                        if trainee.dol:
                            if dol == trainee.dol or (dol and dol < trainee.dol):
                                sheet_duplicate_updated += 1
                                state["skip_count"] += 1
                                continue

                        if trainee.id not in sheet_before_state:
                            sheet_before_state[trainee.id] = cls._serialize_trainee_state(trainee)

                    # 4. Process separation logic without raising errors for null DOJ
                    days_worked = None
                    status_after = "SEPARATED"
                    blocked_reason = None

                    if trainee.doj and dol:
                        days_worked = (dol - trainee.doj).days
                        if days_worked >= 0 and days_worked < 30:
                            status_after = "BLOCKED"
                            blocked_reason = f"Early Separation - Resigned before 30 days (tenure: {days_worked} days)"
                            reason = "Early Separation"
                            state["early_separations_count"] += 1
                            sheet_blocked += 1
                            state["blocked_count"] += 1
                        else:
                            sheet_separated += 1
                            state["separated_count"] += 1
                    else:
                        sheet_separated += 1
                        state["separated_count"] += 1

                    # Log incomplete rows as WARNING
                    all_missing = missing_fields.copy()
                    if trainee.name == "MISSING":
                        all_missing.append("Name")
                    if trainee.doj is None:
                        all_missing.append("DOJ")
                    
                    if all_missing:
                        missing_str = ", ".join(all_missing)
                        print(f"Sheet : {sheet_name} | Excel Row : {row_num} | Ticket : {trainee.ticket_number or trainee_id} | Imported | Status : INCOMPLETE | Missing: {missing_str}")
                        logger.warning(f"Sheet : {sheet_name} | Excel Row : {row_num} | Ticket : {trainee.ticket_number or trainee_id} | Imported | Status : INCOMPLETE | Missing: {missing_str}")

                    # Update reconciliation stats for first-time sheet updates
                    if not is_new:
                        if all_missing:
                            sheet_incomplete += 1
                        else:
                            sheet_updated += 1

                    # Save separation record
                    has_existing_rec = (trainee.id, dol) in existing_seps

                    if not has_existing_rec:
                        sep_rec = SeparationRecord(
                            trainee_id=trainee.id,
                            dol=dol,
                            reason=reason,
                            file_name=file_name,
                            extra_data={
                                "sheet": sheet_name,
                                "sheet_name": sheet_name,
                                "month": sheet_month,
                                "status_before": trainee.status,
                                "status_after": status_after,
                                "tenure": days_worked,
                                "early_exit": days_worked is not None and days_worked < 30,
                                "days_worked": days_worked,
                                "scheme": trainee.scheme
                            }
                        )
                        db.add(sep_rec)
                        existing_seps.add((trainee.id, dol))

                        # Write a TraineeLifecycle record to preserve history
                        sep_lc = TraineeLifecycle(
                            trainee_id=trainee.id,
                            joining_date=trainee.doj,
                            leaving_date=dol,
                            status=status_after,
                            upload_id=state.get("upload_id"),
                            extra_data={
                                "reason": reason,
                                "days_worked": days_worked,
                                "early_exit": days_worked is not None and days_worked < 30,
                                "sheet": sheet_name,
                                "month": sheet_month,
                                "scheme": trainee.scheme,
                                "category": trainee.category,
                            }
                        )
                        db.add(sep_lc)

                    trainee.dol = dol
                    trainee.status = status_after
                    trainee.blocked_reason = blocked_reason
                    
                    sheet_after_state[trainee.id] = cls._serialize_trainee_state(trainee)
                    state["success_count"] += 1
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    sheet_failed += 1
                    state["error_count"] += 1
                    state["errors"].append(f"Sheet '{sheet_name}', Row {row_num}: Database operation failed: {str(e)}")

        # End of process_batch

        # Run process_batch in batches of 500, commit and print progress
        BATCH_SIZE = 500
        total_rows = len(sheet["rows"])
        batch_num = 1
        for i in range(0, total_rows, BATCH_SIZE):
            process_batch(sheet["rows"][i:i+BATCH_SIZE])
            print(f"Processing {min(i + BATCH_SIZE, total_rows)} / {total_rows}")
            db.flush()
            db.commit()
            print(f"Commit Batch {batch_num}")
            batch_num += 1

        sheet_duration = time.time() - sheet_start_time
        db_insert_duration = max(0.0, sheet_duration - dup_resolution_duration)
        state.setdefault("db_insert_time", 0.0)
        state["db_insert_time"] += db_insert_duration

        parsed_rows = len(sheet["rows"]) + sheet["blank_rows"]
        total_accounted = (sheet_inserted + sheet_updated + sheet_incomplete + 
                           sheet_duplicate_updated + sheet_identity_conflict + 
                           sheet["blank_rows"] + sheet_failed)
        reconcile_status = "PASS" if parsed_rows == total_accounted else "FAIL"

        reconcile_msg = (
            f"\n===================================\n"
            f"Sheet : {sheet_name}\n"
            f"Excel Rows          : {sheet.get('total_rows', parsed_rows)}\n"
            f"Parsed Rows         : {parsed_rows}\n"
            f"Inserted            : {sheet_inserted}\n"
            f"Updated             : {sheet_updated}\n"
            f"Incomplete          : {sheet_incomplete}\n"
            f"Duplicate Updated   : {sheet_duplicate_updated}\n"
            f"Identity Conflicts  : {sheet_identity_conflict}\n"
            f"Blank Rows          : {sheet['blank_rows']}\n"
            f"Failed Rows         : {sheet_failed}\n"
            f"--------------------------------\n"
            f"TOTAL ACCOUNTED\n"
            f"Expected            : {parsed_rows}\n"
            f"Actual              : {total_accounted}\n"
            f"PASS / FAIL         : {reconcile_status}\n"
            f"===================================\n"
        )
        print(reconcile_msg)
        logger.info(reconcile_msg)

        # Audit logging
        AuditLogRepository.add_log(
            db=db,
            action="IMPORT_SEPARATION_SHEET",
            module="SEPARATION_UPLOAD",
            details=f"Workbook: {file_name} | Sheet: {sheet_name} | Reconcile: {reconcile_status} (Parsed: {parsed_rows}, Accounted: {total_accounted})",
            operator="Admin",
            workbook=file_name,
            sheet=sheet_name,
            rows_count=parsed_rows,
            duration=sheet_duration,
            inserted=sheet_inserted + sheet_incomplete,
            updated=sheet_updated + sheet_duplicate_updated,
            failed=sheet_failed,
            warnings=sheet_incomplete + sheet_identity_conflict,
            errors=sheet_failed,
            before_state={"trainees": sheet_before_state} if sheet_before_state else None,
            after_state={"trainees": sheet_after_state} if sheet_after_state else None,
            upload_id=state.get("upload_id"),
            commit=False
        )

        state["sheet_summaries"][sheet_name] = {
            "sheet_name": sheet_name,
            "sheet_type": "Separation",
            "scheme": sheet_scheme,
            "rows_read": parsed_rows,
            "inserted": sheet_inserted,
            "updated": sheet_updated + sheet_duplicate_updated,
            "skipped": sheet_blank + sheet_identity_conflict,
            "warnings": sheet_incomplete + sheet_identity_conflict,
            "errors": sheet_failed
        }

    @classmethod
    def import_bdc_workbook(cls, db: Session, file_content: bytes, file_name: str, upload_mode: str = "INCREMENTAL", operator: str = "Admin") -> Dict[str, Any]:
        """Processes the BDC Master Workbook using Universal WorkbookParser."""
        import time
        import uuid
        import hashlib
        start_time = time.time()
        
        # 1. Compute file hash and check duplicate
        file_hash = hashlib.sha256(file_content).hexdigest()
        file_size = len(file_content)
        
        # Unique UUID for the upload history
        current_upload_id = str(uuid.uuid4())
        
        existing_upload = db.query(UploadHistory).filter(UploadHistory.file_hash == file_hash).first()
        is_duplicate = False
        duplicate_warning = None
        if existing_upload:
            is_duplicate = True
            duplicate_warning = f"Workbook '{file_name}' was already uploaded previously (Upload ID: {existing_upload.upload_id})."
            
        # Create UploadHistory first and flush it to the session
        upload = UploadHistory(
            upload_id=current_upload_id,
            file_name=file_name,
            file_hash=file_hash,
            file_size=file_size,
            upload_type="UNKNOWN",
            uploaded_by=operator,
            processing_time=0.0,
            status="PROCESSING",
            is_duplicate=is_duplicate,
            remarks=duplicate_warning
        )
        db.add(upload)
        db.flush()

        # Count before import
        count_before = db.query(Trainee).count()
        print(f"Print count before import: {count_before}")

        parsed_wb = WorkbookParser.parse_workbook(file_content, file_name, upload_mode=upload_mode)
        wb_stats = parsed_wb["stats"]
        
        warnings = list(wb_stats["warnings"])
        if duplicate_warning:
            warnings.append(duplicate_warning)
        errors = list(wb_stats["errors"])
        
        state = {
            "success_count": 0,
            "skipped_count": 0,
            "created_count": 0,
            "updated_count": 0,
            "skip_count": 0,
            "error_count": 0,
            "duplicate_count": 0,
            "blocked_count": 0,
            "separated_count": 0,
            "early_separations_count": 0,
            "unknown_employees_count": 0,
            "rows_processed": 0,
            "rows_no_change": 0,
            "rows_rehired": 0,
            "rows_inactive": 0,
            "warnings": warnings,
            "errors": errors,
            "upload_id": current_upload_id,
            "processed_ids": set(),
            "sheet_summaries": {}
        }

        # Resolve upload_mode scheme scope
        upload_mode_upper = upload_mode.upper()
        is_legacy = upload_mode_upper in ("INCREMENTAL", "FULL_SYNC")
        expected_scheme = None
        if not is_legacy and upload_mode_upper != "MASTER":
            if upload_mode_upper == "BTECH":
                expected_scheme = "B.Tech"
            elif upload_mode_upper == "MTECH":
                expected_scheme = "M.Tech"
            else:
                expected_scheme = upload_mode.title() if len(upload_mode) > 4 else upload_mode.upper()

        sheets_processed = []
        employee_sheets = []
        separation_sheets = []
        skipped_sheets = list(wb_stats["skipped_sheets"])
        failed_sheets = list(wb_stats["failed_sheets"])
        
        BATCH_SIZE = 2000
        
        for sheet in parsed_wb["sheets"]:
            sheet_name = sheet["sheet_name"]
            sheet_type = sheet["sheet_type"]
            
            # Derive sheet scheme
            first_row = sheet["rows"][0] if sheet["rows"] else {}
            _, sheet_scheme = cls.derive_category_and_scheme(sheet_name, first_row, file_name)
            
            # If individual scheme mode selected, validate & warn
            if expected_scheme and sheet_type in ("Employee Master", "BDC", "Separation"):
                if sheet_scheme.upper().replace(".", "") != expected_scheme.upper().replace(".", ""):
                    skipped_sheets.append(sheet_name)
                    state["warnings"].append(
                        f"Workbook contains sheet '{sheet_name}' which appears to belong to scheme '{sheet_scheme}', "
                        f"but '{expected_scheme}' was selected. This sheet was skipped."
                    )
                    continue
            
            if sheet_type in ("Employee Master", "BDC"):
                sheets_processed.append(sheet_name)
                employee_sheets.append(sheet_name)
                cls._process_employee_master_sheet(db, sheet, file_name, BATCH_SIZE, state)
            elif sheet_type == "Separation":
                sheets_processed.append(sheet_name)
                separation_sheets.append(sheet_name)
                cls._process_separation_sheet(db, sheet, file_name, BATCH_SIZE, state)
            else:
                skipped_sheets.append(sheet_name)
                state["warnings"].append(f"Sheet '{sheet_name}' ignored. Classification: '{sheet_type}'. Reason: Sheet type is not supported for BDC Master/Separation sync.")

        # Full Sync Mode: Mark missing ACTIVE employees as INACTIVE
        if upload_mode == "FULL_SYNC":
            if employee_sheets and not failed_sheets:
                # Find currently ACTIVE trainees not present in processed_ids
                missing_trainees = db.query(Trainee).filter(
                    Trainee.status == "ACTIVE",
                    ~Trainee.id.in_(state["processed_ids"])
                ).all()
                for mt in missing_trainees:
                    before_state = cls._serialize_trainee_state(mt)
                    mt.status = "INACTIVE"
                    mt.blocked_reason = "Missing from Latest Master"
                    state["rows_inactive"] += 1
                    
                    AuditLogRepository.add_log(
                        db=db,
                        action="EMPLOYEE_DEACTIVATED",
                        module="BDC_UPLOAD",
                        details=f"Employee '{mt.name}' ({mt.id}) deactivated: Missing from Latest Master.",
                        operator=operator,
                        workbook=file_name,
                        employee_id=mt.id,
                        before_state=before_state,
                        after_state=cls._serialize_trainee_state(mt),
                        upload_id=current_upload_id,
                        commit=False
                    )
                db.flush()
            else:
                state["warnings"].append("Full Sync mode requested, but Employee Master sheets were empty or workbook processing encountered errors. Deactivations skipped.")

        if not sheets_processed:
            raise ValueError("No sheets with required columns found.")
            
        total_duration = time.time() - start_time
        
        # Classify upload type
        if employee_sheets and separation_sheets:
            upload_type = "MIXED"
        elif employee_sheets:
            upload_type = "FULL_SYNC" if upload_mode == "FULL_SYNC" else "INCREMENTAL"
        elif separation_sheets:
            upload_type = "SEPARATION"
        else:
            upload_type = "UNKNOWN"
            
        upload_status = "SUCCESS"
        if failed_sheets:
            upload_status = "PARTIAL" if sheets_processed else "FAILED"
            
        # Update existing UploadHistory summary fields in database
        upload.upload_type = upload_type
        upload.processing_time = total_duration
        upload.status = upload_status
        upload.workbook_version = wb_stats.get("workbook_version")
        upload.parser_version = wb_stats.get("parser_version", "2.0.0")
        upload.sheet_count = wb_stats.get("number_of_sheets", 0)
        upload.visible_sheet_count = len(sheets_processed)
        upload.hidden_sheet_count = len(skipped_sheets)
        upload.rows_processed = state.get("rows_processed", 0)
        upload.rows_inserted = state.get("created_count", 0)
        upload.rows_updated = state.get("updated_count", 0)
        upload.rows_no_change = state.get("rows_no_change", 0)
        upload.rows_skipped = state.get("skip_count", 0)
        upload.rows_failed = state.get("error_count", 0)
        upload.rows_rehired = state.get("rows_rehired", 0)
        upload.rows_inactive = state.get("rows_inactive", 0)
        upload.employee_sheets = employee_sheets
        upload.separation_sheets = separation_sheets
        upload.invoice_sheets = []
        upload.remarks = duplicate_warning
        db.flush()
        
        AuditLogRepository.add_log(
            db=db,
            action="IMPORT_BDC_WORKBOOK",
            module="BDC_UPLOAD",
            details=f"File: {file_name}. Sheets processed: {', '.join(sheets_processed)}. Imported: {state['created_count']}, Updated: {state['updated_count']}, Inactivated: {state['rows_inactive']}, Errors: {state['error_count']}. Duration: {total_duration:.2f}s",
            operator=operator,
            workbook=file_name,
            sheet=", ".join(sheets_processed),
            rows_count=state["rows_processed"],
            duration=total_duration,
            inserted=state["created_count"],
            updated=state["updated_count"],
            failed=state["error_count"],
            warnings=len(state["warnings"]),
            errors=state["error_count"],
            upload_id=current_upload_id,
            commit=False
        )
        
        print(f"rows received: {state['rows_processed']}")
        print(f"rows inserted: {state['created_count']}")
        print(f"rows updated: {state['updated_count']}")
        print(f"rows skipped: {state['skip_count']}")
        print(f"rows failed: {state['error_count']}")

        # Timing report setup
        parser_timings = parsed_wb.get("timings", {})
        wb_load_time = parser_timings.get("workbook_load", 0.0)
        hdr_det_time = parser_timings.get("header_detection", 0.0)
        row_parse_time = parser_timings.get("row_parsing", 0.0)
        dup_res_time = state.get("dup_resolution_time", 0.0)
        db_ins_time = state.get("db_insert_time", 0.0)

        print("commit started")
        t_commit_start = time.time()
        db.commit()
        commit_time = time.time() - t_commit_start
        print("commit completed")

        # Measure dashboard refresh timing
        t_dash_start = time.time()
        db.query(Trainee.status, func.count(Trainee.id)).group_by(Trainee.status).all()
        dash_refresh_time = time.time() - t_dash_start

        total_time = wb_load_time + hdr_det_time + row_parse_time + dup_res_time + db_ins_time + commit_time + dash_refresh_time

        # Print timing report exactly as specified
        print("Workbook Load")
        print(f"{wb_load_time:.1f} sec")
        print("Header Detection")
        print(f"{hdr_det_time:.1f} sec")
        print("Row Parsing")
        print(f"{row_parse_time:.1f} sec")
        print("Duplicate Resolution")
        print(f"{dup_res_time:.1f} sec")
        print("Database Insert")
        print(f"{db_ins_time:.1f} sec")
        print("Commit")
        print(f"{commit_time:.1f} sec")
        print("Dashboard Refresh")
        print(f"{dash_refresh_time:.1f} sec")
        print("Total")
        print(f"{total_time:.1f} sec")

        # Print count after import
        count_after = db.query(Trainee).count()
        state["skipped_count"] = state["skip_count"]
        return {
            "success": True,
            "processed": state["rows_processed"],
            "created": state["created_count"],
            "updated": state["updated_count"] + state["separated_count"] + state["blocked_count"],
            "failed": state["error_count"],
            "message": f"Successfully processed {state['rows_processed']} rows.",
            "workbook_name": file_name,
            "number_of_sheets": len(parsed_wb["sheets"]) + len(skipped_sheets),
            "processed_sheets": sheets_processed,
            "skipped_sheets": skipped_sheets,
            "failed_sheets": failed_sheets,
            "inserted_records": state["created_count"],
            "updated_records": state["updated_count"] + state["separated_count"] + state["blocked_count"],
            "skipped_records": state["skip_count"],
            "failed_records": state["error_count"],
            "warnings": state["warnings"],
            "errors": state["errors"],
            "upload_id": current_upload_id,
            "is_duplicate": is_duplicate,
            "sheet_summaries": list(state.get("sheet_summaries", {}).values()),
            
            # Spec compliance keys
            "employee_sheets": employee_sheets,
            "separation_sheets": separation_sheets,
            "rows_processed": state["rows_processed"],
            "rows_imported": state["created_count"],
            "rows_updated": state["updated_count"] + state["separated_count"] + state["blocked_count"],
            "rows_skipped": state["skip_count"],
            "rows_failed": state["error_count"],
            "rows_no_change": state["rows_no_change"],
            "rows_rehired": state["rows_rehired"],
            "rows_inactive": state["rows_inactive"],
            "processing_time": total_duration,
            
            # Legacy compatibility keys
            "sheets_processed": sheets_processed,
            "employees_updated": state["updated_count"] + state["separated_count"] + state["blocked_count"],
            "early_separations": state["early_separations_count"],
            "blocked_employees": state["blocked_count"],
            "inserted": state["created_count"],
            "updated": state["updated_count"] + state["separated_count"] + state["blocked_count"],
            "blocked": state["blocked_count"],
            "separated": state["separated_count"],
            "duplicates": state["duplicate_count"],
            "unknown_employees": state["unknown_employees_count"],
            
            # original keys
            "success_count": state["success_count"],
            "skipped_count": state["skipped_count"],
            "total_records": state["success_count"] + state["skipped_count"],
            "created_count": state["created_count"],
            "updated_count": state["updated_count"] + state["separated_count"] + state["blocked_count"],
            "insert_count": state["created_count"],
            "update_count": state["updated_count"] + state["separated_count"] + state["blocked_count"],
            "skip_count": state["skip_count"],
            "error_count": state["error_count"],
            
            # Additional UI friendly stats keys
            "Workbook Name": file_name,
            "Processed Sheets": ", ".join(sheets_processed),
            "Skipped Sheets": ", ".join(skipped_sheets),
            "Inserted": state["created_count"],
            "Updated": state["updated_count"] + state["separated_count"] + state["blocked_count"],
            "Skipped": state["skip_count"],
            "Failed": state["error_count"]
        }

    @classmethod
    def _save_upload_history(
        cls,
        db: Session,
        upload_id: str,
        file_name: str,
        file_hash: str,
        file_size: int,
        upload_type: str,
        uploaded_by: str,
        processing_time: float,
        status: str,
        is_duplicate: bool,
        stats: dict,
        state: dict,
        employee_sheets: List[str],
        separation_sheets: List[str],
        invoice_sheets: List[str],
        remarks: Optional[str] = None
    ) -> None:
        hist = UploadHistory(
            upload_id=upload_id,
            file_name=file_name,
            file_hash=file_hash,
            file_size=file_size,
            upload_type=upload_type,
            uploaded_by=uploaded_by,
            processing_time=processing_time,
            status=status,
            is_duplicate=is_duplicate,
            workbook_version=stats.get("workbook_version"),
            parser_version=stats.get("parser_version", "2.0.0"),
            application_version="2.0.0",
            sheet_count=stats.get("number_of_sheets", 0),
            visible_sheet_count=len(stats.get("sheets_processed", [])),
            hidden_sheet_count=len(stats.get("skipped_sheets", [])),
            rows_processed=state.get("rows_processed", 0),
            rows_inserted=state.get("created_count", 0),
            rows_updated=state.get("updated_count", 0),
            rows_no_change=state.get("rows_no_change", 0),
            rows_skipped=state.get("skip_count", 0),
            rows_failed=state.get("error_count", 0),
            rows_rehired=state.get("rows_rehired", 0),
            rows_inactive=state.get("rows_inactive", 0),
            employee_sheets=employee_sheets,
            separation_sheets=separation_sheets,
            invoice_sheets=invoice_sheets,
            remarks=remarks
        )
        db.add(hist)
        db.commit()

    @classmethod
    def import_separation_workbook(cls, db: Session, file_content: bytes, file_name: str, upload_mode: str = "MASTER", operator: str = "Admin") -> Dict[str, Any]:
        """Processes the Separation Workbook (NAPS, B.Tech, M.Tech) using Universal WorkbookParser."""
        import time
        import uuid
        import hashlib
        start_time = time.time()
        
        # Compute file hash and check duplicate
        file_hash = hashlib.sha256(file_content).hexdigest()
        file_size = len(file_content)
        
        # Unique UUID for the upload history
        current_upload_id = str(uuid.uuid4())
        
        existing_upload = db.query(UploadHistory).filter(UploadHistory.file_hash == file_hash).first()
        is_duplicate = False
        duplicate_warning = None
        if existing_upload:
            is_duplicate = True
            duplicate_warning = f"Workbook '{file_name}' was already uploaded previously (Upload ID: {existing_upload.upload_id})."
            
        # Create UploadHistory first and flush it
        upload = UploadHistory(
            upload_id=current_upload_id,
            file_name=file_name,
            file_hash=file_hash,
            file_size=file_size,
            upload_type="SEPARATION",
            uploaded_by=operator,
            processing_time=0.0,
            status="PROCESSING",
            is_duplicate=is_duplicate,
            remarks=duplicate_warning
        )
        db.add(upload)
        db.flush()

        # Count before import
        count_before = db.query(Trainee).count()
        print(f"Print count before import: {count_before}")

        parsed_wb = WorkbookParser.parse_workbook(file_content, file_name, upload_mode=upload_mode)
        wb_stats = parsed_wb["stats"]
        
        warnings = list(wb_stats["warnings"])
        if duplicate_warning:
            warnings.append(duplicate_warning)
        errors = list(wb_stats["errors"])
        
        state = {
            "success_count": 0,
            "skipped_count": 0,
            "created_count": 0,
            "updated_count": 0,
            "skip_count": 0,
            "error_count": 0,
            "duplicate_count": 0,
            "blocked_count": 0,
            "separated_count": 0,
            "early_separations_count": 0,
            "unknown_employees_count": 0,
            "rows_processed": 0,
            "rows_no_change": 0,
            "rows_rehired": 0,
            "rows_inactive": 0,
            "warnings": warnings,
            "errors": errors,
            "upload_id": current_upload_id,
            "sheet_summaries": {}
        }

        # Resolve upload_mode scheme scope
        upload_mode_upper = upload_mode.upper()
        expected_scheme = None
        if upload_mode_upper != "MASTER":
            if upload_mode_upper == "BTECH":
                expected_scheme = "B.Tech"
            elif upload_mode_upper == "MTECH":
                expected_scheme = "M.Tech"
            else:
                expected_scheme = upload_mode.title() if len(upload_mode) > 4 else upload_mode.upper()

        sheets_processed = []
        employee_sheets = []
        separation_sheets = []
        skipped_sheets = list(wb_stats["skipped_sheets"])
        failed_sheets = list(wb_stats["failed_sheets"])
        
        BATCH_SIZE = 2000
        
        for sheet in parsed_wb["sheets"]:
            sheet_name = sheet["sheet_name"]
            sheet_type = sheet["sheet_type"]
            
            # Derive sheet scheme
            first_row = sheet["rows"][0] if sheet["rows"] else {}
            _, sheet_scheme = cls.derive_category_and_scheme(sheet_name, first_row, file_name)
            
            # If individual scheme mode selected, validate & warn
            if expected_scheme and sheet_type in ("Employee Master", "BDC", "Separation"):
                if sheet_scheme.upper().replace(".", "") != expected_scheme.upper().replace(".", ""):
                    skipped_sheets.append(sheet_name)
                    state["warnings"].append(
                        f"Workbook contains sheet '{sheet_name}' which appears to belong to scheme '{sheet_scheme}', "
                        f"but '{expected_scheme}' was selected. This sheet was skipped."
                    )
                    continue
            
            if sheet_type == "Separation":
                sheets_processed.append(sheet_name)
                separation_sheets.append(sheet_name)
                cls._process_separation_sheet(db, sheet, file_name, BATCH_SIZE, state)
            elif sheet_type in ("Employee Master", "BDC"):
                skipped_sheets.append(sheet_name)
                state["warnings"].append(f"Sheet '{sheet_name}' ignored. Classification: '{sheet_type}'. Reason: Sheet type is not supported for Separation sync.")
            else:
                skipped_sheets.append(sheet_name)
                state["warnings"].append(f"Sheet '{sheet_name}' ignored. Classification: '{sheet_type}'. Reason: Sheet type is not supported for Separation sync.")

        if not sheets_processed:
            raise ValueError("No sheets with required columns found.")
            
        total_duration = time.time() - start_time
        
        upload_status = "SUCCESS"
        if failed_sheets:
            upload_status = "PARTIAL" if sheets_processed else "FAILED"
            
        # Update existing UploadHistory summary fields in database
        upload.upload_type = "SEPARATION"
        upload.processing_time = total_duration
        upload.status = upload_status
        upload.workbook_version = wb_stats.get("workbook_version")
        upload.parser_version = wb_stats.get("parser_version", "2.0.0")
        upload.sheet_count = wb_stats.get("number_of_sheets", 0)
        upload.visible_sheet_count = len(sheets_processed)
        upload.hidden_sheet_count = len(skipped_sheets)
        upload.rows_processed = state.get("rows_processed", 0)
        upload.rows_inserted = 0
        upload.rows_updated = state.get("separated_count", 0) + state.get("blocked_count", 0)
        upload.rows_no_change = state.get("rows_no_change", 0)
        upload.rows_skipped = state.get("skip_count", 0)
        upload.rows_failed = state.get("error_count", 0)
        upload.rows_rehired = 0
        upload.rows_inactive = 0
        upload.employee_sheets = []
        upload.separation_sheets = separation_sheets
        upload.invoice_sheets = []
        upload.remarks = duplicate_warning
        db.flush()
        
        AuditLogRepository.add_log(
            db=db,
            action="IMPORT_SEPARATION_WORKBOOK",
            module="SEPARATION_UPLOAD",
            details=f"File: {file_name}. Sheets processed: {', '.join(sheets_processed)}. Separated: {state['separated_count'] + state['blocked_count']}, Errors: {state['error_count']}. Duration: {total_duration:.2f}s",
            operator=operator,
            workbook=file_name,
            sheet=", ".join(sheets_processed),
            rows_count=state["rows_processed"],
            duration=total_duration,
            inserted=0,
            updated=state["separated_count"] + state["blocked_count"],
            failed=state["error_count"],
            warnings=len(state["warnings"]),
            errors=state["error_count"],
            upload_id=current_upload_id,
            commit=False
        )
        
        print(f"rows received: {state['rows_processed']}")
        print(f"rows inserted: {state['separated_count'] + state['blocked_count']}")
        print(f"rows updated: {state['skip_count']}")
        print(f"rows skipped: {state['skipped_count'] - state['error_count']}")
        print(f"rows failed: {state['error_count']}")
        print("commit started")
        db.commit()
        print("commit completed")
        # Print count after import
        count_after = db.query(Trainee).count()
        state["skipped_count"] = state["skip_count"]
        return {
            "success": True,
            "processed": state["rows_processed"],
            "created": state["separated_count"] + state["blocked_count"],
            "updated": state["skip_count"],
            "failed": state["error_count"],
            "message": f"Successfully processed {state['rows_processed']} rows.",
            "workbook_name": file_name,
            "number_of_sheets": len(parsed_wb["sheets"]) + len(skipped_sheets),
            "processed_sheets": sheets_processed,
            "skipped_sheets": skipped_sheets,
            "failed_sheets": failed_sheets,
            "inserted_records": state["separated_count"] + state["blocked_count"],
            "updated_records": state["skip_count"],
            "skipped_records": state["skipped_count"] - state["error_count"],
            "failed_records": state["error_count"],
            "warnings": state["warnings"],
            "errors": state["errors"],
            "upload_id": current_upload_id,
            "is_duplicate": is_duplicate,
            "sheet_summaries": list(state.get("sheet_summaries", {}).values()),
            
            # Spec compliance keys
            "employee_sheets": employee_sheets,
            "separation_sheets": separation_sheets,
            "rows_processed": state["rows_processed"],
            "rows_imported": 0,
            "rows_updated": state["separated_count"] + state["blocked_count"],
            "rows_skipped": state["skip_count"],
            "rows_failed": state["error_count"],
            "rows_no_change": state["rows_no_change"],
            "rows_rehired": state["rows_rehired"],
            "rows_inactive": state["rows_inactive"],
            "processing_time": total_duration,
            
            # Legacy compatibility keys
            "sheets_processed": sheets_processed,
            "employees_updated": state["separated_count"] + state["blocked_count"],
            "early_separations": state["early_separations_count"],
            "blocked_employees": state["blocked_count"],
            "inserted": 0,
            "updated": state["separated_count"] + state["blocked_count"],
            "blocked": state["blocked_count"],
            "separated": state["separated_count"],
            "duplicates": state["duplicate_count"],
            "unknown_employees": state["unknown_employees_count"],
            
            # original keys
            "success_count": state["success_count"],
            "skipped_count": state["skipped_count"],
            "errors": state["errors"],
            "total_records": state["success_count"] + state["skipped_count"],
            "separated_count": state["separated_count"],
            "already_synced_count": state["skip_count"]
        }

    @classmethod
    def parse_pair(cls, pair_val: Any) -> Dict[str, float]:
        """
        Parse a textual pair value into counts of pairs, jeans, and shirts.
        """
        res = {"pair_count": 0.0, "jeans_count": 0.0, "shirt_count": 0.0}
        if pair_val is None:
            return res
        pair_str = str(pair_val).strip().lower()
        if not pair_str or pair_str.lower() in ('nan', 'none', 'null') or pair_str.startswith('#'):
            return res

        # Check for specific jeans and shirt counts
        # E.g. "2 Pair Jeans & 2 Pair T Shirt" or "1.5 pr jeans, 1 shirt"
        import re
        jeans_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:pair|pr)?s?\s*(?:of)?\s*(?:jeans?|jean)', pair_str)
        shirt_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:pair|pr)?s?\s*(?:of)?\s*(?:t\s*shirts?|t-shirts?|tshirts?|shirts?|shirt)', pair_str)
        
        has_specific = False
        if jeans_match:
            res["jeans_count"] = float(jeans_match.group(1))
            has_specific = True
        if shirt_match:
            res["shirt_count"] = float(shirt_match.group(1))
            has_specific = True

        # If it is like "2 Pair" (no specific jeans/shirt count), or has a generic pair prefix
        # E.g. "2 Pair", "1.5 pr", "3"
        generic_match = re.search(r'^(\d+(?:\.\d+)?)\s*(?:pair|pr)?s?$', pair_str)
        val_float = 0.0
        if not generic_match:
            try:
                val_float = float(pair_str)
                generic_match = True
            except ValueError:
                generic_match = False
        else:
            val_float = float(generic_match.group(1))

        if generic_match and not has_specific:
            res["pair_count"] = val_float
            res["jeans_count"] = val_float
            res["shirt_count"] = val_float
        else:
            leading_match = re.match(r'^(\d+(?:\.\d+)?)\s*(?:pair|pr)s?\b', pair_str)
            if leading_match:
                res["pair_count"] = float(leading_match.group(1))
            else:
                res["pair_count"] = max(res["jeans_count"], res["shirt_count"])
                
        return res

    @classmethod
    def parse_pdf_invoice(cls, file_content: bytes, file_name: str) -> Dict[str, Any]:
        """
        Parses a PDF invoice using pdfplumber, extracting tables, matching headers using synonyms,
        and returning a structure compatible with WorkbookParser.parse_workbook.
        """
        import io
        import pdfplumber
        
        stats = {
            "workbook_name": file_name,
            "number_of_sheets": 1,
            "sheets_processed": ["Invoice"],
            "skipped_sheets": [],
            "failed_sheets": [],
            "warnings": [],
            "errors": []
        }
        
        all_rows = []
        headers = None
        header_idx = None
        
        # Define synonyms mapping for header detection
        invoice_synonyms = {
            "ticket_number": ["t no", "ticket no", "ticket number", "pers no", "ticket", "ticketid", "tktno", "tktnumber", "boarding ticket", "ticket id", "persno", "personnelno", "personnelnumber", "emp id", "employee id", "employee_id", "emp_id", "trainee id", "trainee_id"],
            "joining_date": ["date of joining", "joining date", "doj", "joining"],
            "batch": ["batch no", "batch"],
            "candidate_name": ["candidates name", "candidate name", "employee name", "name", "completename", "completname", "traineename", "employeename", "firstname", "fullname", "billed name", "employee_name"],
            "pair": ["pair", "uniform pair", "kit pair"],
            "amount": ["amount", "bill amount", "billing amount", "claimed amount", "total", "billed total", "total amount", "total_amount", "amount_paid"],
            "distribution_date": ["distribution date", "issued date", "date"],
            "page_number": ["page no", "page", "page_number"]
        }
        
        try:
            with pdfplumber.open(io.BytesIO(file_content)) as pdf:
                for page_num, page in enumerate(pdf.pages, start=1):
                    tables = page.extract_tables()
                    for table in tables:
                        if not table:
                            continue
                        
                        # Find headers if not yet found
                        for row_idx, row in enumerate(table):
                            cleaned_row = [str(c).strip() if c is not None else "" for c in row]
                            
                            if headers is None:
                                matched_fields = set()
                                for cell in cleaned_row:
                                    norm_cell = WorkbookParser.normalize_header(cell)
                                    for field, syns in invoice_synonyms.items():
                                        if norm_cell in [WorkbookParser.normalize_header(s) for s in syns]:
                                            matched_fields.add(field)
                                            break
                                
                                if len(matched_fields) >= 2:
                                    headers = cleaned_row
                                    header_idx = row_idx
                                    continue
                            
                            if headers is not None:
                                if row_idx == header_idx and len(all_rows) == 0:
                                    continue
                                if all(c == "" for c in cleaned_row):
                                    continue
                                all_rows.append(cleaned_row)
        except Exception as e:
            stats["failed_sheets"].append("Invoice")
            stats["errors"].append(f"Failed parsing PDF: {str(e)}")
            return {
                "stats": stats,
                "sheets": []
            }
            
        if headers is None:
            stats["skipped_sheets"].append("Invoice")
            stats["warnings"].append("Could not detect invoice headers in PDF.")
            return {
                "stats": stats,
                "sheets": []
            }
            
        # Dedup headers
        normalized_headers = []
        seen_headers = {}
        for idx, orig_h in enumerate(headers):
            norm_h = WorkbookParser.normalize_header(orig_h)
            if not norm_h:
                norm_h = f"empty_col_{idx + 1}"
            if norm_h in seen_headers:
                seen_headers[norm_h] += 1
                norm_h = f"{norm_h}_{seen_headers[norm_h]}"
            else:
                seen_headers[norm_h] = 1
            normalized_headers.append(norm_h)
            
        rows_data = []
        blank_rows_count = 0
        valid_rows_count = 0
        
        for r_idx, row in enumerate(all_rows, start=1):
            row_cells = row + [""] * (len(headers) - len(row))
            row_cells = row_cells[:len(headers)]
            
            if all(val is None or str(val).strip() == "" for val in row_cells):
                blank_rows_count += 1
                continue
                
            row_dict = {}
            for idx, norm_h in enumerate(normalized_headers):
                val = row_cells[idx]
                row_dict[norm_h] = val
                
            row_dict["_row_num"] = r_idx
            rows_data.append(row_dict)
            valid_rows_count += 1
            
        parsed_sheets = [{
            "sheet_name": "Invoice",
            "sheet_type": "Invoice",
            "original_headers": headers,
            "normalized_headers": normalized_headers,
            "rows": rows_data,
            "total_rows": len(all_rows) + 1,
            "valid_rows": valid_rows_count,
            "blank_rows": blank_rows_count
        }]
        
        return {
            "stats": stats,
            "sheets": parsed_sheets
        }

    @classmethod
    def parse_month_year(cls, val_str: str) -> Tuple[Optional[str], Optional[int]]:
        month_pattern = re.compile(
            r'\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\b',
            re.IGNORECASE
        )
        m_mon = month_pattern.search(val_str)
        
        year_pattern = re.compile(r'\b(20\d{2}|\d{2})\b')
        m_yr = year_pattern.search(val_str)
        
        mon = None
        yr = None
        if m_mon:
            months_map_full = {
                "jan": "January", "feb": "February", "mar": "March", "apr": "April",
                "may": "May", "jun": "June", "jul": "July", "aug": "August",
                "sep": "September", "oct": "October", "nov": "November", "dec": "December"
            }
            mon_raw = m_mon.group(1).lower()[:3]
            mon = months_map_full.get(mon_raw, m_mon.group(1).capitalize())
            
        if m_yr:
            yr_str = m_yr.group(1)
            if len(yr_str) == 2:
                yr = 2000 + int(yr_str)
            else:
                yr = int(yr_str)
                
        return mon, yr

    @classmethod
    def detect_sheet_metadata(
        cls, 
        sheet: Dict[str, Any], 
        file_content: bytes,
        file_name: str, 
        invoice_number_override: Optional[str] = None, 
        invoice_date_override: Optional[datetime.date] = None
    ) -> Tuple[str, datetime.date, str, int, str]:
        import openpyxl
        import io
        
        inv_num = invoice_number_override
        inv_date = invoice_date_override
        billing_month = None
        billing_year = None
        vendor_name = None
        
        # 1. Search columns of first row
        if sheet["rows"]:
            first_row = sheet["rows"][0]
            if not inv_num:
                inv_num = cls.get_column_value(first_row, ['invoice number', 'invoice_no', 'invoice no', 'bill no', 'bill number', 'reference number', 'reference no'])
            if not inv_date:
                raw_date = cls.get_column_value(first_row, ['invoice date', 'date', 'bill date'])
                inv_date = WorkbookParser.parse_date(raw_date)
            
            # Check for billing month
            billing_month_val = cls.get_column_value(first_row, ['billing month', 'invoice month', 'month'])
            if billing_month_val:
                billing_month, billing_year = cls.parse_month_year(str(billing_month_val))
                
            # Check for vendor
            vendor_val = cls.get_column_value(first_row, ['vendor', 'vendor name', 'vendor_name', 'supplier', 'supplier name'])
            if vendor_val:
                vendor_name = str(vendor_val).strip()

        # 2. Search cells of top rows if Excel
        if not file_name.lower().endswith(".pdf") and (not inv_num or not inv_date or not billing_month or not vendor_name):
            try:
                # Load sheet specifically
                wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
                if sheet["sheet_name"] in wb.sheetnames:
                    ws = wb[sheet["sheet_name"]]
                    # Scan first 15 rows
                    for r in range(1, 16):
                        for c in range(1, min(ws.max_column + 1, 30)):
                            val = ws.cell(row=r, column=c).value
                            if val and isinstance(val, str):
                                val_clean = val.strip()
                                if not inv_num:
                                    m = re.search(r'(?:invoice\s*no|invoice\s*number|bill\s*no|bill\s*number|reference\s*number)\s*[:\-]?\s*([a-zA-Z0-9\-\/]+)', val_clean, re.IGNORECASE)
                                    if m:
                                        inv_num = m.group(1)
                                if not inv_date:
                                    m = re.search(r'(?:invoice\s*date|bill\s*date|date)\s*[:\-]?\s*([\d\-\.\/]+)', val_clean, re.IGNORECASE)
                                    if m:
                                        inv_date = WorkbookParser.parse_date(m.group(1))
                                if not billing_month:
                                    m = re.search(r'(?:billing\s*month|invoice\s*month|month)\s*[:\-]?\s*([a-zA-Z0-9\s]+)', val_clean, re.IGNORECASE)
                                    if m:
                                        billing_month, billing_year = cls.parse_month_year(m.group(1))
                                if not vendor_name:
                                    m = re.search(r'(?:vendor\s*name|vendor|supplier|supplier\s*name)\s*[:\-]?\s*([a-zA-Z0-9\s\.\,\-\&]+)', val_clean, re.IGNORECASE)
                                    if m:
                                        vendor_name = m.group(1)
                wb.close()
            except Exception:
                pass

        # 3. Fallback from sheet name for billing month/year
        if not billing_month:
            billing_month, billing_year = cls.parse_month_year(sheet["sheet_name"])
            
        # 4. Derive billing month from invoice date if still absent
        if not billing_month and inv_date:
            months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
            billing_month = months[inv_date.month - 1]
            billing_year = inv_date.year

        # 5. Fallback defaults
        if not inv_num:
            base_name = os.path.splitext(file_name)[0]
            inv_num = f"{base_name}_{sheet['sheet_name']}".replace(" ", "_")
        if not inv_date:
            inv_date = datetime.date.today()
        if not billing_month:
            months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
            billing_month = months[inv_date.month - 1]
            billing_year = inv_date.year
        if not vendor_name:
            vendor_name = "Tata Projects"

        return str(inv_num).strip(), inv_date, billing_month, billing_year, vendor_name

    @classmethod
    def import_invoice_workbook(
        cls, 
        db: Session, 
        file_content: bytes, 
        file_name: str,
        invoice_number_override: Optional[str] = None,
        invoice_date_override: Optional[datetime.date] = None,
        operator: str = "Admin"
    ) -> Dict[str, Any]:
        """Processes a Monthly Invoice Workbook/PDF using Universal parser with synonym matching."""
        import time
        import uuid
        import hashlib
        import os
        start_time = time.time()
        
        file_hash = hashlib.sha256(file_content).hexdigest()
        file_size = len(file_content)
        current_upload_id = str(uuid.uuid4())
        
        existing_upload = db.query(UploadHistory).filter(UploadHistory.file_hash == file_hash).first()
        is_duplicate = False
        duplicate_warning = None
        if existing_upload:
            is_duplicate = True
            duplicate_warning = f"Invoice file '{file_name}' was already uploaded previously (Upload ID: {existing_upload.upload_id})."

        # Create UploadHistory first and flush it to the session to prevent FOREIGN KEY constraint issues
        upload = UploadHistory(
            upload_id=current_upload_id,
            file_name=file_name,
            file_hash=file_hash,
            file_size=file_size,
            upload_type="INVOICE",
            uploaded_by=operator,
            processing_time=0.0,
            status="PROCESSING",
            is_duplicate=is_duplicate,
            remarks=duplicate_warning
        )
        db.add(upload)
        db.flush()
        
        is_pdf = file_name.lower().endswith(".pdf")
        # Count before import
        count_before = db.query(InvoiceItem).count()
        print(f"Print count before import: {count_before}")
        if is_pdf:
            parsed_wb = cls.parse_pdf_invoice(file_content, file_name)
        else:
            parsed_wb = WorkbookParser.parse_workbook(file_content, file_name, upload_mode="INVOICE")
            
        wb_stats = parsed_wb["stats"]
        
        warnings = list(wb_stats.get("warnings", []))
        if duplicate_warning:
            warnings.append(duplicate_warning)
        errors = list(wb_stats.get("errors", []))
        
        invoice_synonyms = {
            "ticket_number": ["t no", "ticket no", "ticket number", "pers no", "ticket", "ticketid", "tktno", "tktnumber", "boarding ticket", "ticket id", "persno", "personnelno", "personnelnumber", "emp id", "employee id", "employee_id", "emp_id", "trainee id", "trainee_id"],
            "joining_date": ["date of joining", "joining date", "doj", "joining"],
            "batch": ["batch no", "batch"],
            "candidate_name": ["candidates name", "candidate name", "employee name", "name", "completename", "completname", "traineename", "employeename", "firstname", "fullname", "billed name", "employee_name"],
            "pair": ["pair", "uniform pair", "kit pair"],
            "amount": ["amount", "bill amount", "billing amount", "claimed amount", "total", "billed total", "total amount", "total_amount", "amount_paid"],
            "distribution_date": ["distribution date", "issued date", "date"],
            "page_number": ["page no", "page", "page_number"]
        }
        
        success_count = 0
        skipped_count = 0
        error_count = 0
        
        sheets_processed = []
        skipped_sheets = list(wb_stats["skipped_sheets"])
        failed_sheets = list(wb_stats["failed_sheets"])
        
        first_inv_num = None
        first_inv_date = None
        
        BATCH_SIZE = 2000
        
        for sheet in parsed_wb["sheets"]:
            sheet_name = sheet["sheet_name"]
            sheets_processed.append(sheet_name)
            
            # Detect metadata per sheet
            inv_num, inv_date, billing_month, billing_year, vendor_name = cls.detect_sheet_metadata(
                sheet=sheet,
                file_content=file_content,
                file_name=file_name,
                invoice_number_override=invoice_number_override,
                invoice_date_override=invoice_date_override
            )
            
            if not first_inv_num:
                first_inv_num = inv_num
                first_inv_date = inv_date
                
            # Keep previous invoices but mark them as SUPERSEDED if they are currently ACTIVE
            existing_active = db.query(Invoice).filter(
                Invoice.invoice_number == inv_num,
                Invoice.status == "ACTIVE"
            ).all()
            for ext_inv in existing_active:
                ext_inv.status = "SUPERSEDED"
            db.flush()
            
            # Create parent Invoice
            invoice = Invoice(
                invoice_number=inv_num,
                invoice_date=inv_date,
                billing_month=billing_month,
                billing_year=billing_year,
                vendor_name=vendor_name,
                workbook_name=file_name,
                sheet_name=sheet_name,
                upload_id=current_upload_id,
                uploaded_by=operator,
                uploaded_at=datetime.datetime.utcnow(),
                status="ACTIVE",
                total_amount=0.0,
                approved_amount=0.0,
                rejected_amount=0.0,
                fraud_amount=0.0
            )
            db.add(invoice)
            db.flush()
            
            sheet_total_amount = 0.0
            
            def process_batch(rows_to_process, current_sheet, parent_invoice):
                nonlocal success_count, skipped_count, error_count, sheet_total_amount
                
                batch_trainee_ids = set()
                batch_tickets = set()
                valid_rows = []
                
                for row in rows_to_process:
                    row_num = row["_row_num"]
                    
                    r_ticket = cls.get_column_value(row, invoice_synonyms["ticket_number"])
                    r_ticket = str(r_ticket).strip() if r_ticket is not None else ""
                    
                    trainee_id = r_ticket
                    direct_id = cls.get_column_value(row, ['trainee id', 'emp id', 'employee id', 'trainee_id', 'reg no'])
                    if direct_id:
                        trainee_id = str(direct_id).strip()
                    
                    if not trainee_id and not r_ticket:
                        skipped_count += 1
                        error_count += 1
                        errors.append(f"Sheet '{current_sheet}', Row {row_num}: Missing both Trainee ID and Ticket Number.")
                        continue
                        
                    if trainee_id:
                        batch_trainee_ids.add(trainee_id)
                    if r_ticket:
                        batch_tickets.add(r_ticket)
                        
                    valid_rows.append((row_num, trainee_id, r_ticket, row))
                    
                if not valid_rows:
                    return
                    
                db_trainees_by_id = {}
                if batch_trainee_ids:
                    res_id = db.query(Trainee).filter(Trainee.id.in_(batch_trainee_ids)).all()
                    db_trainees_by_id = {t.id: t for t in res_id}
                    
                db_trainees_by_ticket = {}
                if batch_tickets:
                    res_ticket = db.query(Trainee).filter(Trainee.ticket_number.in_(batch_tickets)).all()
                    db_trainees_by_ticket = {t.ticket_number: t for t in res_ticket}
                    
                for row_num, trainee_id, r_ticket, row in valid_rows:
                    try:
                        trainee = None
                        if trainee_id and trainee_id in db_trainees_by_id:
                            trainee = db_trainees_by_id[trainee_id]
                        elif r_ticket and r_ticket in db_trainees_by_ticket:
                            trainee = db_trainees_by_ticket[r_ticket]
                            
                        exists_in_master = trainee is not None
                        resolved_trainee_id = trainee.id if exists_in_master else None
                        trainee_name = trainee.name if exists_in_master else None
                        
                        billed_name = cls.get_column_value(row, invoice_synonyms["candidate_name"])
                        billed_name = str(billed_name).strip() if billed_name is not None else ""
                        
                        joining_date_raw = cls.get_column_value(row, invoice_synonyms["joining_date"])
                        joining_date = WorkbookParser.parse_date(joining_date_raw) or (trainee.doj if trainee else None)
                        
                        distribution_date_raw = cls.get_column_value(row, invoice_synonyms["distribution_date"])
                        distribution_date = WorkbookParser.parse_date(distribution_date_raw) or parent_invoice.invoice_date
                        
                        batch = cls.get_column_value(row, invoice_synonyms["batch"])
                        batch = str(batch).strip() if batch is not None else (trainee.batch if trainee else "")
                        
                        pair = cls.get_column_value(row, invoice_synonyms["pair"])
                        pair = str(pair).strip() if pair is not None else ""
                        
                        amount_raw = cls.get_column_value(row, invoice_synonyms["amount"])
                        amount = WorkbookParser.parse_float(amount_raw)
                        
                        page_number_raw = cls.get_column_value(row, invoice_synonyms["page_number"])
                        page_number = None
                        if page_number_raw is not None:
                            try:
                                page_number = int(float(page_number_raw))
                            except Exception:
                                page_number = str(page_number_raw).strip()
                        
                        joining_amt = 0.0
                        days180_amt = 0.0
                        other_amt = 0.0
                        
                        stage_idx_val = cls.get_column_value(row, ['billing stage', 'stage', 'billing_stage'])
                        if stage_idx_val:
                            stage_val = str(stage_idx_val).strip().lower()
                            if "join" in stage_val:
                                joining_amt = amount
                            elif "six" in stage_val or "180" in stage_val or "month" in stage_val:
                                days180_amt = amount
                            else:
                                joining_amt = amount
                        else:
                            legacy_joining = cls.get_column_value(row, ['joining', 'joining payment', 'joining reimbursement', 'joining_amount'])
                            legacy_180 = cls.get_column_value(row, ['180 days', '180 days payment', '180 days reimbursement', '180_days_amount'])
                            
                            if legacy_joining is not None or legacy_180 is not None:
                                joining_amt = WorkbookParser.parse_float(legacy_joining)
                                days180_amt = WorkbookParser.parse_float(legacy_180)
                                other_amt = WorkbookParser.parse_float(cls.get_column_value(row, ['other', 'uniform', 'shirt', 'jeans', 'excess']))
                            else:
                                ref_doj = joining_date or (trainee.doj if trainee else None)
                                ref_dist = distribution_date or parent_invoice.invoice_date
                                
                                if ref_doj and ref_dist:
                                    delta_days = (ref_dist - ref_doj).days
                                    if delta_days >= 180:
                                        days180_amt = amount
                                    else:
                                        joining_amt = amount
                                else:
                                    joining_amt = amount
                                    
                        parsed_pair = cls.parse_pair(pair)
                        shirt_qty = parsed_pair["shirt_count"]
                        jean_qty = parsed_pair["jeans_count"]
                        pair_count = parsed_pair["pair_count"]
                        
                        if not shirt_qty:
                            shirt_qty = WorkbookParser.parse_float(cls.get_column_value(row, ['shirt quantity', 'shirt qty', 'shirt_quantity', 'shirt_qty', 'shirt']) or 0.0)
                        if not jean_qty:
                            jean_qty = WorkbookParser.parse_float(cls.get_column_value(row, ['jean quantity', 'jean qty', 'jean_quantity', 'jean_qty', 'jeans quantity', 'jeans qty', 'jeans']) or 0.0)
                        if not pair_count:
                            pair_count = max(shirt_qty, jean_qty)
                        
                        raw_row_data = {
                            "ticket_number": r_ticket,
                            "ticket number": r_ticket,
                            "ticket no": r_ticket,
                            "trainee_id": resolved_trainee_id or trainee_id,
                            "trainee id": resolved_trainee_id or trainee_id,
                            "joining_date": joining_date.strftime("%Y-%m-%d") if joining_date else "",
                            "batch": batch,
                            "candidate_name": billed_name,
                            "pair": pair,
                            "amount": amount,
                            "distribution_date": distribution_date.strftime("%Y-%m-%d") if distribution_date else "",
                            "page_number": str(page_number) if page_number is not None else "",
                            "shirt_quantity": shirt_qty,
                            "jean_quantity": jean_qty,
                            "pair_count": pair_count,
                            "jeans_count": jean_qty,
                            "shirt_count": shirt_qty,
                            "sheet_name": current_sheet
                        }
                        
                        for k, v in row.items():
                            if not k.startswith('_') and k not in raw_row_data:
                                raw_row_data[k] = str(v) if v is not None else ""
                                
                        invoice_item = InvoiceItem(
                            invoice_id=parent_invoice.invoice_id,
                            ticket_number=r_ticket,
                            candidate_name=billed_name or (trainee_name if exists_in_master else f"Unknown ({trainee_id or r_ticket})"),
                            joining_date=joining_date,
                            batch=batch,
                            pair=pair,
                            jeans_count=int(jean_qty),
                            shirt_count=int(shirt_qty),
                            claimed_amount=amount,
                            approved_amount=0.0,
                            rejected_amount=0.0,
                            distribution_date=distribution_date,
                            page_number=int(page_number) if isinstance(page_number, int) else None,
                            status="PENDING",
                            reason=None,
                            validation_summary=None,
                            
                            # Legacy compatibility columns
                            invoice_number=parent_invoice.invoice_number,
                            invoice_date=parent_invoice.invoice_date,
                            trainee_id=resolved_trainee_id,
                            billed_name=billed_name or (trainee_name if exists_in_master else f"Unknown ({trainee_id or r_ticket})"),
                            billed_joining_amount=joining_amt,
                            billed_180_days_amount=days180_amt,
                            billed_other_amount=other_amt,
                            billed_total_amount=amount,
                            approved_joining_amount=0.0,
                            approved_180_days_amount=0.0,
                            approved_total_amount=0.0,
                            file_name=file_name,
                            uploaded_at=parent_invoice.uploaded_at,
                            extra_data=raw_row_data
                        )
                        
                        db.add(invoice_item)
                        success_count += 1
                        sheet_total_amount += amount
                    except Exception as e:
                        skipped_count += 1
                        error_count += 1
                        errors.append(f"Sheet '{current_sheet}', Row {row_num}: {str(e)}")
                
                db.flush()
            
            try:
                for i in range(0, len(sheet["rows"]), BATCH_SIZE):
                    process_batch(sheet["rows"][i:i+BATCH_SIZE], sheet_name, invoice)
                invoice.total_amount = sheet_total_amount
                print(f"rows received: {success_count + skipped_count}")
                print(f"rows inserted: {success_count}")
                print(f"rows updated: 0")
                print(f"rows skipped: {skipped_count - error_count}")
                print(f"rows failed: {error_count}")
                print("commit started")
                db.commit()
                print("commit completed")
                # Print count after import
                count_after = db.query(InvoiceItem).count()
                print(f"Print count after import: {count_after}")
            except Exception as e:
                failed_sheets.append(sheet_name)
                error_count += 1
                errors.append(f"Failed to process sheet '{sheet_name}': {str(e)}")
                print("rollback")
                db.rollback()
                
        if not sheets_processed:
            raise ValueError("No sheets with required columns found.")
            
        total_duration = time.time() - start_time
        
        upload_status = "SUCCESS"
        if failed_sheets:
            upload_status = "PARTIAL" if sheets_processed else "FAILED"

        # Update existing UploadHistory summary fields in database
        upload.status = upload_status
        upload.processing_time = total_duration
        upload.workbook_version = wb_stats.get("workbook_version")
        upload.parser_version = wb_stats.get("parser_version", "2.0.0")
        upload.sheet_count = wb_stats.get("number_of_sheets", 0)
        upload.visible_sheet_count = len(sheets_processed)
        upload.hidden_sheet_count = len(skipped_sheets)
        upload.rows_processed = success_count + skipped_count
        upload.rows_inserted = success_count
        upload.rows_updated = 0
        upload.rows_no_change = 0
        upload.rows_skipped = skipped_count - error_count
        upload.rows_failed = error_count
        upload.employee_sheets = []
        upload.separation_sheets = []
        upload.invoice_sheets = sheets_processed
        upload.remarks = duplicate_warning
        db.flush()
        
        AuditLogRepository.add_log(
            db=db,
            action="IMPORT_INVOICE",
            module="INVOICE_UPLOAD",
            details=f"Invoice(s): {first_inv_num}. Billed: {success_count}, Skipped: {skipped_count}. Duration: {total_duration:.2f}s",
            operator=operator,
            workbook=file_name,
            sheet=", ".join(sheets_processed),
            rows_count=success_count + skipped_count,
            duration=total_duration,
            inserted=success_count,
            updated=0,
            failed=error_count,
            warnings=len(warnings),
            errors=error_count,
            before_state=None,
            after_state=None,
            invoice_number=first_inv_num,
            upload_id=current_upload_id
        )
        
        return {
            "success": True,
            "processed": success_count + skipped_count,
            "created": success_count,
            "updated": 0,
            "failed": error_count,
            "message": f"Successfully processed {success_count + skipped_count} rows.",
            "workbook_name": file_name,
            "number_of_sheets": len(parsed_wb["sheets"]) + len(skipped_sheets),
            "processed_sheets": sheets_processed,
            "skipped_sheets": skipped_sheets,
            "failed_sheets": failed_sheets,
            "inserted_records": success_count,
            "updated_records": 0,
            "skipped_records": skipped_count - error_count,
            "failed_records": error_count,
            "warnings": warnings,
            "errors": errors,
            "upload_id": current_upload_id,
            "is_duplicate": is_duplicate,
            
            # original keys for compatibility
            "invoice_number": first_inv_num,
            "invoice_date": first_inv_date,
            "success_count": success_count,
            "skipped_count": skipped_count,
            "errors": errors,
            "total_records": success_count + skipped_count,
            "records_imported": success_count
        }
