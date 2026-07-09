"""
TMPVL Billing Audit & Fraud Detection System — FastAPI Route Handlers.

All data-access goes through Repository classes.
All business logic goes through Service classes.
Endpoint functions only: validate input → call service/repo → map to HTTP response.
"""
import os
import re
import json
import datetime
from io import BytesIO
from typing import List, Optional, Dict, Any

from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException, Query, Header
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from backend.app.core.db import get_db
from backend.app.api.schemas import BlockTraineeRequest, SettingsPayload
from backend.app.repositories.repositories import (
    TraineeRepository,
    InvoiceRepository,
    LedgerRepository,
    ValidationRepository,
    AuditLogRepository,
)
from backend.app.services.import_service import ImportService
from backend.app.services.validation_service import ValidationService
from backend.app.services.ledger_service import LedgerService
from backend.app.services.report_service import ReportService
from backend.app.services.workbook_parser import WorkbookParser
from backend.app.models.models import Trainee, InvoiceRecord, ValidationResult, PaymentLedger, UploadHistory, TraineeLifecycle
import uuid

router = APIRouter()

def get_offline_user() -> str:
    import os
    import platform
    try:
        return os.getlogin()
    except Exception:
        return os.environ.get("USER") or os.environ.get("USERNAME") or platform.node() or "Administrator"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SETTINGS_FILE = os.path.join(_BASE_DIR, "settings_config.json")

_EXCEL_EXTENSIONS = {".xlsx", ".xls"}
_MAX_LOG_LIMIT = 500
_MAX_UPLOAD_BYTES = 50 * 1024 * 1024  # 50 MB


def _assert_file_extension(filename: str, allowed_extensions: set) -> None:
    """Raise 400 if the uploaded file is not a recognised format."""
    ext = os.path.splitext(filename or "")[-1].lower()
    if ext not in allowed_extensions:
        ext_str = ", ".join(allowed_extensions)
        raise HTTPException(
            status_code=400,
            detail=f"Only files with extensions {ext_str} are accepted."
        )


async def _read_upload_safely(file: UploadFile, allowed_extensions: Optional[set] = None) -> tuple:
    """Validate upload file, enforce size limit, and return (content, safe_filename)."""
    filename = file.filename or "upload.xlsx"
    if allowed_extensions is None:
        allowed_extensions = _EXCEL_EXTENSIONS
    _assert_file_extension(filename, allowed_extensions)

    content = await file.read()
    if len(content) > _MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File too large. Maximum upload size is {_MAX_UPLOAD_BYTES // (1024 * 1024)} MB."
        )

    return content, filename


def _severity_from_flags(statuses: List[str]) -> str:
    """Return the highest-severity label from a list of validation status strings."""
    if "FRAUD" in statuses:
        return "FRAUD"
    if "ERROR" in statuses:
        return "ERROR"
    if "WARNING" in statuses:
        return "WARNING"
    return "OK"


def _sanitize_filename(name: str) -> str:
    """Strip characters that could inject Content-Disposition headers."""
    return re.sub(r'[^\w\-.]', '_', name)


# ---------------------------------------------------------------------------
# DASHBOARD
# ---------------------------------------------------------------------------

@router.get("/dashboard/stats")
def get_dashboard_stats(db: Session = Depends(get_db)):
    """Aggregate KPIs, chart data, and recent alerts for the dashboard."""
    try:
        trainee_counts = TraineeRepository.count_by_status(db)

        # 1. Scheme Breakdown
        scheme_counts = db.query(Trainee.scheme, func.count(Trainee.id)).group_by(Trainee.scheme).all()
        scheme_map = {s: count for s, count in scheme_counts}
        naps_count = scheme_map.get("NAPS", 0)
        btech_count = scheme_map.get("B.Tech", 0)
        mtech_count = scheme_map.get("M.Tech", 0)

        # 2. Join / Separation / Early exit this month
        today = datetime.date.today()
        start_of_month = datetime.date(today.year, today.month, 1)
        joining_this_month = db.query(func.count(Trainee.id)).filter(
            Trainee.doj >= start_of_month,
            Trainee.doj <= today
        ).scalar() or 0

        separations_this_month = db.query(func.count(Trainee.id)).filter(
            Trainee.dol >= start_of_month,
            Trainee.dol <= today
        ).scalar() or 0

        early_separations = db.query(func.count(Trainee.id)).filter(
            Trainee.dol.isnot(None),
            func.julianday(Trainee.dol) - func.julianday(Trainee.doj) < 30
        ).scalar() or 0

        # 3. Pending Invoices
        pending_invoices_count = db.query(func.count(func.distinct(InvoiceRecord.invoice_number))).filter(
            InvoiceRecord.status == "PENDING"
        ).scalar() or 0
        pending_invoices_amount = db.query(func.sum(InvoiceRecord.billed_total_amount)).filter(
            InvoiceRecord.status == "PENDING"
        ).scalar() or 0.0

        # 4. Billed vs Approved vs Rejected
        invoices = InvoiceRepository.get_unique_invoices(db)
        total_billed   = sum(i["billed_amount"]   for i in invoices)
        total_approved = sum(i["approved_amount"] for i in invoices)
        total_rejected = max(0.0, total_billed - total_approved)

        # 5. Total Payments
        total_payments = db.query(func.sum(PaymentLedger.amount_paid)).scalar() or 0.0

        # 6. Kit Distribution
        approved_records = db.query(InvoiceRecord.extra_data).filter(
            InvoiceRecord.status.in_(["VALIDATED", "APPROVED"])
        ).all()
        total_shirts = 0.0
        total_jeans = 0.0
        for r in approved_records:
            if r.extra_data and isinstance(r.extra_data, dict):
                try:
                    total_shirts += float(r.extra_data.get("shirt_quantity") or 0.0)
                except Exception:
                    pass
                try:
                    total_jeans += float(r.extra_data.get("jean_quantity") or 0.0)
                except Exception:
                    pass

        # 7. Validation counts
        exception_count = ValidationRepository.count_by_statuses(db, ["WARNING", "ERROR"])
        fraud_count = ValidationRepository.count_by_statuses(db, ["FRAUD"])

        # Recent alerts
        recent_alerts_raw = (
            db.query(ValidationResult)
            .order_by(ValidationResult.created_at.desc())
            .limit(10)
            .all()
        )
        alerts_list = [
            {
                "id": a.id,
                "trainee_id": a.trainee_id,
                "invoice_number": a.invoice_record.invoice_number if a.invoice_record else "N/A",
                "rule_name": a.rule_name,
                "status": a.status,
                "message": a.message,
                "created_at": a.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            }
            for a in recent_alerts_raw
        ]

        # Billing Chart (existing)
        chart_billing = [
            {
                "name": inv["invoice_number"],
                "billed": inv["billed_amount"],
                "approved": inv["approved_amount"],
            }
            for inv in invoices
        ]

        # Monthly Joining Chart
        joins_by_month = db.query(
            func.strftime('%Y-%m', Trainee.doj).label('month'),
            func.count(Trainee.id)
        ).group_by('month').order_by('month').all()
        chart_joining = [{"month": m, "count": count} for m, count in joins_by_month if m]

        # Monthly Separation Chart
        seps_by_month = db.query(
            func.strftime('%Y-%m', Trainee.dol).label('month'),
            func.count(Trainee.id)
        ).filter(Trainee.dol.isnot(None)).group_by('month').order_by('month').all()
        chart_separation = [{"month": m, "count": count} for m, count in seps_by_month if m]

        # Category Chart
        chart_categories = [
            {"name": "NAPS", "value": naps_count},
            {"name": "B.Tech", "value": btech_count},
            {"name": "M.Tech", "value": mtech_count}
        ]

        # Fraud trend
        fraud_by_month = db.query(
            func.strftime('%Y-%m', ValidationResult.created_at).label('month'),
            func.count(ValidationResult.id)
        ).filter(ValidationResult.status == "FRAUD").group_by('month').order_by('month').all()
        chart_fraud = [{"month": m, "count": count} for m, count in fraud_by_month if m]

        # Payment trend
        payments_by_month = db.query(
            func.strftime('%Y-%m', PaymentLedger.payment_date).label('month'),
            func.sum(PaymentLedger.amount_paid)
        ).group_by('month').order_by('month').all()
        chart_payments = [{"month": m, "amount": float(amount or 0.0)} for m, amount in payments_by_month if m]

        return {
            "total_trainees": trainee_counts["total"],
            "active_trainees": trainee_counts["ACTIVE"],
            "blocked_trainees": trainee_counts["BLOCKED"],
            "separated_trainees": trainee_counts["SEPARATED"],
            "naps_count": naps_count,
            "btech_count": btech_count,
            "mtech_count": mtech_count,
            "joining_this_month": joining_this_month,
            "separations_this_month": separations_this_month,
            "early_separations": early_separations,
            "pending_invoices_count": pending_invoices_count,
            "pending_invoices_amount": float(pending_invoices_amount),
            "total_billed_amount": total_billed,
            "total_approved_amount": total_approved,
            "total_rejected_amount": total_rejected,
            "savings_generated": total_rejected,
            "total_payments": float(total_payments),
            "total_shirts": int(total_shirts),
            "total_jeans": int(total_jeans),
            "exception_count": exception_count,
            "fraud_count": fraud_count,
            "recent_alerts": alerts_list,
            "chart_billing": chart_billing,
            "chart_joining": chart_joining,
            "chart_separation": chart_separation,
            "chart_categories": chart_categories,
            "chart_fraud": chart_fraud,
            "chart_payments": chart_payments
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/dashboard/fraud-alerts")
def get_dashboard_fraud_alerts(db: Session = Depends(get_db)):
    """Return a list of all validation results with status FRAUD."""
    try:
        alerts = db.query(ValidationResult).filter(ValidationResult.status == "FRAUD").all()
        return [
            {
                "id": a.id,
                "trainee_id": a.trainee_id,
                "invoice_number": a.invoice_record.invoice_number if a.invoice_record else "N/A",
                "rule_name": a.rule_name,
                "status": a.status,
                "message": a.message,
                "reason_code": a.reason_code,
                "recommended_action": a.recommended_action,
                "created_at": a.created_at.strftime("%Y-%m-%d %H:%M:%S")
            }
            for a in alerts
        ]
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


# ---------------------------------------------------------------------------
# TRAINEE ENDPOINTS
# ---------------------------------------------------------------------------

def _format_trainee(t: Trainee) -> dict:
    """Serialise a Trainee ORM object to a plain dict (shared by list & detail)."""
    lifecycle_number = 1
    if t.extra_data and isinstance(t.extra_data, dict):
        lifecycle_number = len(t.extra_data.get("lifecycles", [])) + 1
        
    current_eligibility = "Eligible"
    if t.status == "BLOCKED":
        current_eligibility = "Blocked"
    elif t.status == "SEPARATED":
        current_eligibility = "Separated"

    return {
        "id": t.id,
        "name": t.name,
        "doj": t.doj.strftime("%Y-%m-%d") if t.doj else "",
        "dol": t.dol.strftime("%Y-%m-%d") if t.dol else None,
        "scheme": t.scheme,
        "status": t.status,
        "blocked_reason": t.blocked_reason,
        "aadhaar": t.aadhaar,
        "ticket_number": t.ticket_number,
        "category": t.category,
        "batch": t.batch,
        "shop": t.shop,
        "lifecycle_number": lifecycle_number,
        "current_eligibility": current_eligibility,
        "current_workbook": t.current_workbook,
        "current_sheet": t.current_sheet,
    }


@router.get("/trainees")
def get_trainees(
    search: Optional[str] = None,
    status: Optional[str] = None,
    scheme: Optional[str] = None,
    joining_this_month: Optional[bool] = None,
    separations_this_month: Optional[bool] = None,
    early_separations: Optional[bool] = None,
    db: Session = Depends(get_db),
):
    try:
        query = db.query(Trainee)
        if search:
            search_filter = f"%{search}%"
            query = query.filter(
                or_(Trainee.id.like(search_filter), Trainee.name.like(search_filter))
            )
        if status:
            query = query.filter(Trainee.status == status)
        if scheme:
            query = query.filter(Trainee.scheme == scheme)
            
        today = datetime.date.today()
        start_of_month = datetime.date(today.year, today.month, 1)
        if joining_this_month:
            query = query.filter(Trainee.doj >= start_of_month, Trainee.doj <= today)
        if separations_this_month:
            query = query.filter(Trainee.dol >= start_of_month, Trainee.dol <= today)
        if early_separations:
            query = query.filter(
                Trainee.dol.isnot(None),
                func.julianday(Trainee.dol) - func.julianday(Trainee.doj) < 30
            )
            
        trainees = query.order_by(Trainee.id).all()
        return [_format_trainee(t) for t in trainees]
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/trainees/{trainee_id}")
def get_trainee_details(trainee_id: str, db: Session = Depends(get_db)):
    trainee = TraineeRepository.get_by_id(db, trainee_id)
    if not trainee:
        raise HTTPException(status_code=404, detail=f"Trainee '{trainee_id}' not found.")

    ledger = LedgerRepository.get_by_trainee_id(db, trainee_id)
    history_list = [
        {
            "id": h.id,
            "invoice_number": h.invoice_number,
            "payment_type": h.payment_type,
            "amount_paid": h.amount_paid,
            "payment_date": h.payment_date.strftime("%Y-%m-%d"),
        }
        for h in ledger
    ]

    # Calculate payment summary
    joining_paid = sum(h.amount_paid for h in ledger if h.payment_type == "JOINING")
    days180_paid = sum(h.amount_paid for h in ledger if h.payment_type == "180_DAYS")
    total_paid = joining_paid + days180_paid
    payment_summary = {
        "joining_paid": joining_paid,
        "joining_limit": 1200.0,
        "joining_remaining": max(0.0, 1200.0 - joining_paid),
        "days180_paid": days180_paid,
        "days180_limit": 600.0,
        "days180_remaining": max(0.0, 600.0 - days180_paid),
        "total_paid": total_paid,
        "total_limit": 1800.0,
        "total_remaining": max(0.0, 1800.0 - total_paid),
    }

    # Format separation history
    separation_history = []
    for s in trainee.separation_records:
        ext = s.extra_data or {}
        separation_history.append({
            "id": s.id,
            "workbook": s.file_name,
            "sheet": ext.get("sheet", ""),
            "month": ext.get("month", ""),
            "dol": s.dol.strftime("%Y-%m-%d") if s.dol else "",
            "reason": s.reason or "",
            "status_before": ext.get("status_before", ""),
            "status_after": ext.get("status_after", ""),
            "tenure": ext.get("tenure", 0),
        })
    separation_history.sort(key=lambda x: x["dol"])

    # Route through ValidationRepository — no direct db.query in endpoints
    violations = ValidationRepository.get_by_trainee_id(db, trainee_id)
    violations_list = [
        {
            "id": v.id,
            "rule_name": v.rule_name,
            "status": v.status,
            "message": v.message,
            "created_at": v.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        }
        for v in violations
    ]

    # Format invoice timeline
    from backend.app.models.models import InvoiceItem
    timeline_items = db.query(InvoiceItem).filter(InvoiceItem.trainee_id == trainee_id).all()
    timeline_items.sort(key=lambda x: x.invoice_date or datetime.date.min)
    invoice_timeline = [
        {
            "id": item.id,
            "invoice_number": item.invoice_number,
            "invoice_date": item.invoice_date.strftime("%Y-%m-%d") if item.invoice_date else "",
            "claimed_amount": item.claimed_amount,
            "approved_amount": item.approved_amount,
            "rejected_amount": item.rejected_amount,
            "status": item._status,
            "reason": item.reason,
            "validation_summary": item.validation_summary,
            "fraud_score": item.fraud_score,
            "fraud_category": item.fraud_category,
        }
        for item in timeline_items
    ]

    return {
        "profile": _format_trainee(trainee),
        "payment_history": history_list,
        "payment_summary": payment_summary,
        "separation_history": separation_history,
        "violations": violations_list,
        "invoice_timeline": invoice_timeline,
    }


@router.post("/trainees/{trainee_id}/block")
def block_trainee(
    trainee_id: str,
    payload: BlockTraineeRequest,
    db: Session = Depends(get_db),
):
    try:
        trainee = TraineeRepository.block_trainee(db, trainee_id, payload.reason)
        if not trainee:
            raise HTTPException(status_code=404, detail="Trainee not found.")
        AuditLogRepository.add_log(
            db=db,
            action="BLOCK_TRAINEE",
            module="EMPLOYEE_MASTER",
            details=f"Manually blocked trainee '{trainee_id}'. Reason: {payload.reason}",
            operator="Admin",
            employee_id=trainee_id
        )
        return {"message": f"Trainee '{trainee_id}' has been blocked."}
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/trainees/{trainee_id}/unblock")
def unblock_trainee(trainee_id: str, db: Session = Depends(get_db)):
    try:
        trainee = TraineeRepository.unblock_trainee(db, trainee_id)
        if not trainee:
            raise HTTPException(status_code=404, detail="Trainee not found.")
        AuditLogRepository.add_log(
            db=db,
            action="UNBLOCK_TRAINEE",
            module="EMPLOYEE_MASTER",
            details=f"Manually unblocked trainee '{trainee_id}'",
            operator="Admin",
            employee_id=trainee_id
        )
        return {"message": f"Trainee '{trainee_id}' has been unblocked."}
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(exc))


# ---------------------------------------------------------------------------
# EXCEL UPLOAD ENDPOINTS
# ---------------------------------------------------------------------------

@router.post("/uploads/bdc")
async def upload_bdc(
    file: UploadFile = File(...),
    upload_mode: str = Form("INCREMENTAL"),
    admin: str = Depends(get_offline_user),
    db: Session = Depends(get_db)
):
    content, filename = await _read_upload_safely(file)
    try:
        return ImportService.import_bdc_workbook(db, content, filename, upload_mode=upload_mode, operator=admin)
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/uploads/separation")
async def upload_separation(
    file: UploadFile = File(...),
    admin: str = Depends(get_offline_user),
    db: Session = Depends(get_db)
):
    content, filename = await _read_upload_safely(file)
    try:
        return ImportService.import_separation_workbook(db, content, filename, operator=admin)
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/uploads/invoice")
async def upload_invoice(
    file: UploadFile = File(...),
    invoice_number: Optional[str] = Form(None),
    invoice_date: Optional[str] = Form(None),
    admin: str = Depends(get_offline_user),
    db: Session = Depends(get_db),
):
    content, filename = await _read_upload_safely(file, allowed_extensions={".xlsx", ".xls", ".pdf"})

    inv_date: Optional[datetime.date] = None
    if invoice_date:
        try:
            inv_date = datetime.datetime.strptime(invoice_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="invoice_date must be in YYYY-MM-DD format."
            )

    try:
        return ImportService.import_invoice_workbook(
            db=db,
            file_content=content,
            file_name=filename,
            invoice_number_override=invoice_number,
            invoice_date_override=inv_date,
            operator=admin
        )
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/uploads/analyze")
async def analyze_workbook(file: UploadFile = File(...)):
    """Runs a dry-run analysis on an uploaded Excel sheet and returns schema stats/types."""
    content, filename = await _read_upload_safely(file)
    try:
        parsed = WorkbookParser.parse_workbook(content, filename)
        # Convert sheet rows to exclude private keys (like _row_num) for clean output
        for s in parsed["sheets"]:
            s["rows"] = [{k: v for k, v in r.items() if not k.startswith('_')} for r in s["rows"]]
        return parsed
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Analysis failed: {str(exc)}")


# ---------------------------------------------------------------------------
# INVOICE & VALIDATION ENDPOINTS
# ---------------------------------------------------------------------------

@router.get("/invoices")
def get_invoices(status: Optional[str] = None, db: Session = Depends(get_db)):
    try:
        invoices = InvoiceRepository.get_unique_invoices(db)
        if status:
            invoices = [inv for inv in invoices if inv["status"] == status]
        return invoices
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/invoices/{invoice_number}")
def get_invoice_records(invoice_number: str, db: Session = Depends(get_db)):
    records = InvoiceRepository.get_by_invoice_number(db, invoice_number)
    if not records:
        raise HTTPException(status_code=404, detail="Invoice not found.")

    # Batch-load all validation flags for this invoice in ONE query (avoids N+1)
    all_flags = ValidationRepository.get_by_invoice_number(db, invoice_number)
    flags_by_record: dict = {}
    for f in all_flags:
        flags_by_record.setdefault(f.invoice_record_id, []).append(f.status)

    result_list = []
    for r in records:
        flag_statuses = flags_by_record.get(r.id, [])
        result_list.append({
            "id": r.id,
            "trainee_id": r.trainee_id,
            "billed_name": r.billed_name,
            "billed_joining": r.billed_joining_amount,
            "billed_180_days": r.billed_180_days_amount,
            "billed_other": r.billed_other_amount,
            "billed_total": r.billed_total_amount,
            "approved_joining": r.approved_joining_amount,
            "approved_180_days": r.approved_180_days_amount,
            "approved_total": r.approved_total_amount,
            "status": r.status,
            "severity": _severity_from_flags(flag_statuses),
            "flags_count": len(flag_statuses),
        })
    return result_list


@router.get("/invoices/{invoice_number}/exceptions")
def get_invoice_exceptions(invoice_number: str, db: Session = Depends(get_db)):
    exceptions = ValidationRepository.get_by_invoice_number(db, invoice_number)
    return [
        {
            "id": e.id,
            "invoice_record_id": e.invoice_record_id,
            "trainee_id": e.trainee_id,
            "rule_name": e.rule_name,
            "status": e.status,
            "message": e.message,
            "reason_code": e.reason_code,
            "recommended_action": e.recommended_action,
        }
        for e in exceptions
    ]


@router.get("/invoices/{invoice_number}/validation-details/{record_id}")
def get_validation_details(
    invoice_number: str, record_id: int, db: Session = Depends(get_db)
):
    # Verify the record belongs to the given invoice to prevent cross-invoice leaks
    record = InvoiceRepository.get_by_id(db, record_id)
    if not record or record.invoice_number != invoice_number:
        raise HTTPException(
            status_code=404,
            detail=f"Record {record_id} not found under invoice '{invoice_number}'."
        )

    flags = ValidationRepository.get_by_invoice_record(db, record_id)
    return [
        {
            "id": f.id,
            "rule_name": f.rule_name,
            "status": f.status,
            "message": f.message,
            "reason_code": f.reason_code,
            "recommended_action": f.recommended_action,
        }
        for f in flags
    ]


@router.post("/invoices/{invoice_number}/validate")
def validate_invoice(invoice_number: str, db: Session = Depends(get_db)):
    try:
        return ValidationService.validate_invoice(db, invoice_number)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/invoices/{invoice_number}/approve")
def approve_invoice(invoice_number: str, db: Session = Depends(get_db)):
    try:
        success = LedgerService.approve_invoice_and_post_to_ledger(db, invoice_number)
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(exc))
    if not success:
        raise HTTPException(status_code=404, detail=f"Invoice '{invoice_number}' not found.")
    return {"message": f"Invoice '{invoice_number}' approved and payouts posted to ledger."}


@router.post("/invoices/{invoice_number}/reject")
def reject_invoice(invoice_number: str, db: Session = Depends(get_db)):
    try:
        success = LedgerService.reject_invoice(db, invoice_number)
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(exc))
    if not success:
        raise HTTPException(status_code=404, detail=f"Invoice '{invoice_number}' not found.")
    return {"message": f"Invoice '{invoice_number}' marked as rejected."}


@router.delete("/invoices/{invoice_number}")
def delete_invoice(invoice_number: str, db: Session = Depends(get_db)):
    """
    Atomically removes ledger entries, validation results, and invoice records
    for the given invoice number in a single transaction.
    """
    try:
        # Check existence first so we can return 404 before touching data
        records = InvoiceRepository.get_by_invoice_number(db, invoice_number)
        if not records:
            raise HTTPException(status_code=404, detail=f"Invoice '{invoice_number}' not found.")

        # Delete ledger then invoice records — both deferred to a single commit
        LedgerRepository.delete_by_invoice_number(db, invoice_number, commit=False)
        InvoiceRepository.delete_by_invoice_number(db, invoice_number, commit=False)
        db.commit()

        workbook_name = records[0].file_name if records else None
        AuditLogRepository.add_log(
            db=db,
            action="DELETE_INVOICE",
            module="INVOICE_UPLOAD",
            details=f"Deleted invoice '{invoice_number}' and associated records.",
            operator="Admin",
            workbook=workbook_name,
            rows_count=len(records),
            invoice_number=invoice_number
        )
        return {"message": f"Invoice '{invoice_number}' has been permanently deleted."}
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/invoices/compare/{invoice_a}/{invoice_b}")
def compare_invoices(invoice_a: str, invoice_b: str, db: Session = Depends(get_db)):
    from backend.app.models.models import InvoiceItem
    
    items_a = db.query(InvoiceItem).filter(InvoiceItem.invoice_number == invoice_a).all()
    items_b = db.query(InvoiceItem).filter(InvoiceItem.invoice_number == invoice_b).all()
    
    if not items_a:
        raise HTTPException(status_code=404, detail=f"Invoice '{invoice_a}' not found or has no records.")
    if not items_b:
        raise HTTPException(status_code=404, detail=f"Invoice '{invoice_b}' not found or has no records.")
        
    map_a = {item.trainee_id or item.ticket_number: item for item in items_a}
    map_b = {item.trainee_id or item.ticket_number: item for item in items_b}
    
    added = []
    removed = []
    changed = []
    common_count = 0
    
    for key, item_b in map_b.items():
        if key not in map_a:
            added.append({
                "trainee_id": item_b.trainee_id,
                "ticket_number": item_b.ticket_number,
                "name": item_b.candidate_name,
                "billed_joining": item_b.billed_joining_amount,
                "billed_180_days": item_b.billed_180_days_amount,
                "billed_total": item_b.claimed_amount
            })
        else:
            common_count += 1
            item_a = map_a[key]
            if abs(item_b.claimed_amount - item_a.claimed_amount) > 0.01:
                changed.append({
                    "trainee_id": item_b.trainee_id,
                    "ticket_number": item_b.ticket_number,
                    "name": item_b.candidate_name,
                    "a_billed": item_a.claimed_amount,
                    "b_billed": item_b.claimed_amount,
                    "diff": item_b.claimed_amount - item_a.claimed_amount
                })
                
    for key, item_a in map_a.items():
        if key not in map_b:
            removed.append({
                "trainee_id": item_a.trainee_id,
                "ticket_number": item_a.ticket_number,
                "name": item_a.candidate_name,
                "billed_joining": item_a.billed_joining_amount,
                "billed_180_days": item_a.billed_180_days_amount,
                "billed_total": item_a.claimed_amount
            })
            
    return {
        "invoice_a_number": invoice_a,
        "invoice_b_number": invoice_b,
        "added_employees": added,
        "removed_employees": removed,
        "changed_employees": changed,
        "common_count": common_count,
        "invoice_a_total": sum(item.claimed_amount for item in items_a),
        "invoice_b_total": sum(item.claimed_amount for item in items_b)
    }


@router.get("/invoices/{invoice_number}/reconciliation")
def get_invoice_reconciliation(invoice_number: str, db: Session = Depends(get_db)):
    from backend.app.models.models import InvoiceItem, PaymentLedger, Invoice
    
    items = db.query(InvoiceItem).filter(InvoiceItem.invoice_number == invoice_number).all()
    if not items:
        raise HTTPException(status_code=404, detail=f"Invoice '{invoice_number}' not found.")
        
    trainee_ids = [item.trainee_id for item in items if item.trainee_id]
    
    # Batch query ledger and other invoices
    ledger_entries = db.query(PaymentLedger).filter(PaymentLedger.trainee_id.in_(trainee_ids)).all() if trainee_ids else []
    ledger_map = {}
    for entry in ledger_entries:
        ledger_map.setdefault(entry.trainee_id, []).append(entry)
        
    other_items = db.query(InvoiceItem).join(Invoice).filter(
        InvoiceItem.trainee_id.in_(trainee_ids),
        InvoiceItem.invoice_number != invoice_number,
        Invoice.status == "ACTIVE"
    ).all() if trainee_ids else []
    
    other_claims_map = {}
    for item in other_items:
        other_claims_map.setdefault(item.trainee_id, []).append(item)
        
    reconciliation_details = []
    
    for item in items:
        t_id = item.trainee_id
        t_ledger = ledger_map.get(t_id, []) if t_id else []
        t_other = other_claims_map.get(t_id, []) if t_id else []
        
        already_paid_joining = sum(h.amount_paid for h in t_ledger if h.payment_type == "JOINING")
        already_paid_180 = sum(h.amount_paid for h in t_ledger if h.payment_type == "180_DAYS")
        
        already_claimed_joining = sum(c.billed_joining_amount for c in t_other)
        already_claimed_180 = sum(c.billed_180_days_amount for c in t_other)
        
        # Reconciliation status derivation
        rec_status = "CLEARED"
        if item._status == "FRAUD":
            rec_status = "FRAUD_SUSPECT"
        elif (item.billed_joining_amount > 0 and (already_paid_joining > 0 or already_claimed_joining > 0)) or \
             (item.billed_180_days_amount > 0 and (already_paid_180 > 0 or already_claimed_180 > 0)):
            rec_status = "DOUBLE_BILLED"
        elif item._status == "REJECTED":
            rec_status = "REJECTED"
        elif item._status == "PARTIALLY_APPROVED":
            rec_status = "PARTIALLY_APPROVED"
        elif item.approved_amount < item.claimed_amount:
            rec_status = "OUTSTANDING"
            
        reconciliation_details.append({
            "id": item.id,
            "trainee_id": item.trainee_id,
            "name": item.candidate_name,
            "billed_total": item.claimed_amount,
            "approved_total": item.approved_amount,
            "already_paid_joining": already_paid_joining,
            "already_paid_180": already_paid_180,
            "already_claimed_joining": already_claimed_joining,
            "already_claimed_180": already_claimed_180,
            "reconciliation_status": rec_status,
            "status": item._status,
            "fraud_score": item.fraud_score,
            "fraud_category": item.fraud_category
        })
        
    return {
        "invoice_number": invoice_number,
        "total_records": len(items),
        "details": reconciliation_details
    }


# ---------------------------------------------------------------------------
# PAYMENT LEDGER
# ---------------------------------------------------------------------------

@router.get("/ledger")
def get_ledger(db: Session = Depends(get_db)):
    try:
        entries = LedgerRepository.get_all_entries(db)
        return [
            {
                "id": e.id,
                "trainee_id": e.trainee_id,
                "trainee_name": e.trainee.name if e.trainee else "N/A",
                "invoice_number": e.invoice_number,
                "payment_type": e.payment_type,
                "amount_paid": e.amount_paid,
                "payment_date": e.payment_date.strftime("%Y-%m-%d"),
            }
            for e in entries
        ]
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


# ---------------------------------------------------------------------------
# EXCEL REPORT DOWNLOADS
# ---------------------------------------------------------------------------

def _excel_streaming_response(file_bytes: bytes, filename: str) -> StreamingResponse:
    safe_filename = _sanitize_filename(filename)
    return StreamingResponse(
        BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe_filename}"},
    )


@router.get("/reports/invoice-history")
def download_invoice_history(admin: str = Depends(get_offline_user), db: Session = Depends(get_db)):
    try:
        file_bytes = ReportService.generate_invoice_history_report(db)
        return _excel_streaming_response(file_bytes, "invoice_history_report.xlsx")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/monthly-billing-summary")
def download_monthly_billing_summary(admin: str = Depends(get_offline_user), db: Session = Depends(get_db)):
    try:
        file_bytes = ReportService.generate_monthly_billing_summary_report(db)
        return _excel_streaming_response(file_bytes, "monthly_billing_summary_report.xlsx")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/duplicate-billing")
def download_duplicate_billing(admin: str = Depends(get_offline_user), db: Session = Depends(get_db)):
    try:
        file_bytes = ReportService.generate_duplicate_billing_report(db)
        return _excel_streaming_response(file_bytes, "duplicate_billing_report.xlsx")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/repeated-payments")
def download_repeated_payments(admin: str = Depends(get_offline_user), db: Session = Depends(get_db)):
    try:
        file_bytes = ReportService.generate_repeated_payment_report(db)
        return _excel_streaming_response(file_bytes, "repeated_payments_report.xlsx")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/outstanding-payments")
def download_outstanding_payments(admin: str = Depends(get_offline_user), db: Session = Depends(get_db)):
    try:
        file_bytes = ReportService.generate_outstanding_payment_report(db)
        return _excel_streaming_response(file_bytes, "outstanding_payments_report.xlsx")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/invoice-comparison/{invoice_a}/{invoice_b}")
def download_invoice_comparison(invoice_a: str, invoice_b: str, admin: str = Depends(get_offline_user), db: Session = Depends(get_db)):
    try:
        file_bytes = ReportService.generate_invoice_comparison_report(db, invoice_a, invoice_b)
        return _excel_streaming_response(file_bytes, f"comparison_{invoice_a}_vs_{invoice_b}.xlsx")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/vendor-summary")
def download_vendor_summary(admin: str = Depends(get_offline_user), db: Session = Depends(get_db)):
    try:
        file_bytes = ReportService.generate_vendor_summary_report(db)
        return _excel_streaming_response(file_bytes, "vendor_summary_report.xlsx")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/employee-billing-history")
def download_employee_billing_history(admin: str = Depends(get_offline_user), db: Session = Depends(get_db)):
    try:
        file_bytes = ReportService.generate_employee_billing_history_report(db)
        return _excel_streaming_response(file_bytes, "employee_billing_history_report.xlsx")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/approved-invoice/{invoice_number}")
def download_approved_invoice(invoice_number: str, admin: str = Depends(get_offline_user), db: Session = Depends(get_db)):
    try:
        file_bytes = ReportService.generate_approved_invoice_excel(db, invoice_number)
        return _excel_streaming_response(file_bytes, f"approved_invoice_{invoice_number}.xlsx")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/exceptions")
def download_exception_report(
    invoice_number: Optional[str] = None,
    format_version: str = "legacy",
    admin: str = Depends(get_offline_user),
    db: Session = Depends(get_db),
):
    try:
        file_bytes = ReportService.generate_exception_report_excel(db, invoice_number, format_version=format_version)
        filename = f"exceptions_{invoice_number}.xlsx" if invoice_number else "exceptions_report_all.xlsx"
        return _excel_streaming_response(file_bytes, filename)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/fraud")
def download_fraud_report(
    invoice_number: Optional[str] = None,
    admin: str = Depends(get_offline_user),
    db: Session = Depends(get_db)
):
    try:
        file_bytes = ReportService.generate_fraud_report_excel(db, invoice_number=invoice_number)
        filename = f"fraud_report_{invoice_number}.xlsx" if invoice_number else "fraud_incidents_report.xlsx"
        return _excel_streaming_response(file_bytes, filename)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/payment-summary")
def download_payment_summary(
    invoice_number: Optional[str] = None,
    format_version: str = "legacy",
    admin: str = Depends(get_offline_user),
    db: Session = Depends(get_db)
):
    try:
        file_bytes = ReportService.generate_payment_summary_excel(db, invoice_number=invoice_number, format_version=format_version)
        filename = f"payment_summary_{invoice_number}.xlsx" if invoice_number else "payment_summary_ledger.xlsx"
        return _excel_streaming_response(file_bytes, filename)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/rejected-invoice/{invoice_number}")
def download_rejected_invoice(
    invoice_number: str,
    admin: str = Depends(get_offline_user),
    db: Session = Depends(get_db)
):
    try:
        file_bytes = ReportService.generate_rejected_invoice_excel(db, invoice_number)
        return _excel_streaming_response(file_bytes, f"rejected_invoice_{invoice_number}.xlsx")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/finance-summary")
def download_finance_summary(
    invoice_number: Optional[str] = None,
    admin: str = Depends(get_offline_user),
    db: Session = Depends(get_db)
):
    try:
        file_bytes = ReportService.generate_finance_summary_excel(db, invoice_number=invoice_number)
        filename = f"finance_summary_{invoice_number}.xlsx" if invoice_number else "finance_summary_all.xlsx"
        return _excel_streaming_response(file_bytes, filename)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/corrected-invoice/{invoice_number}")
def download_corrected_invoice(invoice_number: str, admin: str = Depends(get_offline_user), db: Session = Depends(get_db)):
    try:
        file_bytes = ReportService.generate_corrected_invoice_excel(db, invoice_number)
        return _excel_streaming_response(file_bytes, f"corrected_invoice_{invoice_number}.xlsx")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/vendor-payment-summary/{invoice_number}")
def download_vendor_payment_summary(invoice_number: str, admin: str = Depends(get_offline_user), db: Session = Depends(get_db)):
    try:
        file_bytes = ReportService.generate_vendor_payment_summary_excel(db, invoice_number)
        return _excel_streaming_response(file_bytes, f"vendor_payment_summary_{invoice_number}.xlsx")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/audit-logs")
def download_audit_logs(admin: str = Depends(get_offline_user), db: Session = Depends(get_db)):
    try:
        file_bytes = ReportService.generate_audit_report_excel(db)
        return _excel_streaming_response(file_bytes, "system_audit_logs.xlsx")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/finance-analytics")
def get_finance_analytics(admin: str = Depends(get_offline_user), db: Session = Depends(get_db)):
    """Retrieve all consolidated financial metrics and savings in JSON format."""
    try:
        return ReportService.get_finance_analytics_data(db)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/finance-analytics/export")
def export_finance_analytics(
    format: str = Query(default="excel", pattern="^(excel|csv|pdf)$"),
    metric: str = Query(default="summary"),
    admin: str = Depends(get_offline_user),
    db: Session = Depends(get_db)
):
    """Export the consolidated finance analytics report in Excel, PDF, or CSV format."""
    from fastapi.responses import Response, StreamingResponse
    try:
        if format == "excel":
            file_bytes = ReportService.generate_finance_analytics_excel(db)
            safe_filename = _sanitize_filename("finance_analytics_report.xlsx")
            return StreamingResponse(
                BytesIO(file_bytes),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={safe_filename}"},
            )
        elif format == "csv":
            file_bytes = ReportService.generate_finance_analytics_csv(db, metric=metric)
            safe_filename = _sanitize_filename(f"finance_analytics_{metric}.csv")
            return Response(
                content=file_bytes,
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f"attachment; filename={safe_filename}"},
            )
        elif format == "pdf":
            file_bytes = ReportService.generate_finance_analytics_pdf(db)
            safe_filename = _sanitize_filename("finance_analytics_report.pdf")
            return Response(
                content=file_bytes,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename={safe_filename}"},
            )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/reports/validation-errors")
def download_validation_errors(
    invoice_number: Optional[str] = None,
    admin: str = Depends(get_offline_user),
    db: Session = Depends(get_db)
):
    """Generate downloadable Excel report of row-level errors for Trainees and Invoices."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    query = db.query(ValidationResult)
    if invoice_number:
        query = query.join(InvoiceRecord).filter(InvoiceRecord.invoice_number == invoice_number)
    results = query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation Errors"
    ws.views.sheetView[0].showGridLines = True
    
    headers = ["Sheet", "Row", "Ticket", "Employee", "Rule", "Reason", "Suggested Fix", "Severity"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[1].height = 25
    
    for res in results:
        sheet_val = "N/A"
        row_val = "N/A"
        ticket_val = "N/A"
        emp_name = "N/A"
        
        trainee = None
        if res.trainee_id:
            trainee = db.query(Trainee).filter(Trainee.id == res.trainee_id).first()
        elif res.invoice_record_id:
            inv_rec = db.query(InvoiceRecord).filter(InvoiceRecord.id == res.invoice_record_id).first()
            if inv_rec and inv_rec.trainee_id:
                trainee = db.query(Trainee).filter(Trainee.id == inv_rec.trainee_id).first()
                sheet_val = inv_rec.file_name or "N/A"
                if inv_rec.extra_data:
                    row_val = inv_rec.extra_data.get("_row_num") or "N/A"
        
        if trainee:
            ticket_val = trainee.ticket_number or "N/A"
            emp_name = trainee.name or "N/A"
            if sheet_val == "N/A":
                sheet_val = trainee.current_sheet or "N/A"
            if row_val == "N/A" and trainee.extra_data:
                row_val = trainee.extra_data.get("_row_num") or "N/A"
                
        ws.append([
            sheet_val,
            row_val,
            ticket_val,
            emp_name,
            res.rule_name,
            res.message,
            res.recommended_action or "N/A",
            res.status
        ])
        
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    return _excel_streaming_response(out.getvalue(), "validation_errors_report.xlsx")


@router.get("/uploads/history")
def get_upload_history(
    upload_type: Optional[str] = None,
    query: Optional[str] = None,
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    admin: str = Depends(get_offline_user),
    db: Session = Depends(get_db)
):
    """Retrieve permanent upload history logs for the dashboard."""
    q = db.query(UploadHistory)
    if upload_type:
        q = q.filter(UploadHistory.upload_type == upload_type)
    if query:
        search_filter = f"%{query}%"
        q = q.filter(
            or_(
                UploadHistory.file_name.like(search_filter),
                UploadHistory.uploaded_by.like(search_filter),
                UploadHistory.remarks.like(search_filter)
            )
        )
    total = q.count()
    uploads = q.order_by(UploadHistory.upload_time.desc()).offset(offset).limit(limit).all()
    
    return {
        "total": total,
        "uploads": [
            {
                "upload_id": u.upload_id,
                "file_name": u.file_name,
                "file_hash": u.file_hash,
                "file_size": u.file_size,
                "upload_type": u.upload_type,
                "uploaded_by": u.uploaded_by,
                "upload_time": u.upload_time.strftime("%Y-%m-%d %H:%M:%S"),
                "processing_time": u.processing_time,
                "status": u.status,
                "is_duplicate": u.is_duplicate,
                "workbook_version": u.workbook_version,
                "parser_version": u.parser_version,
                "application_version": u.application_version,
                "sheet_count": u.sheet_count,
                "visible_sheet_count": u.visible_sheet_count,
                "hidden_sheet_count": u.hidden_sheet_count,
                "rows_processed": u.rows_processed,
                "rows_inserted": u.rows_inserted,
                "rows_updated": u.rows_updated,
                "rows_no_change": u.rows_no_change,
                "rows_skipped": u.rows_skipped,
                "rows_failed": u.rows_failed,
                "rows_rehired": u.rows_rehired,
                "rows_inactive": u.rows_inactive,
                "employee_sheets": u.employee_sheets,
                "separation_sheets": u.separation_sheets,
                "invoice_sheets": u.invoice_sheets,
                "remarks": u.remarks
            }
            for u in uploads
        ]
    }



# ---------------------------------------------------------------------------
# AUDIT LOGS
# ---------------------------------------------------------------------------

@router.get("/logs")
def get_logs(
    limit: int = Query(default=100, ge=1, le=_MAX_LOG_LIMIT),
    offset: int = Query(default=0, ge=0),
    module: Optional[str] = None,
    employee_id: Optional[str] = None,
    invoice_number: Optional[str] = None,
    workbook: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    action: Optional[str] = None,
    db: Session = Depends(get_db),
):
    try:
        logs = AuditLogRepository.get_logs(
            db=db,
            limit=limit,
            offset=offset,
            module=module,
            employee_id=employee_id,
            invoice_number=invoice_number,
            workbook=workbook,
            date_from=date_from,
            date_to=date_to,
            action=action
        )
        return [
            {
                "id": log.id,
                "timestamp": log.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                "action": log.action,
                "module": log.module,
                "details": log.details,
                "operator": log.operator,
                "workbook": log.workbook,
                "sheet": log.sheet,
                "rows_count": log.rows_count,
                "duration": log.duration,
                "inserted": log.inserted,
                "updated": log.updated,
                "failed": log.failed,
                "warnings": log.warnings,
                "errors": log.errors,
                "before_state": log.before_state,
                "after_state": log.after_state,
                "employee_id": log.employee_id,
                "invoice_number": log.invoice_number,
            }
            for log in logs
        ]
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


# ---------------------------------------------------------------------------
# SETTINGS  (stored as a local JSON file — no DB dependency)
# ---------------------------------------------------------------------------

@router.get("/settings")
def get_settings(admin: str = Depends(get_offline_user)):
    """Return persisted policy thresholds or system defaults."""
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except (json.JSONDecodeError, OSError):
            pass  # Fall through to defaults on corrupt file
    return SettingsPayload().model_dump()


@router.post("/settings")
def save_settings(payload: SettingsPayload, admin: str = Depends(get_offline_user)):
    """Persist validated policy thresholds to a local JSON file."""
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as fh:
            json.dump(payload.model_dump(), fh, indent=2)
        return {"message": "Settings updated successfully.", "settings": payload.model_dump()}
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"Could not write settings file: {exc}")


# ---------------------------------------------------------------------------
# RECONCILIATION ENGINE
# ---------------------------------------------------------------------------

def _build_reconciliation_summary(db: Session, invoice_number: str) -> Dict[str, Any]:
    """Compute reconciliation KPIs for a single invoice from the database."""
    records = db.query(
        InvoiceRecord.billed_total_amount,
        InvoiceRecord.approved_total_amount,
        InvoiceRecord.status,
        InvoiceRecord.invoice_date,
    ).filter(InvoiceRecord.invoice_number == invoice_number).all()

    if not records:
        raise HTTPException(status_code=404, detail=f"Invoice '{invoice_number}' not found.")

    invoice_date = records[0].invoice_date

    billed_total     = sum(r.billed_total_amount   for r in records)
    approved_total   = sum(r.approved_total_amount for r in records)
    rejected_total   = max(0.0, billed_total - approved_total)
    pending_total    = sum(r.billed_total_amount   for r in records if r.status == "PENDING")
    money_saved      = rejected_total
    approved_count   = sum(1 for r in records if r.approved_total_amount > 0.0)
    rejected_count   = sum(1 for r in records if r.approved_total_amount == 0.0 and r.status != "PENDING")
    exception_count  = sum(1 for r in records if r.status == "EXCEPTION")
    pending_count    = sum(1 for r in records if r.status == "PENDING")

    # Count validation result statuses
    vr_counts = {"FRAUD": 0, "ERROR": 0, "WARNING": 0}
    all_flags = ValidationRepository.get_by_invoice_number(db, invoice_number)
    for f in all_flags:
        if f.status in vr_counts:
            vr_counts[f.status] += 1

    # Determine overall invoice status
    statuses = [r.status for r in records]
    if all(s == "APPROVED" for s in statuses):
        overall_status = "APPROVED"
    elif "EXCEPTION" in statuses:
        overall_status = "EXCEPTION"
    elif all(s == "PENDING" for s in statuses):
        overall_status = "PENDING"
    elif "VALIDATED" in statuses:
        overall_status = "VALIDATED"
    else:
        overall_status = "MIXED"

    return {
        "invoice_number":  invoice_number,
        "invoice_date":    invoice_date.strftime("%Y-%m-%d") if invoice_date else "",
        "overall_status":  overall_status,
        "total_rows":      len(records),
        "billed_total":    round(billed_total,   2),
        "approved_total":  round(approved_total, 2),
        "rejected_total":  round(rejected_total, 2),
        "pending_total":   round(pending_total,  2),
        "money_saved":     round(money_saved,    2),
        "approved_count":  approved_count,
        "rejected_count":  rejected_count,
        "exception_count": exception_count,
        "pending_count":   pending_count,
        "fraud_count":     vr_counts["FRAUD"],
        "error_count":     vr_counts["ERROR"],
        "warning_count":   vr_counts["WARNING"],
    }


@router.get("/reconciliation/summary/{invoice_number}")
def get_reconciliation_summary(invoice_number: str, db: Session = Depends(get_db)):
    """Return a live, fully computed reconciliation summary for a single invoice."""
    try:
        return _build_reconciliation_summary(db, invoice_number)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/reconciliation/run/{invoice_number}")
def run_reconciliation(invoice_number: str, db: Session = Depends(get_db)):
    """
    Run the full reconciliation pipeline for an invoice:
      1. Validate (rule engine across Employee Master + Ledger + Separation)
      2. Auto-approve if there are no FRAUD or ERROR violations
      3. Return the reconciliation summary
    """
    try:
        # Step 1 — Run validation rules engine
        validation_result = ValidationService.validate_invoice(db, invoice_number)

        # Step 2 — Auto-approve only when there are zero FRAUD or ERROR flags
        auto_approved = False
        if validation_result["fraud_count"] == 0 and validation_result["error_count"] == 0:
            success = LedgerService.approve_invoice_and_post_to_ledger(db, invoice_number)
            auto_approved = success

        AuditLogRepository.add_log(
            db=db,
            action="RUN_RECONCILIATION",
            module="RECONCILIATION",
            details=(
                f"Invoice: {invoice_number}. "
                f"Validated {validation_result['total_records']} records. "
                f"Fraud: {validation_result['fraud_count']}, "
                f"Errors: {validation_result['error_count']}, "
                f"Warnings: {validation_result['warning_count']}. "
                f"Auto-approved: {auto_approved}."
            ),
            operator="Admin",
            rows_count=validation_result['total_records'],
            invoice_number=invoice_number
        )

        # Step 3 — Return live summary
        summary = _build_reconciliation_summary(db, invoice_number)
        summary["auto_approved"] = auto_approved
        summary["validation"] = validation_result
        return summary

    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(exc))
